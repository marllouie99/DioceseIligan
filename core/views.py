from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib import messages
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, F, Exists, OuterRef, Value, BooleanField, ExpressionWrapper, IntegerField, Sum, Avg
from django.urls import reverse
from django.utils.text import slugify
from django.core.cache import cache
from django.db import IntegrityError
from django.contrib.auth import get_user_model

User = get_user_model()
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

from .models import (
    Church,
    ChurchFollow,
    ChurchStaff,
    BookableService,
    ServiceCategory,
    Availability,
    ServiceImage,
    Booking,
    Post,
    PostLike,
    PostBookmark,
    PostComment,
    CommentLike,
    PostView,
    PostReport,
    ChurchVerificationRequest,
    ChurchVerificationDocument,
    Notification,
    DeclineReason,
    Donation,
    ServiceReview,
)
from .forms import (
    ChurchCreateForm,
    ChurchUpdateForm,
    ChurchSearchForm,
    BookableServiceForm,
    AvailabilityForm,
    AvailabilityBulkForm,
    ServiceImageForm,
    BookingForm,
    PostForm,
    ChurchVerificationUploadForm,
    SuperAdminChurchCreateForm,
)
from .utils import get_user_display_data, get_essential_profile_status, optimize_image
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import os
from django.utils import timezone
from .notifications import create_booking_notification, NotificationTemplates


# Permission Helper Functions
def log_staff_activity(user, church, action, category, description, target_id=None, target_type=None, request=None):
    """
    Helper function to log staff activities.
    
    Args:
        user: The user performing the action
        church: The church instance
        action: Action type (use StaffActivityLog.ACTION_* constants)
        category: Category type (use StaffActivityLog.CATEGORY_* constants)
        description: Human-readable description of the action
        target_id: Optional ID of the affected object
        target_type: Optional type of the affected object
        request: Optional request object to get IP address
    """
    from core.models import ChurchStaff, StaffActivityLog
    
    try:
        # Get the staff member record
        staff_member = ChurchStaff.objects.filter(
            user=user,
            church=church,
            status=ChurchStaff.STATUS_ACTIVE
        ).first()
        
        if staff_member:
            # Get IP address from request
            ip_address = None
            if request:
                ip_address = request.META.get('REMOTE_ADDR') or request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
            
            StaffActivityLog.objects.create(
                staff=staff_member,
                church=church,
                action=action,
                category=category,
                description=description,
                target_id=target_id,
                target_type=target_type,
                ip_address=ip_address
            )
    except Exception as e:
        # Silently fail - don't break the main operation if logging fails
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Failed to log staff activity: {str(e)}")


def user_can_manage_church(user, church, required_permissions=None):
    """
    Check if user has permission to manage church based on ownership or staff role.
    
    Args:
        user: The user to check
        church: The church instance
        required_permissions: List of required permissions like ['appointments', 'services', 'content', 'events']
                            If None, just checks if user has any access
    
    Returns:
        tuple: (has_permission: bool, user_role: str or None)
    """
    # Owner has full access
    if church.owner == user:
        return (True, 'owner')
    
    # Check staff positions
    staff_position = ChurchStaff.objects.filter(
        user=user,
        church=church,
        status=ChurchStaff.STATUS_ACTIVE
    ).first()
    
    if not staff_position:
        return (False, None)
    
    role = staff_position.role
    
    # If no specific permissions required, return True (has access)
    if required_permissions is None:
        return (True, role)
    
    # Define role permissions
    role_permissions = {
        ChurchStaff.ROLE_SECRETARY: ['appointments', 'services', 'availability', 'transactions', 'messaging'],
        ChurchStaff.ROLE_VOLUNTEER: ['events', 'content']
    }
    
    user_permissions = role_permissions.get(role, [])
    
    # Check if user has all required permissions
    has_permission = all(perm in user_permissions for perm in required_permissions)
    
    return (has_permission, role)


def _app_context(request):
    """Common context for app pages: user display, initial, placeholder activity counts."""
    from accounts.views import ACTIVITY_COUNTS
    
    user = request.user
    try:
        profile = user.profile
    except Exception:
        profile = None
    
    user_display_name, user_initial = get_user_display_data(user, profile)
    # Compute essential profile status for global UI indicators
    try:
        essential = get_essential_profile_status(user, profile)
        essentials_incomplete = not essential.get('is_complete')
        essentials_missing = list(essential.get('missing', []))
    except Exception:
        essential = {'is_complete': False, 'missing': []}
        essentials_incomplete = True
        essentials_missing = []

    # Get recent user activities for sidebar (post interactions)
    recent_activities = []
    if user.is_authenticated:
        from core.models import UserInteraction
        recent_activities = UserInteraction.objects.filter(
            user=user
        ).select_related('content_type').prefetch_related('content_object').order_by('-created_at')[:3]

    # PayPal configuration for donation integration
    from django.conf import settings
    PAYPAL_CLIENT_ID = getattr(settings, 'PAYPAL_CLIENT_ID', '')
    PAYPAL_CURRENCY = getattr(settings, 'PAYPAL_CURRENCY', 'PHP')
    
    # Stripe configuration for credit card donations
    STRIPE_PUBLISHABLE_KEY = getattr(settings, 'STRIPE_PUBLISHABLE_KEY', '')
    
    # Get unread booking notifications count for church owners
    unread_booking_notifications = 0
    if user.is_authenticated and user.owned_churches.exists():
        from core.models import Notification, Booking
        unread_booking_notifications = Notification.objects.filter(
            user=user,
            is_read=False,
            notification_type__in=[
                Notification.TYPE_BOOKING_REQUESTED,
                Notification.TYPE_BOOKING_REVIEWED,
                Notification.TYPE_BOOKING_APPROVED,
                Notification.TYPE_BOOKING_DECLINED,
                Notification.TYPE_BOOKING_CANCELED,
                Notification.TYPE_BOOKING_COMPLETED
            ]
        ).count()
    
    return {
        'user_display_name': user_display_name,
        'user_initial': user_initial,
        'activity_counts': ACTIVITY_COUNTS,
        'is_admin_mode': bool(request.session.get('super_admin_mode', False)) if getattr(user, 'is_superuser', False) else False,
        'profile_essentials_incomplete': essentials_incomplete,
        'profile_essentials_missing': essentials_missing,
        'essential_status': essential,
        'recent_activities': recent_activities,
        'PAYPAL_CLIENT_ID': PAYPAL_CLIENT_ID,
        'PAYPAL_CURRENCY': PAYPAL_CURRENCY,
        'STRIPE_PUBLISHABLE_KEY': STRIPE_PUBLISHABLE_KEY,
        'unread_booking_notifications': unread_booking_notifications,
    }

def home(request):
    context = {"title": "ChurchIligan"}
    return render(request, "home.html", context)


@login_required
def toggle_super_admin_mode(request):
    """Toggle or set Super Admin mode for superusers via session.
    If mode is 'admin' -> enable admin mode. If 'user' -> disable. Otherwise toggle.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to use this feature.')
        return redirect('core:home')

    mode = request.GET.get('mode')
    if mode == 'admin':
        request.session['super_admin_mode'] = True
        messages.info(request, 'Switched to Super Admin mode.')
        # Always go to Super Admin dashboard when enabling admin mode
        return redirect('core:super_admin_dashboard')
    elif mode == 'user':
        request.session['super_admin_mode'] = False
        messages.info(request, 'Switched to User mode.')
        next_url = request.GET.get('next')
        return redirect(next_url) if next_url else redirect('dashboard')
    else:
        current = bool(request.session.get('super_admin_mode', False))
        request.session['super_admin_mode'] = not current
        messages.info(request, 'Switched mode.')
        if request.session.get('super_admin_mode'):
            # Toggled into admin mode -> go to admin dashboard
            return redirect('core:super_admin_dashboard')
        # Toggled into user mode -> honor next, else dashboard
        next_url = request.GET.get('next')
        return redirect(next_url) if next_url else redirect('dashboard')

@login_required
def discover(request):
    """Discover churches page with search and filtering."""
    search_form = ChurchSearchForm(request.GET)
    churches = Church.objects.filter(is_active=True).select_related('owner')
    
    # Apply search filters
    if search_form.is_valid():
        query = search_form.cleaned_data.get('query')
        denomination = search_form.cleaned_data.get('denomination')
        city = search_form.cleaned_data.get('city')
        size = search_form.cleaned_data.get('size')
        
        if query:
            churches = churches.filter(
                Q(name__icontains=query) |
                Q(description__icontains=query) |
                Q(city__icontains=query) |
                Q(denomination__icontains=query)
            )
        
        if denomination:
            churches = churches.filter(denomination=denomination)
        
        if city:
            churches = churches.filter(city__icontains=city)
        
        if size:
            churches = churches.filter(size=size)
    
    # Order by follower count and creation date
    churches = churches.annotate(
        followers_count=Count('followers')
    ).order_by('-followers_count', '-created_at')
    
    # Pagination
    paginator = Paginator(churches, 12)  # 12 churches per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get user's followed churches for context
    user_followed_churches = set()
    if request.user.is_authenticated:
        user_followed_churches = set(
            ChurchFollow.objects.filter(user=request.user)
            .values_list('church_id', flat=True)
        )
    
    # Get user address for distance calculation
    user_address_data = {}
    if request.user.is_authenticated:
        try:
            profile = request.user.profile
            user_address_data = {
                'user_street': profile.street_address or '',
                'user_barangay': profile.barangay or '',
                'user_city': profile.city_municipality or '',
                'user_province': profile.province or '',
            }
        except:
            user_address_data = {
                'user_street': '',
                'user_barangay': '',
                'user_city': '',
                'user_province': '',
            }
    
    ctx = {
        'active': 'discover',
        'page_title': 'Discover Churches',
        'search_form': search_form,
        'churches': page_obj,
        'user_followed_churches': user_followed_churches,
        'total_churches': paginator.count,
    }
    ctx.update(user_address_data)
    ctx.update(_app_context(request))
    return render(request, 'core/discover.html', ctx)


@login_required
def create_church(request):
    """Create a new church (legacy entrypoint).
    Church creation is now restricted to Super Admins who can create a church and assign a manager.
    This endpoint remains to gracefully handle existing links and redirects users appropriately.
    """
    if request.user.is_superuser:
        # Super Admins should use the new creation flow
        return redirect('core:super_admin_create_church')
    
    messages.info(
        request,
        'Church creation is restricted to Super Admin. Please contact the platform administrator to create a church and assign you as manager.'
    )
    return redirect('core:home')


@login_required
def church_detail(request, slug):
    """View church details."""
    church = get_object_or_404(Church, slug=slug, is_active=True)
    
    # Check if user is following this church
    is_following = False
    if request.user.is_authenticated:
        is_following = ChurchFollow.objects.filter(
            user=request.user, 
            church=church
        ).exists()
    
    
    # Active services for Request Appointments tab (only show if verified)
    services = []
    if church.is_verified:
        services = list(
            church.bookable_services.filter(is_active=True)
            .prefetch_related('service_images')
            .order_by('name')
        )
    
    # Posts for Posts & Updates tab (avoid N+1 for liked/bookmarked; include counts)
    post_qs = (
        church.posts
        .filter(is_active=True)
        .select_related('church')
    )
    if request.user.is_authenticated:
        post_qs = post_qs.annotate(
            is_liked=Exists(
                PostLike.objects.filter(user=request.user, post_id=OuterRef('pk'))
            ),
            is_bookmarked=Exists(
                PostBookmark.objects.filter(user=request.user, post_id=OuterRef('pk'))
            ),
        )
    else:
        post_qs = post_qs.annotate(
            is_liked=Value(False, output_field=BooleanField()),
            is_bookmarked=Value(False, output_field=BooleanField()),
        )
    post_qs = post_qs.annotate(
        likes_count=Count('likes', distinct=True),
        comments_count=Count('comments', filter=Q(comments__is_active=True), distinct=True),
    ).order_by('-created_at')[:10]
    posts = list(post_qs)
    # Note: The like_count and comment_count properties will automatically use
    # the annotated values (likes_count, comments_count) to avoid N+1 queries
    
    # Reviews for Reviews tab
    recent_reviews = church.service_reviews.filter(is_active=True).select_related(
        'user', 'service'
    ).order_by('-created_at')[:10]
    
    # Calculate overall church rating statistics (DB-level aggregation)
    all_church_reviews = church.service_reviews.filter(is_active=True)
    total_church_reviews = all_church_reviews.count()
    
    if total_church_reviews > 0:
        from django.db.models import Avg
        overall_rating = all_church_reviews.aggregate(avg=Avg('rating'))['avg'] or 0
        overall_rating = round(overall_rating, 1)
        dist_rows = all_church_reviews.values('rating').annotate(count=Count('id'))
        overall_rating_distribution = {i: 0 for i in range(1, 6)}
        for row in dist_rows:
            try:
                rating_val = int(row['rating'])
                if rating_val in overall_rating_distribution:
                    overall_rating_distribution[rating_val] = row['count']
            except Exception:
                continue
    else:
        overall_rating = 0
        overall_rating_distribution = {i: 0 for i in range(1, 6)}
    
    # Determine if the current user can request appointments
    try:
        profile = request.user.profile
    except Exception:
        profile = None
    essential = get_essential_profile_status(request.user, profile)
    can_request_appointment = bool(essential.get('is_complete'))
    essential_missing = list(essential.get('missing', []))
    
    # Get posts with images for photo gallery (limit to recent 9 photos for a 3x3 grid, can show more)
    photo_posts = church.posts.filter(
        is_active=True,
        image__isnull=False
    ).exclude(image='').order_by('-created_at')[:9]
    
    # Get event posts for Events tab
    event_posts = church.posts.filter(
        is_active=True,
        post_type='event'
    ).select_related('church').order_by('-event_start_date')[:10]
    
    # Check if user can manage content (Owner or Ministry Leader/Volunteer)
    can_manage_content, user_role = user_can_manage_church(request.user, church, ['content'])
    
    ctx = {
        'page_title': church.name,
        'church': church,
        'is_following': is_following,
        'bookable_services': services,
        'posts': posts,
        'recent_reviews': recent_reviews,
        'overall_rating': overall_rating,
        'total_church_reviews': total_church_reviews,
        'overall_rating_distribution': overall_rating_distribution,
        'can_request_appointment': can_request_appointment,
        'essential_missing': essential_missing,
        'photo_posts': photo_posts,
        'event_posts': event_posts,
        'can_manage_content': can_manage_content,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/church_detail.html', ctx)


@login_required
def select_church(request):
    """Allow user to select which church to manage if they manage multiple churches or staff positions."""
    from core.models import ChurchStaff
    
    # Get churches user owns using consistent query
    owned_churches = Church.objects.filter(owner=request.user)
    
    # Get churches where user is active staff
    staff_churches = Church.objects.filter(
        staff_members__user=request.user,
        staff_members__status=ChurchStaff.STATUS_ACTIVE
    )
    
    # Combine both querysets using distinct to avoid duplicates
    all_churches = (owned_churches | staff_churches).distinct()
    
    if not all_churches.exists():
        if request.user.is_superuser:
            return redirect('core:super_admin_create_church')
        messages.info(
            request,
            "You don't have access to manage any parishes yet. Please contact a Parish Manager."
        )
        return redirect('core:home')
    
    # If user only manages one church, redirect directly to manage_church
    if all_churches.count() == 1:
        return redirect('core:manage_church', church_id=all_churches.first().id)
    
    # Annotate churches with user role for display
    churches_with_roles = []
    for church in all_churches:
        if church in owned_churches:
            role = 'Parish Manager'
        else:
            staff_position = ChurchStaff.objects.filter(
                user=request.user,
                church=church,
                status=ChurchStaff.STATUS_ACTIVE
            ).first()
            role = staff_position.get_role_display() if staff_position else 'Staff'
        
        churches_with_roles.append({
            'church': church,
            'role': role
        })
    
    ctx = {
        'churches_with_roles': churches_with_roles,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/select_church.html', ctx)


@login_required
def manage_church(request, church_id=None):
    """Manage user's church or staff position."""
    from core.models import ChurchStaff
    
    # Determine user's role and churches they can manage
    user_role = None  # 'owner', 'secretary', 'volunteer'
    staff_position = None
    
    try:
        # If church_id is provided, get that specific church
        if church_id:
            # Try to get as owner first
            try:
                church = request.user.owned_churches.get(id=church_id)
                user_role = 'owner'
            except Church.DoesNotExist:
                # Try to get as staff member
                staff_position = ChurchStaff.objects.filter(
                    user=request.user,
                    church_id=church_id,
                    status=ChurchStaff.STATUS_ACTIVE
                ).select_related('church').first()
                
                if staff_position:
                    church = staff_position.church
                    user_role = staff_position.role
                else:
                    raise Church.DoesNotExist
        else:
            # If no church_id, redirect to select_church page
            owned_churches = list(request.user.owned_churches.all())
            staff_churches = list(Church.objects.filter(
                staff_members__user=request.user,
                staff_members__status=ChurchStaff.STATUS_ACTIVE
            ).distinct())
            
            # Combine owned and staff churches using list
            # Remove duplicates while preserving order
            seen_ids = set()
            all_manageable_churches = []
            for church in owned_churches + staff_churches:
                if church.id not in seen_ids:
                    seen_ids.add(church.id)
                    all_manageable_churches.append(church)
            
            if not all_manageable_churches:
                if request.user.is_superuser:
                    return redirect('core:super_admin_create_church')
                messages.info(
                    request,
                    "You don't have access to manage any parishes. Please contact a Parish Manager."
                )
                return redirect('core:home')
            # If only one church, use it directly
            elif len(all_manageable_churches) == 1:
                church = all_manageable_churches[0]
                # Determine role for this church
                if church in owned_churches:
                    user_role = 'owner'
                else:
                    staff_position = ChurchStaff.objects.filter(
                        user=request.user,
                        church=church,
                        status=ChurchStaff.STATUS_ACTIVE
                    ).first()
                    user_role = staff_position.role if staff_position else None
            else:
                # Multiple churches, redirect to selection page
                return redirect('core:select_church')
    except Church.DoesNotExist:
        messages.error(request, "You don't have permission to manage this parish.")
        return redirect('core:select_church')
    
    # AJAX: mark booking notifications as read when appointments tab is activated via JS
    # This avoids requiring a full page reload just to clear the badge
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('ajax_mark_read') == '1':
        # Only clear new booking request notifications for this church (owner-facing)
        Notification.objects.filter(
            user=request.user,
            is_read=False,
            notification_type=Notification.TYPE_BOOKING_REQUESTED,
            church=church
        ).update(is_read=True, read_at=timezone.now())
        return JsonResponse({'success': True})
    
    # Mark booking notifications as read when viewing appointments tab
    # Only mark notifications for THIS church's bookings
    current_tab = request.GET.get('tab', 'overview')
    if current_tab == 'appointments':
        # Clear owner-facing booking request notifications for this church
        Notification.objects.filter(
            user=request.user,
            is_read=False,
            notification_type=Notification.TYPE_BOOKING_REQUESTED,
            church=church
        ).update(is_read=True, read_at=timezone.now())
    
    if request.method == 'POST' and request.POST.get('form_type') != 'verification':
        form = ChurchUpdateForm(request.POST, request.FILES, instance=church)
        if form.is_valid():
            # Save the form directly - Django ModelForm handles partial updates correctly
            # when initialized with instance=church
            church = form.save()
            messages.success(request, 'Church information has been updated successfully!')
            return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ChurchUpdateForm(instance=church)
    
    # Verification UI form (separate submission endpoint)
    verif_form = ChurchVerificationUploadForm()

    # Calculate date ranges (needed for analytics)
    from datetime import datetime, timedelta
    today = timezone.now().date()
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)
    
    # Get church statistics
    follower_count = church.followers.count()
    recent_followers = church.followers.select_related('user')[:10]
    
    # Followers analytics for Followers tab
    # New followers this week
    followers_this_week = ChurchFollow.objects.filter(
        church=church,
        followed_at__date__gte=last_7_days
    ).count()
    
    # Followers from previous week (for comparison)
    two_weeks_ago = today - timedelta(days=14)
    followers_previous_week = ChurchFollow.objects.filter(
        church=church,
        followed_at__date__gte=two_weeks_ago,
        followed_at__date__lt=last_7_days
    ).count()
    
    # Calculate growth rate (percentage increase from previous week)
    if followers_previous_week > 0:
        growth_rate = round(((followers_this_week - followers_previous_week) / followers_previous_week) * 100, 1)
    elif followers_this_week > 0:
        growth_rate = 100  # 100% growth if we had 0 before and now have some
    else:
        growth_rate = 0
    
    # Active followers (users who have interacted with church posts in last 30 days)
    from django.db.models import Q
    active_follower_ids = set(
        list(PostLike.objects.filter(
            post__church=church,
            created_at__date__gte=last_30_days
        ).values_list('user_id', flat=True)) +
        list(PostComment.objects.filter(
            post__church=church,
            created_at__date__gte=last_30_days
        ).values_list('user_id', flat=True))
    )
    active_followers = ChurchFollow.objects.filter(
        church=church,
        user_id__in=active_follower_ids
    ).count()

    # Appointments tab data
    appt_status = request.GET.get('appt_status', 'all')
    bookings_all = (
        Booking.objects
        .filter(church=church)
        .select_related('service', 'service__category', 'user', 'user__profile')
        .order_by('-created_at')
    )
    counts = {
        'all': bookings_all.count(),
        Booking.STATUS_REQUESTED: bookings_all.filter(status=Booking.STATUS_REQUESTED).count(),
        Booking.STATUS_REVIEWED: bookings_all.filter(status=Booking.STATUS_REVIEWED).count(),
        Booking.STATUS_APPROVED: bookings_all.filter(status=Booking.STATUS_APPROVED).count(),
        Booking.STATUS_COMPLETED: bookings_all.filter(status=Booking.STATUS_COMPLETED).count(),
    }
    bookings = bookings_all
    if appt_status != 'all':
        bookings = bookings.filter(status=appt_status)

    # Conflict detection among active statuses on same church/date
    active_statuses = [Booking.STATUS_REQUESTED, Booking.STATUS_REVIEWED, Booking.STATUS_APPROVED]
    active_bookings = bookings_all.filter(status__in=active_statuses)
    
    # Detect date conflicts (multiple bookings on same date)
    active_keys = list(active_bookings.values('church_id', 'date'))
    from collections import Counter
    key_counter = Counter((k['church_id'], k['date']) for k in active_keys)
    date_conflicts = {(cid, d) for (cid, d), c in key_counter.items() if c > 1}
    
    # Detect user conflicts (same user booking multiple services on same date)
    user_date_keys = list(active_bookings.values('user_id', 'date'))
    user_date_counter = Counter((k['user_id'], k['date']) for k in user_date_keys)
    user_conflicts = {(uid, d) for (uid, d), c in user_date_counter.items() if c > 1}
    
    # Mark conflicts for template usage
    booking_list = []
    for b in bookings:
        has_date_conflict = (b.church_id, b.date) in date_conflicts
        has_user_conflict = (b.user_id, b.date) in user_conflicts
        setattr(b, 'is_conflict', has_date_conflict or has_user_conflict)
        setattr(b, 'conflict_type', 'date' if has_date_conflict else ('user' if has_user_conflict else None))
        booking_list.append(b)

    # Posts data for Content tab with counts
    from django.db.models import Count
    posts = church.posts.filter(is_active=True).annotate(
        likes_count=Count('likes', distinct=True),
        comments_count=Count('comments', filter=Q(comments__is_active=True), distinct=True),
        bookmarks_count=Count('bookmarks', distinct=True)
    ).order_by('-created_at')[:20]
    posts_count = church.posts.filter(is_active=True).count()
    
    # Add liked and bookmarked status to each post for consistency
    if request.user.is_authenticated:
        for post in posts:
            post.is_liked = post.is_liked_by(request.user)
            post.is_bookmarked = post.is_bookmarked_by(request.user)
    else:
        for post in posts:
            post.is_liked = False
            post.is_bookmarked = False
    
    # Analytics data for Content tab
    from django.db.models import Sum
    
    # Post analytics (date ranges already calculated above)
    total_posts = posts_count
    posts_last_30_days = church.posts.filter(
        is_active=True, 
        created_at__date__gte=last_30_days
    ).count()
    posts_last_7_days = church.posts.filter(
        is_active=True, 
        created_at__date__gte=last_7_days
    ).count()
    
    # Engagement analytics
    total_likes = PostLike.objects.filter(post__church=church, post__is_active=True).count()
    total_comments = PostComment.objects.filter(post__church=church, post__is_active=True).count()
    
    # Calculate total views for all church posts
    from django.db.models import Sum
    total_views = church.posts.filter(is_active=True).aggregate(
        total_views=Sum('view_count')
    )['total_views'] or 0
    
    likes_last_30_days = PostLike.objects.filter(
        post__church=church,
        post__is_active=True,
        created_at__date__gte=last_30_days
    ).count()
    
    comments_last_30_days = PostComment.objects.filter(
        post__church=church,
        post__is_active=True,
        created_at__date__gte=last_30_days
    ).count()
    
    # Calculate views from last 30 days
    views_last_30_days = PostView.objects.filter(
        post__church=church,
        post__is_active=True,
        viewed_at__date__gte=last_30_days
    ).count()
    
    # Most engaged post and Top 5 (by views + likes + comments)
    analytics_posts_qs = church.posts.filter(is_active=True).annotate(
        likes_count=Count('likes', distinct=True),
        comments_count=Count('comments', filter=Q(comments__is_active=True), distinct=True),
    ).annotate(
        total_engagement=ExpressionWrapper(
            F('view_count') + F('likes_count') + F('comments_count'),
            output_field=IntegerField()
        )
    )
    most_engaged_post = analytics_posts_qs.order_by('-total_engagement', '-created_at').first()
    top_5_posts = list(analytics_posts_qs.order_by('-total_engagement', '-created_at')[:5])
    max_engagement = top_5_posts[0].total_engagement if top_5_posts else 0
    
    # Top performing posts (by likes + comments)
    top_posts = church.posts.filter(is_active=True).annotate(
        total_engagement=Count('likes') + Count('comments')
    ).order_by('-total_engagement')[:5]
    
    # Add engagement data to top posts
    if request.user.is_authenticated:
        for post in top_posts:
            post.is_liked = post.is_liked_by(request.user)
            post.is_bookmarked = post.is_bookmarked_by(request.user)
            post.likes_count = post.likes.count()
            post.comments_count = post.comments.count()
    else:
        for post in top_posts:
            post.is_liked = False
            post.is_bookmarked = False
            post.likes_count = post.likes.count()
            post.comments_count = post.comments.count()
    
    # Growth metrics
    followers_last_30_days = ChurchFollow.objects.filter(
        church=church,
        followed_at__date__gte=last_30_days
    ).count()
    
    bookings_last_30_days = Booking.objects.filter(
        church=church,
        created_at__date__gte=last_30_days
    ).count()
    
    # Event posts data for Events tab
    event_posts = church.posts.filter(
        is_active=True,
        post_type='event'
    ).order_by('-created_at')[:20]
    event_posts_count = church.posts.filter(
        is_active=True,
        post_type='event'
    ).count()
    
    # Add liked and bookmarked status to event posts
    if request.user.is_authenticated:
        for post in event_posts:
            post.is_liked = post.is_liked_by(request.user)
            post.is_bookmarked = post.is_bookmarked_by(request.user)
    else:
        for post in event_posts:
            post.is_liked = False
            post.is_bookmarked = False

    # Donations data for Donations tab
    from decimal import Decimal
    from django.db.models import Sum
    
    # Get all donations for this church's posts
    donations_queryset = Donation.objects.filter(
        post__church=church
    ).select_related('donor', 'donor__profile', 'post').order_by('-created_at')
    
    # Recent donations (last 20)
    recent_donations = list(donations_queryset[:20])
    
    # Add follower status to each donation
    for donation in recent_donations:
        if donation.donor:
            donation.is_follower = ChurchFollow.objects.filter(
                user=donation.donor,
                church=church
            ).exists()
        else:
            donation.is_follower = False
    
    # Donation statistics 
    total_donations = donations_queryset.filter(payment_status='completed').count()
    total_amount = donations_queryset.filter(payment_status='completed').aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    # This month's donations
    current_month_start = today.replace(day=1)
    this_month_donations = donations_queryset.filter(
        payment_status='completed',
        created_at__date__gte=current_month_start
    )
    this_month_amount = this_month_donations.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    this_month_count = this_month_donations.count()
    
    # This year's donations
    current_year_start = today.replace(month=1, day=1)
    this_year_donations = donations_queryset.filter(
        payment_status='completed',
        created_at__date__gte=current_year_start
    )
    this_year_amount = this_year_donations.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    this_year_count = this_year_donations.count()
    
    # Unique donors count
    unique_donors = donations_queryset.filter(payment_status='completed').values('donor').distinct().count()
    
    # Calculate donation trends for last 6 months
    donation_monthly_data = []
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        
        # Calculate next month start
        if month_start.month == 12:
            next_month_start = month_start.replace(year=month_start.year + 1, month=1)
        else:
            next_month_start = month_start.replace(month=month_start.month + 1)
        
        month_donations = donations_queryset.filter(
            payment_status='completed',
            created_at__gte=month_start,
            created_at__lt=next_month_start
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        donation_monthly_data.append(float(month_donations))
    
    # Calculate average monthly donations
    average_monthly = sum(donation_monthly_data) / len(donation_monthly_data) if donation_monthly_data else 0
    
    # Find highest month
    highest_month = max(donation_monthly_data) if donation_monthly_data else 0
    
    # Calculate growth rate (comparing last month vs previous month)
    if len(donation_monthly_data) >= 2:
        last_month = donation_monthly_data[-1]
        previous_month = donation_monthly_data[-2]
        if previous_month > 0:
            growth_rate_donations = round(((last_month - previous_month) / previous_month) * 100, 1)
        elif last_month > 0:
            growth_rate_donations = 100
        else:
            growth_rate_donations = 0
    else:
        growth_rate_donations = 0
    
    # Top donors (non-anonymous) - get unique donors with totals
    from django.db.models import Subquery, OuterRef
    top_donors_data = donations_queryset.filter(
        payment_status='completed',
        is_anonymous=False,
        donor__isnull=False
    ).values('donor').annotate(
        total_donated=Sum('amount'),
        donation_count=Count('id')
    ).order_by('-total_donated')[:5]
    
    # Get the actual donor objects with profile data
    top_donors = []
    for donor_data in top_donors_data:
        donor = User.objects.select_related('profile').get(id=donor_data['donor'])
        donor.total_donated = donor_data['total_donated']
        donor.donation_count = donor_data['donation_count']
        top_donors.append(donor)
    
    # Services statistics
    total_services = church.bookable_services.count()
    
    # Most booked service (only if it has bookings)
    from django.db.models import Avg
    most_booked_service = church.bookable_services.annotate(
        booking_count=Count('bookings')
    ).filter(booking_count__gt=0).order_by('-booking_count').first()
    
    # Highest rated service
    highest_rated_service = church.bookable_services.annotate(
        avg_rating=Avg('reviews__rating', filter=Q(reviews__is_active=True))
    ).filter(avg_rating__isnull=False).order_by('-avg_rating').first()
    
    # Lowest rated service
    lowest_rated_service = church.bookable_services.annotate(
        avg_rating=Avg('reviews__rating', filter=Q(reviews__is_active=True))
    ).filter(avg_rating__isnull=False).order_by('avg_rating').first()
    
    # Service bookings over time (last 30 days)
    from datetime import timedelta
    service_bookings_by_day = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        count = Booking.objects.filter(
            church=church,
            created_at__date=day
        ).count()
        service_bookings_by_day.append({
            'date': day.strftime('%b %d'),
            'count': count
        })
    
    # Top 5 services by bookings
    top_services_by_bookings = church.bookable_services.annotate(
        booking_count=Count('bookings')
    ).filter(booking_count__gt=0).order_by('-booking_count')[:5]
    
    # Transactions data for Transactions tab
    # Get only bookings where online payment was actually initiated (has payment method)
    transactions_queryset = Booking.objects.filter(
        church=church,
        payment_method__isnull=False  # Only show bookings with payment method (online payment)
    ).exclude(
        payment_method=''  # Exclude empty payment methods
    ).select_related('service', 'service__category', 'user', 'user__profile').order_by('-payment_date', '-created_at')
    
    # Recent transactions (last 20)
    recent_transactions = list(transactions_queryset[:20])
    
    # Transaction statistics
    total_revenue = transactions_queryset.filter(payment_status='paid').aggregate(
        total=Sum('payment_amount')
    )['total'] or Decimal('0.00')
    
    completed_transactions = transactions_queryset.filter(payment_status='paid').count()
    pending_transactions = transactions_queryset.filter(payment_status='pending').count()
    
    # This month's transactions
    this_month_transactions = transactions_queryset.filter(
        payment_status='paid',
        payment_date__gte=current_month_start
    )
    this_month_revenue = this_month_transactions.aggregate(total=Sum('payment_amount'))['total'] or Decimal('0.00')
    
    # Get unread booking notifications for THIS specific church
    # Only count REQUESTED bookings (pending appointments) for the badge
    unread_booking_notifications_for_church = Notification.objects.filter(
        user=request.user,
        is_read=False,
        notification_type=Notification.TYPE_BOOKING_REQUESTED,
        church=church
    ).count()
    
    # Get staff members for Parish Admins tab
    secretaries = ChurchStaff.objects.filter(
        church=church,
        role=ChurchStaff.ROLE_SECRETARY,
        status=ChurchStaff.STATUS_ACTIVE
    ).select_related('user', 'user__profile').order_by('-added_at')
    
    volunteers = ChurchStaff.objects.filter(
        church=church,
        role=ChurchStaff.ROLE_VOLUNTEER,
        status=ChurchStaff.STATUS_ACTIVE
    ).select_related('user', 'user__profile').order_by('-added_at')

    ctx = {
        'active': 'manage',
        'page_title': 'Manage Church',
        'church': church,
        'form': form,
        'verif_form': verif_form,
        'follower_count': follower_count,
        'recent_followers': recent_followers,
        'followers_this_week': followers_this_week,
        'growth_rate': growth_rate,
        'active_followers': active_followers,
        'bookings': booking_list,
        'booking_counts': counts,
        'appt_status': appt_status,
        'latest_verification': church.verification_requests.order_by('-created_at').first(),
        'decline_reasons': list(church.decline_reasons.order_by('order', 'id')),
        'unread_booking_notifications': unread_booking_notifications_for_church,
        'posts': posts,
        'posts_count': posts_count,
        'event_posts': event_posts,
        'event_posts_count': event_posts_count,
        # Analytics data
        'analytics': {
            'total_posts': total_posts,
            'posts_last_30_days': posts_last_30_days,
            'posts_last_7_days': posts_last_7_days,
            'total_likes': total_likes,
            'total_comments': total_comments,
            'total_views': total_views,
            'likes_last_30_days': likes_last_30_days,
            'comments_last_30_days': comments_last_30_days,
            'views_last_30_days': views_last_30_days,
            # For content_sub_tabs charts and ranking
            'most_engaged_post': most_engaged_post,
            'top_5_posts': top_5_posts,
            'max_engagement': max_engagement,
            'top_posts': top_posts,
            'followers_last_30_days': followers_last_30_days,
            'bookings_last_30_days': bookings_last_30_days,
        },
        # Donations data
        'donations': {
            'recent_donations': recent_donations,
            'total_donations': total_donations,
            'total_amount': total_amount,
            'this_month_amount': this_month_amount,
            'this_month_count': this_month_count,
            'this_year_amount': this_year_amount,
            'this_year_count': this_year_count,
            'unique_donors': unique_donors,
            'top_donors': top_donors,
            'average_monthly': average_monthly,
            'highest_month': highest_month,
            'growth_rate': growth_rate_donations,
        },
        # Transactions data
        'transactions': {
            'recent_transactions': recent_transactions,
            'total_revenue': total_revenue,
            'completed_transactions': completed_transactions,
            'pending_transactions': pending_transactions,
            'this_month_revenue': this_month_revenue,
        },
        # Services statistics
        'total_services': total_services,
        'most_booked_service': most_booked_service,
        'highest_rated_service': highest_rated_service,
        'lowest_rated_service': lowest_rated_service,
        'service_bookings_by_day': service_bookings_by_day,
        'top_services_by_bookings': top_services_by_bookings,
        'user_role': user_role,  # 'owner', 'secretary', 'volunteer'
        'staff_position': staff_position,  # ChurchStaff object if staff member
        'secretaries': secretaries,  # Parish Secretaries list
        'volunteers': volunteers,  # Ministry Leaders/Volunteers list
    }
    ctx.update(_app_context(request))
    # Override global unread count with church-specific pending count for this page
    ctx['unread_booking_notifications'] = unread_booking_notifications_for_church
    # AJAX partial for appointments list
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('partial') == 'appointments_list':
        return render(request, 'core/partials/appointments_list.html', {
            'bookings': booking_list,
        })
    return render(request, 'core/manage_church.html', ctx)


@login_required
def get_church_followers_list(request, church_id):
    """API endpoint to get followers list for adding staff members."""
    try:
        church = get_object_or_404(Church, id=church_id)
        
        # Only church owner can manage staff
        can_manage, role = user_can_manage_church(request.user, church)
        
        if not can_manage:
            return JsonResponse({'success': False, 'message': 'You do not have permission to manage this church'}, status=403)
        
        if role != 'owner':
            return JsonResponse({'success': False, 'message': 'Only the parish manager can add staff members'}, status=403)
        
        # Get all followers
        followers = ChurchFollow.objects.filter(church=church).select_related('user', 'user__profile')
        
        # Get existing staff members
        from core.models import ChurchStaff
        existing_staff_ids = ChurchStaff.objects.filter(
            church=church,
            status=ChurchStaff.STATUS_ACTIVE
        ).values_list('user_id', flat=True)
        
        followers_data = []
        for follow in followers:
            user = follow.user
            profile = user.profile if hasattr(user, 'profile') else None
            
            followers_data.append({
                'id': user.id,
                'name': user.get_full_name() or user.username,
                'email': user.email,
                'avatar': profile.profile_image.url if profile and profile.profile_image else None,
                'is_staff': user.id in existing_staff_ids
            })
        
        return JsonResponse({
            'success': True,
            'followers': followers_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def add_church_staff(request, church_id):
    """API endpoint to add a staff member to the church."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=405)
    
    try:
        import json
        
        church = get_object_or_404(Church, id=church_id)
        
        # Only church owner can manage staff
        can_manage, role = user_can_manage_church(request.user, church)
        if not can_manage or role != 'owner':
            return JsonResponse({'success': False, 'message': 'Only the parish manager can add staff members'}, status=403)
        
        data = json.loads(request.body)
        user_id = data.get('user_id')
        role = data.get('role')
        
        if not user_id or not role:
            return JsonResponse({'success': False, 'message': 'Missing required fields'}, status=400)
        
        # Validate role
        if role not in [ChurchStaff.ROLE_SECRETARY, ChurchStaff.ROLE_VOLUNTEER]:
            return JsonResponse({'success': False, 'message': 'Invalid role'}, status=400)
        
        # Get the user
        staff_user = get_object_or_404(User, id=user_id)
        
        # Check if user follows the church
        if not ChurchFollow.objects.filter(church=church, user=staff_user).exists():
            return JsonResponse({'success': False, 'message': 'User must follow the church first'}, status=400)
        
        # Check if already exists
        existing = ChurchStaff.objects.filter(
            church=church,
            user=staff_user,
            role=role
        ).first()
        
        if existing:
            if existing.status == ChurchStaff.STATUS_INACTIVE:
                # Reactivate
                existing.status = ChurchStaff.STATUS_ACTIVE
                existing.save()
                return JsonResponse({'success': True, 'message': 'Staff member reactivated'})
            else:
                return JsonResponse({'success': False, 'message': 'User is already a staff member with this role'}, status=400)
        
        # Create new staff member
        ChurchStaff.objects.create(
            church=church,
            user=staff_user,
            role=role,
            added_by=request.user
        )
        
        return JsonResponse({'success': True, 'message': 'Staff member added successfully'})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def remove_church_staff(request, church_id, staff_id):
    """API endpoint to remove a staff member from the church."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=405)
    
    try:
        church = get_object_or_404(Church, id=church_id)
        
        # Only church owner can manage staff
        can_manage, role = user_can_manage_church(request.user, church)
        if not can_manage or role != 'owner':
            return JsonResponse({'success': False, 'message': 'Only the parish manager can remove staff members'}, status=403)
        
        # Get the staff member
        staff_member = get_object_or_404(ChurchStaff, id=staff_id, church=church)
        
        # Prevent removing if already inactive
        if staff_member.status == ChurchStaff.STATUS_INACTIVE:
            return JsonResponse({'success': False, 'message': 'Staff member is already inactive'}, status=400)
        
        # Set status to inactive instead of deleting
        staff_member.status = ChurchStaff.STATUS_INACTIVE
        staff_member.save()
        
        return JsonResponse({'success': True, 'message': 'Staff member removed successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def get_staff_activities(request, church_id, staff_id):
    """API endpoint to get activity logs for a staff member."""
    try:
        church = get_object_or_404(Church, id=church_id)
        
        # Only church owner can view staff activities
        can_manage, role = user_can_manage_church(request.user, church)
        if not can_manage or role != 'owner':
            return JsonResponse({'success': False, 'message': 'Only the parish manager can view staff activities'}, status=403)
        
        # Get the staff member
        staff_member = get_object_or_404(ChurchStaff, id=staff_id, church=church)
        
        # Get recent activities (last 50)
        from core.models import StaffActivityLog
        activities = StaffActivityLog.objects.filter(
            staff=staff_member
        ).select_related('staff', 'church')[:50]
        
        activities_data = []
        for activity in activities:
            activities_data.append({
                'id': activity.id,
                'action': activity.get_action_display(),
                'category': activity.get_category_display(),
                'description': activity.description,
                'created_at': activity.created_at.strftime('%b %d, %Y at %I:%M %p'),
                'ip_address': activity.ip_address or 'N/A',
            })
        
        return JsonResponse({
            'success': True,
            'activities': activities_data,
            'total_count': staff_member.activities.count()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def update_church_logo(request):
    """AJAX endpoint to update the church logo from Manage Church page."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

    # Get church_id from POST data
    church_id = request.POST.get('church_id')
    try:
        if church_id:
            church = request.user.owned_churches.get(id=church_id)
        else:
            church = request.user.owned_churches.first()
            if not church:
                return JsonResponse({'success': False, 'message': "You don't own any church."}, status=400)
    except Church.DoesNotExist:
        return JsonResponse({'success': False, 'message': "You don't have permission to update this church."}, status=403)

    file_obj = request.FILES.get('logo') or request.FILES.get('file') or request.FILES.get('image')
    if not file_obj:
        return JsonResponse({'success': False, 'message': 'No file uploaded.'}, status=400)

    # Basic validation
    allowed_types = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
    content_type = getattr(file_obj, 'content_type', '')
    if content_type not in allowed_types:
        return JsonResponse({'success': False, 'message': 'Unsupported file type.'}, status=400)
    max_size = 10 * 1024 * 1024  # 10MB
    if getattr(file_obj, 'size', 0) > max_size:
        return JsonResponse({'success': False, 'message': 'File too large (max 10MB).'}, status=400)

    import logging, traceback
    logger = logging.getLogger(__name__)
    try:
        # Optimize and save explicitly under churches/logos/
        try:
            optimized = optimize_image(file_obj, max_size=(400, 400), quality=85, format='JPEG')
        except Exception:
            logger.exception("[Update Logo] optimize_image failed; saving original")
            file_obj.seek(0)
            optimized = ContentFile(file_obj.read(), name=os.path.basename(getattr(file_obj, 'name', 'logo.jpg') or 'logo.jpg'))

        # Assign optimized content directly; upload_to will handle path
        church.logo = optimized
        church.save(update_fields=['logo'])

        # Build URL via field storage (Cloudinary in production)
        try:
            logo_url = church.logo.storage.url(church.logo.name)
        except Exception as url_err:
            logger.exception("[Update Logo] URL build error")
            logo_url = church.logo.url if getattr(church, 'logo', None) else None
        # Debug logging
        logger.info(f"[Update Logo] Default storage: {default_storage.__class__.__name__}")
        try:
            field_storage_name = church.logo.storage.__class__.__name__ if church.logo else 'None'
        except Exception:
            field_storage_name = 'Unknown'
        logger.info(f"[Update Logo] Field storage: {field_storage_name}")
        logger.info(f"[Update Logo] Name: {church.logo.name if church.logo else 'None'}")
        logger.info(f"[Update Logo] URL: {logo_url}")
        return JsonResponse({'success': True, 'url': logo_url, 'message': 'Logo updated successfully.'})
    except Exception as e:
        logger.exception("[Update Logo] Failed")
        return JsonResponse({'success': False, 'message': 'Failed to save logo.'}, status=500)


@login_required
def update_church_cover(request):
    """AJAX endpoint to update the church cover image from Manage Church page."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

    # Get church_id from POST data
    church_id = request.POST.get('church_id')
    
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        if church_id:
            try:
                church_id = int(church_id)
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'message': 'Invalid church ID.'}, status=400)
            church = request.user.owned_churches.get(id=church_id)
            logger.info(f"update_church_cover: Using church_id={church_id} from POST data")
        else:
            logger.warning(f"update_church_cover: No church_id provided for user {request.user.id}, falling back to first church")
            church = request.user.owned_churches.first()
            if not church:
                return JsonResponse({'success': False, 'message': "You don't own any church."}, status=400)
    except Church.DoesNotExist:
        return JsonResponse({'success': False, 'message': "You don't have permission to update this church."}, status=403)

    file_obj = request.FILES.get('cover_image') or request.FILES.get('file') or request.FILES.get('image')
    if not file_obj:
        return JsonResponse({'success': False, 'message': 'No file uploaded.'}, status=400)

    # Basic validation
    allowed_types = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
    content_type = getattr(file_obj, 'content_type', '')
    if content_type not in allowed_types:
        return JsonResponse({'success': False, 'message': 'Unsupported file type.'}, status=400)
    max_size = 10 * 1024 * 1024  # 10MB
    if getattr(file_obj, 'size', 0) > max_size:
        return JsonResponse({'success': False, 'message': 'File too large (max 10MB).'}, status=400)

    import logging, traceback
    logger = logging.getLogger(__name__)
    try:
        # Optimize and save explicitly under churches/covers/
        try:
            optimized = optimize_image(file_obj, max_size=(1280, 720), quality=85, format='JPEG')
        except Exception:
            logger.exception("[Update Cover] optimize_image failed; saving original")
            file_obj.seek(0)
            optimized = ContentFile(file_obj.read(), name=os.path.basename(getattr(file_obj, 'name', 'cover.jpg') or 'cover.jpg'))

        # Assign optimized content directly; upload_to will handle path
        church.cover_image = optimized
        church.save(update_fields=['cover_image'])

        # Build URL via field storage
        try:
            cover_url = church.cover_image.storage.url(church.cover_image.name)
        except Exception as url_err:
            logger.exception("[Update Cover] URL build error")
            cover_url = church.cover_image.url if getattr(church, 'cover_image', None) else None
        # Debug logging
        logger.info(f"[Update Cover] Default storage: {default_storage.__class__.__name__}")
        try:
            field_storage_name = church.cover_image.storage.__class__.__name__ if church.cover_image else 'None'
        except Exception:
            field_storage_name = 'Unknown'
        logger.info(f"[Update Cover] Field storage: {field_storage_name}")
        logger.info(f"[Update Cover] Name: {church.cover_image.name if church.cover_image else 'None'}")
        logger.info(f"[Update Cover] URL: {cover_url}")
        return JsonResponse({'success': True, 'url': cover_url, 'message': 'Cover image updated successfully.'})
    except Exception as e:
        logger.exception("[Update Cover] Failed")
        return JsonResponse({'success': False, 'message': 'Failed to save cover image.'}, status=500)


@login_required
def follow_church(request, church_id):
    """Follow or unfollow a church."""
    if request.method == 'POST':
        church = get_object_or_404(Church, id=church_id, is_active=True)
        
        # Check if already following
        follow_obj, created = ChurchFollow.objects.get_or_create(
            user=request.user,
            church=church
        )
        
        if created:
            # Update follower count
            church.follower_count = church.followers.count()
            church.save(update_fields=['follower_count'])
            
            # Log activity
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_CHURCH_FOLLOW,
                content_object=church,
                request=request
            )
            
            return JsonResponse({
                'success': True,
                'action': 'followed',
                'message': f'You are now following {church.name}',
                'follower_count': church.follower_count
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'You are already following this church'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


@login_required
def unfollow_church(request, church_id):
    """Unfollow a church."""
    if request.method == 'POST':
        church = get_object_or_404(Church, id=church_id)
        
        try:
            follow_obj = ChurchFollow.objects.get(user=request.user, church=church)
            follow_obj.delete()
            
            # Update follower count
            church.follower_count = church.followers.count()
            church.save(update_fields=['follower_count'])
            
            # Log activity
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_CHURCH_UNFOLLOW,
                content_object=church,
                request=request
            )
            
            return JsonResponse({
                'success': True,
                'action': 'unfollowed',
                'message': f'You have unfollowed {church.name}',
                'follower_count': church.follower_count
            })
        except ChurchFollow.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'You are not following this church'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


@login_required
def events(request):
    """
    Events page:
    - Main feed (1 column): upcoming event posts from churches the user follows
    - Right sidebar: other upcoming events from churches the user is NOT following
    """
    now = timezone.now()

    # Get followed church IDs
    followed_ids = list(
        ChurchFollow.objects.filter(user=request.user).values_list('church_id', flat=True)
    )

    # Common filter for active event posts including:
    # - Upcoming events (start_date >= now)
    # - Ongoing events (end_date >= now)
    # - Undated events (both start and end are null)
    base_q = (
        Q(is_active=True, post_type='event', church__is_active=True)
        & (
            Q(event_start_date__gte=now)
            | Q(event_end_date__gte=now)
            | (Q(event_start_date__isnull=True) & Q(event_end_date__isnull=True))
        )
    )

    # Main events feed: from followed churches
    events_qs = (
        Post.objects.filter(base_q, church_id__in=followed_ids)
        .select_related('church')
    )

    # Annotate like/bookmark flags and counts (avoid N+1)
    if request.user.is_authenticated:
        events_qs = events_qs.annotate(
            is_liked=Exists(
                PostLike.objects.filter(user=request.user, post_id=OuterRef('pk'))
            ),
            is_bookmarked=Exists(
                PostBookmark.objects.filter(user=request.user, post_id=OuterRef('pk'))
            ),
        )
    else:
        events_qs = events_qs.annotate(
            is_liked=Value(False, output_field=BooleanField()),
            is_bookmarked=Value(False, output_field=BooleanField()),
        )

    events_qs = events_qs.annotate(
        likes_count=Count('likes', distinct=True),
        comments_count=Count('comments', filter=Q(comments__is_active=True), distinct=True),
    ).order_by('event_start_date', '-created_at')

    # Paginate main feed (1 column list)
    paginator = Paginator(events_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Sidebar: other upcoming events from non-followed churches
    # Use a small limit for quick sidebar rendering
    other_events = (
        Post.objects.filter(base_q)
        .exclude(church_id__in=followed_ids or [-1])
        .select_related('church')
        .order_by('event_start_date')[:5]
    )

    ctx = {
        'active': 'events',
        'page_title': 'Events',
        'events': page_obj,
        'other_events': list(other_events),
        'total_events': paginator.count,
        'followed_count': len(followed_ids),
    }
    ctx.update(_app_context(request))
    return render(request, 'core/events.html', ctx)


@login_required
def appointments(request):
    """User's appointments page with status filters and counts."""
    status_filter = request.GET.get('status', 'all')
    bookings_qs = Booking.objects.filter(user=request.user).select_related('service', 'service__category', 'church', 'church__owner', 'handled_by').order_by('-created_at')

    counts = {
        'all': bookings_qs.count(),
        Booking.STATUS_REQUESTED: bookings_qs.filter(status=Booking.STATUS_REQUESTED).count(),
        Booking.STATUS_REVIEWED: bookings_qs.filter(status=Booking.STATUS_REVIEWED).count(),
        Booking.STATUS_APPROVED: bookings_qs.filter(status=Booking.STATUS_APPROVED).count(),
        Booking.STATUS_COMPLETED: bookings_qs.filter(status=Booking.STATUS_COMPLETED).count(),
    }
    if status_filter and status_filter != 'all':
        bookings_qs = bookings_qs.filter(status=status_filter)

    # Add review status for each booking
    bookings = list(bookings_qs)
    for booking in bookings:
        booking.user_has_reviewed = booking.service.has_user_reviewed(request.user)

    ctx = {
        'active': 'appointments',
        'page_title': 'My Appointments',
        'bookings': bookings,
        'counts': counts,
        'status_filter': status_filter,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/my_appointments.html', ctx)


@login_required
@require_http_methods(["POST"])
def cancel_booking(request, booking_id):
    """Cancel a booking (user can only cancel their own pending/requested bookings)."""
    try:
        booking = Booking.objects.get(id=booking_id, user=request.user)
        
        # Only allow cancellation of pending or requested bookings
        if booking.status not in ['pending', 'requested']:
            return JsonResponse({
                'success': False,
                'message': 'Only pending or requested bookings can be cancelled.'
            }, status=400)
        
        # Update booking status to canceled
        booking.status = 'canceled'
        booking.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Booking cancelled successfully.'
        })
        
    except Booking.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Booking not found or you do not have permission to cancel it.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
def following(request):
    """View churches the user is following."""
    followed_churches = Church.objects.filter(
        followers__user=request.user,
        is_active=True
    ).select_related('owner').annotate(
        followers_count=Count('followers'),
        followed_at=F('followers__followed_at')
    ).order_by('-followed_at')
    
    # Get user address for distance calculation
    user_address_data = {}
    if request.user.is_authenticated:
        try:
            profile = request.user.profile
            user_address_data = {
                'user_street': profile.street_address or '',
                'user_barangay': profile.barangay or '',
                'user_city': profile.city_municipality or '',
                'user_province': profile.province or '',
            }
        except:
            user_address_data = {
                'user_street': '',
                'user_barangay': '',
                'user_city': '',
                'user_province': '',
            }
    
    ctx = {
        'active': 'following',
        'page_title': 'Following',
        'followed_churches': followed_churches,
    }
    ctx.update(user_address_data)
    ctx.update(_app_context(request))
    return render(request, 'core/following.html', ctx)


@login_required
def manage(request):
    """Redirect to manage church or create church."""
    if request.user.owned_churches.exists():
        return redirect('core:manage_church')
    else:
        if request.user.is_superuser:
            return redirect('core:super_admin_create_church')
        messages.info(
            request,
            "Church creation is restricted to Super Admin. Please contact an administrator to create a church and assign you as manager."
        )
        return redirect('core:home')


@login_required
def settings_page(request):
    ctx = {
        'active': 'settings',
        'page_title': 'Settings',
    }
    ctx.update(_app_context(request))
    return render(request, 'app/page.html', ctx)


@login_required
def super_admin_dashboard(request):
    """Super Admin dashboard for system-wide management.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    User = get_user_model()
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Count, Sum, Q
    from accounts.models import Profile
    from .models import Post, PostLike, PostComment, PostBookmark, ChurchFollow

    # User Statistics
    total_users = User.objects.count()
    
    # Active users (logged in within last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    active_users = User.objects.filter(last_login__gte=thirty_days_ago).count() if hasattr(User, 'last_login') else 0
    
    # New users this month
    now = timezone.now()
    first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    new_users_this_month = User.objects.filter(date_joined__gte=first_day_of_month).count() if hasattr(User, 'date_joined') else 0
    
    # Recent registrations (last 10 users)
    recent_registrations = User.objects.order_by('-date_joined')[:10] if hasattr(User, 'date_joined') else User.objects.order_by('-id')[:10]
    
    # Daily active users (last 14 days)
    daily_active_users = []
    daily_labels = []
    if hasattr(User, 'last_login'):
        for i in range(13, -1, -1):  # Last 14 days
            day = timezone.now() - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            
            count = User.objects.filter(
                last_login__gte=day_start,
                last_login__lt=day_end
            ).count()
            
            daily_active_users.append(count)
            daily_labels.append(day.strftime('%b %d'))  # e.g., "Oct 05"
    
    # Church Statistics
    total_churches = Church.objects.count()
    
    # Verified churches (assuming there's a verification status)
    verified_churches = Church.objects.filter(is_verified=True).count() if hasattr(Church, 'is_verified') else 0
    
    # New churches this month
    new_churches_this_month = Church.objects.filter(created_at__gte=first_day_of_month).count()
    
    # Top churches by followers (Top 10)
    # Using existing follower_count field instead of annotation
    top_churches_by_followers = Church.objects.order_by('-follower_count')[:10]
    
    # Other stats
    total_services = BookableService.objects.count()
    total_bookings = Booking.objects.count()

    # Recent activity
    recent_churches = Church.objects.select_related('owner').order_by('-created_at')[:8]
    recent_users = User.objects.order_by('-date_joined')[:8] if hasattr(User, 'date_joined') else User.objects.order_by('-id')[:8]
    recent_bookings = Booking.objects.select_related('service', 'service__category', 'church', 'user').order_by('-created_at')[:8]

    # Booking status breakdown
    booking_by_status = list(
        Booking.objects.values('status').annotate(count=Count('id')).order_by('-count')
    )

    # Church verification snapshot for dashboard
    try:
        pending_verifications = list(
            ChurchVerificationRequest.objects
            .select_related('church', 'submitted_by')
            .filter(status=ChurchVerificationRequest.STATUS_PENDING)
            .order_by('created_at')[:5]
        )
        verif_counts = {
            'pending': ChurchVerificationRequest.objects.filter(status=ChurchVerificationRequest.STATUS_PENDING).count(),
            'approved': ChurchVerificationRequest.objects.filter(status=ChurchVerificationRequest.STATUS_APPROVED).count(),
            'rejected': ChurchVerificationRequest.objects.filter(status=ChurchVerificationRequest.STATUS_REJECTED).count(),
        }
    except Exception:
        pending_verifications = []
        verif_counts = {'pending': 0, 'approved': 0, 'rejected': 0}
    
    # Post Statistics
    total_posts = Post.objects.filter(is_active=True).count()
    posts_this_month = Post.objects.filter(
        created_at__gte=first_day_of_month,
        is_active=True
    ).count()
    
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    posts_today = Post.objects.filter(
        created_at__gte=today_start,
        is_active=True
    ).count()
    
    # Debug: Print to console to verify
    print(f"DEBUG - Total Posts: {total_posts}, Posts This Month: {posts_this_month}, Posts Today: {posts_today}")
    print(f"DEBUG - First day of month: {first_day_of_month}")
    print(f"DEBUG - Today start: {today_start}")
    
    # Show when the 3 posts were created
    recent_posts = Post.objects.filter(is_active=True).order_by('-created_at')[:5]
    for p in recent_posts:
        print(f"DEBUG - Post ID {p.id}: created_at = {p.created_at}")
    
    # Total engagement (likes + comments + bookmarks)
    total_likes = PostLike.objects.count()
    total_comments = PostComment.objects.filter(is_active=True).count()
    total_bookmarks = PostBookmark.objects.count()
    total_engagement = total_likes + total_comments + total_bookmarks
    
    # Church followers
    total_followers = ChurchFollow.objects.count()
    seven_days_ago = timezone.now() - timedelta(days=7)
    new_followers_this_week = ChurchFollow.objects.filter(
        followed_at__gte=seven_days_ago
    ).count()
    
    # Total donations (using Donation model)
    try:
        total_donations = Donation.objects.filter(payment_status='completed').aggregate(
            total=Sum('amount')
        )['total'] or 0
        donations_this_month = Donation.objects.filter(
            created_at__gte=first_day_of_month,
            payment_status='completed'
        ).aggregate(total=Sum('amount'))['total'] or 0
    except Exception as e:
        total_donations = 0
        donations_this_month = 0
    
    # Daily Post Engagement Data (last 7 days)
    daily_post_engagement = {
        'labels': [],
        'posts': [],
        'likes': [],
        'comments': []
    }
    
    for i in range(6, -1, -1):  # Last 7 days (Mon to Sun)
        day = timezone.now() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        # Day label
        daily_post_engagement['labels'].append(day.strftime('%a'))  # Mon, Tue, Wed, etc.
        
        # Posts count
        posts_count = Post.objects.filter(
            created_at__gte=day_start,
            created_at__lt=day_end,
            is_active=True
        ).count()
        daily_post_engagement['posts'].append(posts_count)
        
        # Likes count
        likes_count = PostLike.objects.filter(
            created_at__gte=day_start,
            created_at__lt=day_end
        ).count()
        daily_post_engagement['likes'].append(likes_count)
        
        # Comments count
        comments_count = PostComment.objects.filter(
            created_at__gte=day_start,
            created_at__lt=day_end,
            is_active=True
        ).count()
        daily_post_engagement['comments'].append(comments_count)
    
    # Parish Outreach Data (Active members by city/municipality)
    # Get active users (logged in within last 30 days) grouped by city
    parish_outreach = Profile.objects.filter(
        user__last_login__gte=thirty_days_ago,
        city_municipality__isnull=False
    ).exclude(
        city_municipality=''
    ).values('city_municipality').annotate(
        count=Count('id')
    ).order_by('-count')[:6]  # Top 6 cities
    
    parish_outreach_data = {
        'labels': [],
        'counts': []
    }
    
    for item in parish_outreach:
        city = item['city_municipality'] or 'Unknown'
        # Truncate long city names
        if len(city) > 20:
            city = city[:17] + '...'
        parish_outreach_data['labels'].append(city)
        parish_outreach_data['counts'].append(item['count'])
    
    # If we have less than 6 cities, ensure we have at least some data
    if len(parish_outreach_data['labels']) == 0:
        parish_outreach_data['labels'] = ['No Data']
        parish_outreach_data['counts'] = [0]
    
    # Service Statistics (for new stat cards)
    try:
        from .models import ServiceReview
        total_revenue = Booking.objects.filter(
            payment_status='paid'
        ).aggregate(total=Sum('service__price'))['total'] or 0
        
        avg_rating = ServiceReview.objects.aggregate(avg=Avg('rating'))['avg']
        if avg_rating:
            avg_rating = round(avg_rating, 1)
        else:
            avg_rating = 0
        
        total_reviews = ServiceReview.objects.count()
    except:
        total_revenue = 0
        avg_rating = 0
        total_reviews = 0
    
    # User Activity Trends (last 4 weeks for new chart)
    activity_labels = []
    active_users_data = []
    new_users_data = []
    
    for i in range(3, -1, -1):
        week_start = timezone.now() - timedelta(weeks=i+1)
        week_end = timezone.now() - timedelta(weeks=i)
        
        activity_labels.append(f'Week {4-i}')
        
        # Active users in that week
        if hasattr(User, 'last_login'):
            active_count = User.objects.filter(
                last_login__gte=week_start,
                last_login__lt=week_end,
                is_active=True
            ).count()
        else:
            active_count = 0
        active_users_data.append(active_count)
        
        # New users in that week
        if hasattr(User, 'date_joined'):
            new_count = User.objects.filter(
                date_joined__gte=week_start,
                date_joined__lt=week_end
            ).count()
        else:
            new_count = 0
        new_users_data.append(new_count)
    
    # Booking Trends (last 6 months for new chart)
    booking_trends_labels = []
    booking_trends_bookings = []
    booking_trends_revenue = []
    
    for i in range(5, -1, -1):
        month_date = timezone.now() - timedelta(days=30 * i)
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1)
        
        booking_trends_labels.append(month_start.strftime('%b'))
        
        # Booking count
        booking_count = Booking.objects.filter(
            created_at__gte=month_start,
            created_at__lt=month_end
        ).count()
        booking_trends_bookings.append(booking_count)
        
        # Revenue
        revenue = Booking.objects.filter(
            created_at__gte=month_start,
            created_at__lt=month_end,
            payment_status='paid'
        ).aggregate(total=Sum('service__price'))['total'] or 0
        booking_trends_revenue.append(float(revenue))
    
    # Top Donors (for new chart)
    top_donors_labels = []
    top_donors_amounts = []
    
    try:
        # Get top donors by completed donations only
        top_donors = Donation.objects.filter(
            payment_status='completed',
            donor__isnull=False
        ).values(
            'donor__first_name', 
            'donor__last_name',
            'donor__username'
        ).annotate(
            total=Sum('amount')
        ).order_by('-total')[:10]
        
        for donor in top_donors:
            first_name = donor.get('donor__first_name', '')
            last_name = donor.get('donor__last_name', '')
            username = donor.get('donor__username', 'Unknown')
            
            if first_name and last_name:
                name = f"{first_name} {last_name}"
            elif first_name:
                name = first_name
            else:
                name = username
            
            top_donors_labels.append(name)
            top_donors_amounts.append(float(donor['total']))
        
        # If no donors found, show "No Data"
        if not top_donors_labels:
            top_donors_labels = ['No Data']
            top_donors_amounts = [0]
    except Exception as e:
        # In case of any error, show "No Data"
        top_donors_labels = ['No Data']
        top_donors_amounts = [0]
    
    # Monthly growth data (last 6 months)
    monthly_growth_labels = []
    monthly_user_growth = []
    monthly_church_growth = []
    monthly_booking_data = {
        'completed': [],
        'pending': [],
        'cancelled': []
    }
    
    for i in range(5, -1, -1):  # Last 6 months
        month_date = timezone.now() - timedelta(days=30 * i)
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Calculate next month start
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1)
        
        # Month label
        monthly_growth_labels.append(month_start.strftime('%b'))
        
        # User growth
        user_count = User.objects.filter(date_joined__lt=month_end).count() if hasattr(User, 'date_joined') else 0
        monthly_user_growth.append(user_count)
        
        # Church growth
        church_count = Church.objects.filter(created_at__lt=month_end).count()
        monthly_church_growth.append(church_count)
        
        # Booking status data
        completed = Booking.objects.filter(
            created_at__gte=month_start,
            created_at__lt=month_end,
            status='completed'
        ).count()
        pending = Booking.objects.filter(
            created_at__gte=month_start,
            created_at__lt=month_end,
            status='pending'
        ).count()
        cancelled = Booking.objects.filter(
            created_at__gte=month_start,
            created_at__lt=month_end,
            status='cancelled'
        ).count()
        
        monthly_booking_data['completed'].append(completed)
        monthly_booking_data['pending'].append(pending)
        monthly_booking_data['cancelled'].append(cancelled)

    ctx = {
        'active': 'super_admin',
        'page_title': 'System Overview',
        'user_stats': {
            'total_users': total_users,
            'active_users': active_users,
            'new_users_this_month': new_users_this_month,
            'active_percentage': round((active_users / total_users * 100) if total_users > 0 else 0, 1),
            'inactive_users': max(total_users - active_users, 0),
        },
        'daily_active_users': daily_active_users,
        'daily_labels': daily_labels,
        'church_stats': {
            'total_churches': total_churches,
            'verified_churches': verified_churches,
            'new_churches_this_month': new_churches_this_month,
            'verified_percentage': round((verified_churches / total_churches * 100) if total_churches > 0 else 0, 1),
            'unverified_churches': max(total_churches - verified_churches, 0),
        },
        'top_churches_by_followers': top_churches_by_followers,
        'recent_registrations': recent_registrations,
        'stats': {
            'users': total_users,
            'churches': total_churches,
            'services': total_services,
            'bookings': total_bookings,
        },
        'booking_by_status': booking_by_status,
        'recent_churches': recent_churches,
        'recent_users': recent_users,
        'recent_bookings': recent_bookings,
        'verif': {
            'counts': verif_counts,
            'pending_list': pending_verifications,
        },
        'monthly_growth': {
            'labels': monthly_growth_labels,
            'users': monthly_user_growth,
            'churches': monthly_church_growth,
        },
        'monthly_bookings': monthly_booking_data,
        'post_stats': {
            'total_posts': total_posts,
            'posts_this_month': posts_this_month,
            'posts_today': posts_today,
            'total_engagement': total_engagement,
            'total_likes': total_likes,
            'total_comments': total_comments,
            'total_bookmarks': total_bookmarks,
            'total_followers': total_followers,
            'new_followers_this_week': new_followers_this_week,
            'total_donations': total_donations,
            'donations_this_month': donations_this_month,
        },
        'daily_post_engagement': daily_post_engagement,
        'parish_outreach': parish_outreach_data,
        # New data for enhanced Overview tab
        'service_stats': {
            'total_services': total_services,
            'total_revenue': total_revenue,
            'avg_rating': avg_rating,
            'total_reviews': total_reviews,
        },
        'activity_labels': activity_labels,
        'active_users_data': active_users_data,
        'new_users_data': new_users_data,
        'booking_trends_labels': booking_trends_labels,
        'booking_trends_bookings': booking_trends_bookings,
        'booking_trends_revenue': booking_trends_revenue,
        'top_donors_labels': top_donors_labels,
        'top_donors_amounts': top_donors_amounts,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin.html', ctx)


@login_required
def super_admin_export(request, format_type):
    """Export super admin overview data in various formats (CSV, Excel, PDF).
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    import csv
    import io
    from django.utils import timezone
    from datetime import timedelta, datetime
    
    User = get_user_model()
    thirty_days_ago = timezone.now() - timedelta(days=30)
    first_day_of_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Gather all statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(last_login__gte=thirty_days_ago).count() if hasattr(User, 'last_login') else 0
    new_users_this_month = User.objects.filter(date_joined__gte=first_day_of_month).count() if hasattr(User, 'date_joined') else 0
    
    total_churches = Church.objects.count()
    verified_churches = Church.objects.filter(is_verified=True).count() if hasattr(Church, 'is_verified') else 0
    new_churches_this_month = Church.objects.filter(created_at__gte=first_day_of_month).count()
    
    total_services = BookableService.objects.count()
    total_bookings = Booking.objects.count()
    
    total_posts = Post.objects.filter(is_active=True).count()
    posts_this_month = Post.objects.filter(created_at__gte=first_day_of_month, is_active=True).count()
    
    total_likes = PostLike.objects.count()
    total_comments = PostComment.objects.filter(is_active=True).count()
    total_bookmarks = PostBookmark.objects.count()
    total_engagement = total_likes + total_comments + total_bookmarks
    
    total_followers = ChurchFollow.objects.count()
    
    # Try to get donation data
    try:
        from .models import Donation
        total_donations = Donation.objects.filter(status='completed').aggregate(total=Sum('amount'))['total'] or 0
        donations_this_month = Donation.objects.filter(created_at__gte=first_day_of_month, status='completed').aggregate(total=Sum('amount'))['total'] or 0
    except:
        total_donations = 0
        donations_this_month = 0
    
    if format_type == 'csv':
        # Create CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="ChurchConnect_Overview_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ChurchConnect System Overview Report'])
        writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        
        writer.writerow(['Category', 'Metric', 'Value'])
        writer.writerow(['Users', 'Total Users', total_users])
        writer.writerow(['Users', 'Active Users (30 days)', active_users])
        writer.writerow(['Users', 'New Users This Month', new_users_this_month])
        writer.writerow([])
        
        writer.writerow(['Churches', 'Total Churches', total_churches])
        writer.writerow(['Churches', 'Verified Churches', verified_churches])
        writer.writerow(['Churches', 'New Churches This Month', new_churches_this_month])
        writer.writerow([])
        
        writer.writerow(['Services & Bookings', 'Total Services', total_services])
        writer.writerow(['Services & Bookings', 'Total Bookings', total_bookings])
        writer.writerow([])
        
        writer.writerow(['Posts & Engagement', 'Total Posts', total_posts])
        writer.writerow(['Posts & Engagement', 'Posts This Month', posts_this_month])
        writer.writerow(['Posts & Engagement', 'Total Likes', total_likes])
        writer.writerow(['Posts & Engagement', 'Total Comments', total_comments])
        writer.writerow(['Posts & Engagement', 'Total Bookmarks', total_bookmarks])
        writer.writerow(['Posts & Engagement', 'Total Engagement', total_engagement])
        writer.writerow(['Posts & Engagement', 'Total Followers', total_followers])
        writer.writerow([])
        
        writer.writerow(['Donations', 'Total Donations (₱)', total_donations])
        writer.writerow(['Donations', 'Donations This Month (₱)', donations_this_month])
        
        return response
    
    elif format_type == 'excel':
        # Create Excel file using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Overview"
            
            # Title
            ws['A1'] = 'ChurchConnect System Overview Report'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
            
            # Headers
            row = 4
            ws[f'A{row}'] = 'Category'
            ws[f'B{row}'] = 'Metric'
            ws[f'C{row}'] = 'Value'
            
            for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            
            # Data
            data = [
                ['Users', 'Total Users', total_users],
                ['Users', 'Active Users (30 days)', active_users],
                ['Users', 'New Users This Month', new_users_this_month],
                ['', '', ''],
                ['Churches', 'Total Churches', total_churches],
                ['Churches', 'Verified Churches', verified_churches],
                ['Churches', 'New Churches This Month', new_churches_this_month],
                ['', '', ''],
                ['Services & Bookings', 'Total Services', total_services],
                ['Services & Bookings', 'Total Bookings', total_bookings],
                ['', '', ''],
                ['Posts & Engagement', 'Total Posts', total_posts],
                ['Posts & Engagement', 'Posts This Month', posts_this_month],
                ['Posts & Engagement', 'Total Likes', total_likes],
                ['Posts & Engagement', 'Total Comments', total_comments],
                ['Posts & Engagement', 'Total Bookmarks', total_bookmarks],
                ['Posts & Engagement', 'Total Engagement', total_engagement],
                ['Posts & Engagement', 'Total Followers', total_followers],
                ['', '', ''],
                ['Donations', 'Total Donations (₱)', total_donations],
                ['Donations', 'Donations This Month (₱)', donations_this_month],
            ]
            
            for idx, row_data in enumerate(data, start=row+1):
                ws[f'A{idx}'] = row_data[0]
                ws[f'B{idx}'] = row_data[1]
                ws[f'C{idx}'] = row_data[2]
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            
            # Save to response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="ChurchConnect_Overview_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl library. Please install it.')
            return redirect('core:super_admin_dashboard')
    
    elif format_type == 'pdf':
        # Create PDF using reportlab
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.HexColor('#1E90FF'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            elements.append(Paragraph('ChurchConnect System Overview Report', title_style))
            elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Data table
            data = [
                ['Category', 'Metric', 'Value'],
                ['Users', 'Total Users', str(total_users)],
                ['Users', 'Active Users (30 days)', str(active_users)],
                ['Users', 'New Users This Month', str(new_users_this_month)],
                ['', '', ''],
                ['Churches', 'Total Churches', str(total_churches)],
                ['Churches', 'Verified Churches', str(verified_churches)],
                ['Churches', 'New Churches This Month', str(new_churches_this_month)],
                ['', '', ''],
                ['Services & Bookings', 'Total Services', str(total_services)],
                ['Services & Bookings', 'Total Bookings', str(total_bookings)],
                ['', '', ''],
                ['Posts & Engagement', 'Total Posts', str(total_posts)],
                ['Posts & Engagement', 'Posts This Month', str(posts_this_month)],
                ['Posts & Engagement', 'Total Likes', str(total_likes)],
                ['Posts & Engagement', 'Total Comments', str(total_comments)],
                ['Posts & Engagement', 'Total Bookmarks', str(total_bookmarks)],
                ['Posts & Engagement', 'Total Engagement', str(total_engagement)],
                ['Posts & Engagement', 'Total Followers', str(total_followers)],
                ['', '', ''],
                ['Donations', 'Total Donations (₱)', f'₱{total_donations}'],
                ['Donations', 'Donations This Month (₱)', f'₱{donations_this_month}'],
            ]
            
            table = Table(data, colWidths=[2*inch, 3*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E90FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="ChurchConnect_Overview_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response
            
        except ImportError:
            messages.error(request, 'PDF export requires reportlab library. Please install it.')
            return redirect('core:super_admin_dashboard')
    
    else:
        messages.error(request, 'Invalid export format.')
        return redirect('core:super_admin_dashboard')


@login_required
def super_admin_churches_export(request, format_type):
    """Export churches data in various formats (CSV, Excel).
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    import csv
    from django.utils import timezone
    
    # Get all churches with related data
    churches = Church.objects.select_related('owner').order_by('-created_at')
    
    if format_type == 'csv':
        # Create CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="ChurchConnect_Churches_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ChurchConnect Churches Export'])
        writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        
        # Headers
        writer.writerow([
            'Church Name',
            'Denomination',
            'City/Municipality',
            'Province',
            'Region',
            'Owner Name',
            'Owner Email',
            'Followers',
            'Verified',
            'Created Date',
            'Website',
            'Phone'
        ])
        
        # Data rows
        for church in churches:
            writer.writerow([
                church.name,
                church.denomination or 'N/A',
                church.city_municipality or church.city or 'N/A',
                church.province or 'N/A',
                church.region or 'N/A',
                church.owner.get_full_name() if church.owner else 'N/A',
                church.owner.email if church.owner else 'N/A',
                church.follower_count,
                'Yes' if church.is_verified else 'No',
                church.created_at.strftime('%Y-%m-%d'),
                church.website or 'N/A',
                church.phone or 'N/A'
            ])
        
        return response
    
    elif format_type == 'excel':
        # Create Excel file using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Churches"
            
            # Title
            ws['A1'] = 'ChurchConnect Churches Export'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
            
            # Headers
            headers = [
                'Church Name', 'Denomination', 'City/Municipality', 'Province', 'Region',
                'Owner Name', 'Owner Email', 'Followers', 'Verified', 'Created Date',
                'Website', 'Phone'
            ]
            
            row = 4
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            for idx, church in enumerate(churches, start=row+1):
                ws.cell(row=idx, column=1, value=church.name)
                ws.cell(row=idx, column=2, value=church.denomination or 'N/A')
                ws.cell(row=idx, column=3, value=church.city_municipality or church.city or 'N/A')
                ws.cell(row=idx, column=4, value=church.province or 'N/A')
                ws.cell(row=idx, column=5, value=church.region or 'N/A')
                ws.cell(row=idx, column=6, value=church.owner.get_full_name() if church.owner else 'N/A')
                ws.cell(row=idx, column=7, value=church.owner.email if church.owner else 'N/A')
                ws.cell(row=idx, column=8, value=church.follower_count)
                ws.cell(row=idx, column=9, value='Yes' if church.is_verified else 'No')
                ws.cell(row=idx, column=10, value=church.created_at.strftime('%Y-%m-%d'))
                ws.cell(row=idx, column=11, value=church.website or 'N/A')
                ws.cell(row=idx, column=12, value=church.phone or 'N/A')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 15
            ws.column_dimensions['K'].width = 25
            ws.column_dimensions['L'].width = 15
            
            # Save to response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="ChurchConnect_Churches_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl library. Please install it.')
            return redirect('core:super_admin_churches')
    
    else:
        messages.error(request, 'Invalid export format.')
        return redirect('core:super_admin_churches')


@login_required
def super_admin_posts_engagement_data(request):
    """API endpoint to fetch engagement data for different time ranges.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    from datetime import timedelta
    import json
    
    days = int(request.GET.get('days', 7))
    today = timezone.now().date()
    
    engagement_comments = []
    engagement_likes = []
    engagement_shares = []
    labels = []
    
    for i in range(days - 1, -1, -1):
        day = today - timedelta(days=i)
        day_comments = PostComment.objects.filter(created_at__date=day, is_active=True).count()
        day_likes = PostLike.objects.filter(created_at__date=day).count()
        day_shares = PostView.objects.filter(viewed_at__date=day).count()
        
        engagement_comments.append(day_comments)
        engagement_likes.append(day_likes)
        engagement_shares.append(day_shares)
        
        # Format label based on time range
        if days <= 7:
            labels.append(day.strftime('%a'))  # Mon, Tue, Wed
        elif days <= 30:
            labels.append(day.strftime('%m/%d'))  # 01/15
        else:
            labels.append(day.strftime('%m/%d'))  # 01/15
    
    return JsonResponse({
        'labels': labels,
        'comments': engagement_comments,
        'likes': engagement_likes,
        'shares': engagement_shares
    })


@login_required
def super_admin_posts_stats_data(request):
    """API endpoint to fetch posts statistics for different time periods.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    from datetime import timedelta
    from django.db.models import Sum
    
    period = request.GET.get('period', 'all')
    today = timezone.now().date()
    
    # Determine date filter
    if period == '7':
        start_date = today - timedelta(days=7)
        period_text = 'this week'
    elif period == '30':
        start_date = today - timedelta(days=30)
        period_text = 'this month'
    elif period == '90':
        start_date = today - timedelta(days=90)
        period_text = 'last 90 days'
    else:  # 'all'
        start_date = None
        period_text = 'all time'
    
    # Filter posts
    if start_date:
        posts_qs = Post.objects.filter(created_at__date__gte=start_date)
        likes_qs = PostLike.objects.filter(created_at__date__gte=start_date)
        comments_qs = PostComment.objects.filter(created_at__date__gte=start_date, is_active=True)
        views_qs = PostView.objects.filter(viewed_at__date__gte=start_date)
    else:
        posts_qs = Post.objects.all()
        likes_qs = PostLike.objects.all()
        comments_qs = PostComment.objects.filter(is_active=True)
        views_qs = PostView.objects.all()
    
    # Calculate stats
    total_posts = posts_qs.count()
    total_likes = likes_qs.count()
    total_comments = comments_qs.count()
    total_shares = views_qs.count()
    
    # Calculate averages
    avg_likes_per_post = round(total_likes / total_posts) if total_posts > 0 else 0
    avg_comments_per_post = round(total_comments / total_posts) if total_posts > 0 else 0
    avg_shares_per_post = round(total_shares / total_posts) if total_posts > 0 else 0
    
    # New posts text
    if period == 'all':
        new_posts_text = f'{total_posts} total'
    else:
        new_posts_text = f'+{total_posts} {period_text}'
    
    return JsonResponse({
        'total_posts': total_posts,
        'new_posts_text': new_posts_text,
        'total_likes': total_likes,
        'avg_likes_per_post': avg_likes_per_post,
        'total_comments': total_comments,
        'avg_comments_per_post': avg_comments_per_post,
        'total_shares': total_shares,
        'avg_shares_per_post': avg_shares_per_post
    })


@login_required
def super_admin_profile(request):
    """Super Admin's own profile page with admin-focused info and shortcuts.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin Profile.')
        return redirect('core:home')

    user = request.user
    
    # Handle POST request for updating credentials
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_credentials':
            username = request.POST.get('username', '').strip()
            email = request.POST.get('email', '').strip()
            
            # Validate inputs
            errors = []
            
            if username and username != user.username:
                # Check if username already exists
                from django.contrib.auth import get_user_model
                User = get_user_model()
                if User.objects.filter(username=username).exclude(pk=user.pk).exists():
                    errors.append('Username already exists.')
                else:
                    user.username = username
            
            if email and email != user.email:
                # Basic email validation
                import re
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_pattern, email):
                    errors.append('Invalid email format.')
                else:
                    # Check if email already exists
                    from django.contrib.auth import get_user_model
                    User = get_user_model()
                    if User.objects.filter(email=email).exclude(pk=user.pk).exists():
                        errors.append('Email already exists.')
                    else:
                        user.email = email
            
            if errors:
                for error in errors:
                    messages.error(request, error)
            else:
                user.save()
                messages.success(request, 'Credentials updated successfully.')
                return redirect('core:super_admin_profile')
        
        elif action == 'change_password':
            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')
            
            # Super admin can change password without current password verification
            # since they're already authenticated
            if len(new_password) < 8:
                messages.error(request, 'New password must be at least 8 characters long.')
            elif new_password != confirm_password:
                messages.error(request, 'New passwords do not match.')
            else:
                user.set_password(new_password)
                user.save()
                # Update session to prevent logout
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, user)
                messages.success(request, 'Password changed successfully.')
                return redirect('core:super_admin_profile')
    
    # Collect admin-related context
    groups = list(user.groups.values_list('name', flat=True))
    # Count direct and group-derived permissions
    try:
        from django.contrib.auth.models import Permission
        group_perm_count = Permission.objects.filter(group__user=user).distinct().count()
    except Exception:
        group_perm_count = 0
    direct_perm_count = getattr(user, 'user_permissions', None).count() if hasattr(user, 'user_permissions') else 0

    ctx = {
        'active': 'super_admin_profile',
        'page_title': 'Super Admin Profile',
        'admin_info': {
            'username': user.get_username(),
            'email': user.email,
            'last_login': user.last_login,
            'date_joined': getattr(user, 'date_joined', None),
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'is_active': user.is_active,
        },
        'admin_privileges': {
            'groups': groups,
            'direct_permissions_count': direct_perm_count,
            'group_permissions_count': group_perm_count,
        },
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_profile.html', ctx)


@login_required
def super_admin_churches(request):
    """Super Admin churches management page.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    from django.db.models import Count, Avg
    from django.core.paginator import Paginator
    
    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    verified_filter = request.GET.get('verified', '')
    denom_filter = request.GET.get('denomination', '')
    
    # Get all churches with related data
    churches = Church.objects.select_related('owner').annotate(
        posts_count=Count('posts', filter=Q(posts__is_active=True)),
        verification_pending=Exists(
            ChurchVerificationRequest.objects.filter(
                church_id=OuterRef('pk'),
                status=ChurchVerificationRequest.STATUS_PENDING
            )
        )
    )
    
    # Apply filters
    if search_query:
        churches = churches.filter(
            Q(name__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(owner__email__icontains=search_query) |
            Q(owner__first_name__icontains=search_query) |
            Q(owner__last_name__icontains=search_query)
        )
    
    if verified_filter == 'verified':
        churches = churches.filter(is_verified=True)
    elif verified_filter == 'unverified':
        churches = churches.filter(is_verified=False)
    
    if denom_filter:
        churches = churches.filter(denomination=denom_filter)
    
    churches = churches.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(churches, 20)  # 20 churches per page
    page_number = request.GET.get('page', 1)
    churches_page = paginator.get_page(page_number)
    
    # Statistics (on all churches, not filtered)
    all_churches = Church.objects.all()
    total_churches = all_churches.count()
    total_followers = ChurchFollow.objects.count()
    total_cities = all_churches.values('city').distinct().count()
    avg_followers_per_church = round(total_followers / total_churches) if total_churches > 0 else 0
    pending_verifications = ChurchVerificationRequest.objects.filter(
        status=ChurchVerificationRequest.STATUS_PENDING
    ).count()
    
    # Top churches by followers (limit to 6 for chart)
    top_churches = Church.objects.order_by('-follower_count')[:6]
    
    # Denomination choices for filter dropdown
    denomination_choices = Church.DENOMINATION_CHOICES
    
    # Get verification requests for the verifications tab
    verification_requests = ChurchVerificationRequest.objects.select_related(
        'church', 'submitted_by', 'reviewed_by'
    ).prefetch_related('documents').order_by('-created_at')
    
    ctx = {
        'active': 'super_admin_churches',
        'page_title': 'Churches Management',
        'churches_page': churches_page,
        'top_churches': top_churches,
        'stats': {
            'total_churches': total_churches,
            'total_followers': total_followers,
            'total_cities': total_cities,
            'avg_followers_per_church': avg_followers_per_church,
            'pending_verifications': pending_verifications,
        },
        'search_query': search_query,
        'verified_filter': verified_filter,
        'denom_filter': denom_filter,
        'denomination_choices': denomination_choices,
        'verification_requests': verification_requests,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_churches.html', ctx)


@login_required
def super_admin_create_church(request):
    """Super Admin view to create a church and assign a user as manager.
    Access is restricted to superusers.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    if request.method == 'POST':
        form = SuperAdminChurchCreateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                church = form.save()
                assigned_user = form.cleaned_data.get('assigned_user')
                if assigned_user:
                    messages.success(
                        request, 
                        f'Church "{church.name}" has been created successfully and assigned to {assigned_user.get_full_name() or assigned_user.email}!'
                    )
                else:
                    messages.success(
                        request, 
                        f'Church "{church.name}" has been created successfully without a manager.'
                    )
                return redirect('core:super_admin_church_detail', church_id=church.id)
            except Exception as e:
                logger.error(f"Error creating church: {e}", exc_info=True)
                messages.error(request, f'An error occurred while creating the church: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
            logger.warning(f"Form validation errors: {form.errors}")
    else:
        form = SuperAdminChurchCreateForm()
    
    ctx = {
        'active': 'super_admin_churches',
        'page_title': 'Create New Church',
        'form': form,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_create_church.html', ctx)


@login_required
def super_admin_edit_church(request, church_id):
    """Super Admin view to edit a church and reassign manager if needed.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    church = get_object_or_404(Church, id=church_id)
    
    if request.method == 'POST':
        form = SuperAdminChurchCreateForm(request.POST, request.FILES, instance=church)
        if form.is_valid():
            church = form.save()
            assigned_user = form.cleaned_data.get('assigned_user')
            if assigned_user:
                messages.success(
                    request, 
                    f'Church "{church.name}" has been updated successfully! Manager: {assigned_user.get_full_name() or assigned_user.email}'
                )
            else:
                messages.success(
                    request, 
                    f'Church "{church.name}" has been updated successfully! Church is now managerless.'
                )
            return redirect('core:super_admin_church_detail', church_id=church.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SuperAdminChurchCreateForm(instance=church)
    
    ctx = {
        'active': 'super_admin_churches',
        'page_title': f'Edit Church - {church.name}',
        'form': form,
        'church': church,
        'is_edit': True,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_create_church.html', ctx)


@login_required
def api_user_profile(request, user_id):
    """API endpoint to fetch user profile data for auto-fill.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = get_object_or_404(User, id=user_id)
        profile = user.profile
        
        return JsonResponse({
            'success': True,
            'full_name': user.get_full_name() or profile.display_name or '',
            'email': user.email or '',
            'phone': profile.phone or '',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
def super_admin_users(request):
    """Super Admin users management page.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    from django.db.models import Count
    from django.core.paginator import Paginator
    from datetime import timedelta
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    
    # Get all users with related data
    users = User.objects.annotate(
        churches_count=Count('owned_churches', distinct=True),
        posts_count=Count('owned_churches__posts', distinct=True),
        appointments_count=Count('bookings', distinct=True)
    ).order_by('-date_joined')
    
    # Pagination
    paginator = Paginator(users, 20)  # 20 users per page
    page_number = request.GET.get('page', 1)
    users_page = paginator.get_page(page_number)
    
    # Statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    church_admins = User.objects.filter(owned_churches__isnull=False).distinct().count()
    inactive_users = User.objects.filter(is_active=False).count()
    
    # Calculate percentages
    active_percentage = round((active_users / total_users * 100)) if total_users > 0 else 0
    inactive_percentage = round((inactive_users / total_users * 100)) if total_users > 0 else 0
    
    # New users this week
    week_ago = timezone.now() - timedelta(days=7)
    new_users_this_week = User.objects.filter(date_joined__gte=week_ago).count()
    
    # Total churches for church admins stat
    total_churches = Church.objects.count()
    
    # Activity trends data (last 4 weeks)
    activity_labels = []
    active_users_data = []
    new_users_data = []
    
    for i in range(3, -1, -1):
        week_start = timezone.now() - timedelta(weeks=i+1)
        week_end = timezone.now() - timedelta(weeks=i)
        activity_labels.append(f'Week {4-i}')
        
        # Active users (users who logged in during that week)
        active_count = User.objects.filter(
            last_login__gte=week_start,
            last_login__lt=week_end
        ).count()
        active_users_data.append(active_count)
        
        # New users (users who joined during that week)
        new_count = User.objects.filter(
            date_joined__gte=week_start,
            date_joined__lt=week_end
        ).count()
        new_users_data.append(new_count)
    
    # Users by role data
    regular_users = User.objects.filter(
        is_superuser=False,
        owned_churches__isnull=True
    ).count()
    
    role_labels = ['Regular Users', 'Church Admins', 'Service Providers', 'Inactive']
    role_data = [
        regular_users,
        church_admins,
        0,  # Service providers (if you have this role)
        inactive_users
    ]
    
    import json
    
    ctx = {
        'active': 'super_admin_users',
        'page_title': 'Users Management',
        'users_page': users_page,
        'stats': {
            'total_users': total_users,
            'active_users': active_users,
            'church_admins': church_admins,
            'suspended_users': inactive_users,
            'active_percentage': active_percentage,
            'suspended_percentage': inactive_percentage,
            'new_users_this_week': new_users_this_week,
            'total_churches': total_churches,
        },
        'activity_labels': json.dumps(activity_labels),
        'active_users_data': json.dumps(active_users_data),
        'new_users_data': json.dumps(new_users_data),
        'role_labels': json.dumps(role_labels),
        'role_data': json.dumps(role_data),
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_users.html', ctx)


@login_required
def super_admin_managers(request):
    """Super Admin parish managers management page.
    Shows all users who manage parishes.
    Access is restricted to superusers.
    """
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    from django.db.models import Count, Q
    from django.core.paginator import Paginator
    from datetime import timedelta
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    
    # Get all users who own/manage at least one church
    managers = User.objects.filter(
        owned_churches__isnull=False
    ).annotate(
        churches_count=Count('owned_churches', distinct=True),
        verified_churches_count=Count('owned_churches', filter=Q(owned_churches__is_verified=True), distinct=True),
        active_churches_count=Count('owned_churches', filter=Q(owned_churches__is_active=True), distinct=True),
        posts_count=Count('owned_churches__posts', distinct=True),
        services_count=Count('owned_churches__bookable_services', distinct=True),
        bookings_count=Count('owned_churches__bookable_services__bookings', distinct=True)
    ).distinct().order_by('-date_joined')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        managers = managers.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(owned_churches__name__icontains=search_query)
        ).distinct()
    
    # Filter by status
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'active':
        managers = managers.filter(is_active=True)
    elif status_filter == 'inactive':
        managers = managers.filter(is_active=False)
    elif status_filter == 'verified':
        managers = managers.filter(owned_churches__is_verified=True).distinct()
    
    # Pagination
    paginator = Paginator(managers, 20)  # 20 managers per page
    page_number = request.GET.get('page', 1)
    managers_page = paginator.get_page(page_number)
    
    # Statistics
    total_managers = User.objects.filter(owned_churches__isnull=False).distinct().count()
    active_managers = User.objects.filter(owned_churches__isnull=False, is_active=True).distinct().count()
    verified_managers = User.objects.filter(owned_churches__is_verified=True).distinct().count()
    inactive_managers = User.objects.filter(owned_churches__isnull=False, is_active=False).distinct().count()
    
    # Calculate percentages
    active_percentage = round((active_managers / total_managers * 100)) if total_managers > 0 else 0
    verified_percentage = round((verified_managers / total_managers * 100)) if total_managers > 0 else 0
    
    # New managers this week
    week_ago = timezone.now() - timedelta(days=7)
    new_managers_this_week = User.objects.filter(
        owned_churches__isnull=False,
        date_joined__gte=week_ago
    ).distinct().count()
    
    # Total parishes managed
    total_parishes = Church.objects.filter(owner__isnull=False).count()
    
    # Manager activity trends (last 4 weeks)
    activity_labels = []
    active_managers_data = []
    new_managers_data = []
    
    for i in range(3, -1, -1):
        week_start = timezone.now() - timedelta(weeks=i+1)
        week_end = timezone.now() - timedelta(weeks=i)
        activity_labels.append(f'Week {4-i}')
        
        # Active managers (managers who logged in during that week)
        active_count = User.objects.filter(
            owned_churches__isnull=False,
            last_login__gte=week_start,
            last_login__lt=week_end
        ).distinct().count()
        active_managers_data.append(active_count)
        
        # New managers (users who became managers during that week)
        new_count = User.objects.filter(
            owned_churches__isnull=False,
            date_joined__gte=week_start,
            date_joined__lt=week_end
        ).distinct().count()
        new_managers_data.append(new_count)
    
    # Managers by parish count
    managers_with_1_parish = User.objects.filter(owned_churches__isnull=False).annotate(
        church_count=Count('owned_churches')
    ).filter(church_count=1).count()
    
    managers_with_2_parishes = User.objects.filter(owned_churches__isnull=False).annotate(
        church_count=Count('owned_churches')
    ).filter(church_count=2).count()
    
    managers_with_3plus_parishes = User.objects.filter(owned_churches__isnull=False).annotate(
        church_count=Count('owned_churches')
    ).filter(church_count__gte=3).count()
    
    parish_count_labels = ['1 Parish', '2 Parishes', '3+ Parishes']
    parish_count_data = [
        managers_with_1_parish,
        managers_with_2_parishes,
        managers_with_3plus_parishes
    ]
    
    import json
    
    ctx = {
        'active': 'super_admin_managers',
        'page_title': 'Parish Managers',
        'managers_page': managers_page,
        'search_query': search_query,
        'status_filter': status_filter,
        'stats': {
            'total_managers': total_managers,
            'active_managers': active_managers,
            'verified_managers': verified_managers,
            'inactive_managers': inactive_managers,
            'active_percentage': active_percentage,
            'verified_percentage': verified_percentage,
            'new_managers_this_week': new_managers_this_week,
            'total_parishes': total_parishes,
        },
        'activity_labels': json.dumps(activity_labels),
        'active_managers_data': json.dumps(active_managers_data),
        'new_managers_data': json.dumps(new_managers_data),
        'parish_count_labels': json.dumps(parish_count_labels),
        'parish_count_data': json.dumps(parish_count_data),
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_managers.html', ctx)


# Bookable Services Management
@login_required
def manage_services(request):
    """Manage bookable services for a church."""
    # Get church_id from GET parameters
    church_id = request.GET.get('church_id')
    
    if church_id:
        church = get_object_or_404(Church, id=church_id)
        # Check if user can manage services (Owner or Secretary)
        can_manage, role = user_can_manage_church(request.user, church, ['services'])
        if not can_manage:
            messages.error(request, "You don't have permission to manage services.")
            return redirect('core:select_church')
    else:
        # Try to get first owned church
        church = request.user.owned_churches.first()
        if not church:
            # Try to get first church where user is staff with services permission
            staff_position = ChurchStaff.objects.filter(
                user=request.user,
                status=ChurchStaff.STATUS_ACTIVE,
                role=ChurchStaff.ROLE_SECRETARY
            ).select_related('church').first()
            
            if staff_position:
                church = staff_position.church
            else:
                if request.user.is_superuser:
                    return redirect('core:super_admin_create_church')
                messages.info(
                    request,
                    "You don't have permission to manage any parishes. Please contact a Parish Manager."
                )
                return redirect('core:home')
    
    # Search and filter functionality
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    price_filter = request.GET.get('price', '')
    
    services = church.bookable_services.all()
    
    # Apply search filter
    if search_query:
        services = services.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Apply status filter
    if status_filter == 'active':
        services = services.filter(is_active=True)
    elif status_filter == 'inactive':
        services = services.filter(is_active=False)
    
    # Apply price filter
    if price_filter == 'free':
        services = services.filter(is_free=True)
    elif price_filter == 'paid':
        services = services.filter(is_free=False)
    
    # Order by creation date
    services = services.order_by('-created_at')
    
    ctx = {
        'active': 'manage',
        'page_title': 'Manage Services',
        'church': church,
        'services': services,
        'search_query': search_query,
        'status_filter': status_filter,
        'price_filter': price_filter,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/manage_services.html', ctx)


@login_required
def create_service(request):
    """Create a new bookable service."""
    # Get church_id from GET or POST parameters
    church_id = request.GET.get('church_id') or request.POST.get('church_id')
    
    if church_id:
        church = get_object_or_404(Church, id=church_id)
        # Check if user can manage services (Owner or Secretary)
        can_manage, role = user_can_manage_church(request.user, church, ['services'])
        if not can_manage:
            messages.error(request, "You don't have permission to manage services.")
            return redirect('core:select_church')
    else:
        # Try to get first owned church
        church = request.user.owned_churches.first()
        if not church:
            # Try to get first church where user is staff with services permission
            staff_position = ChurchStaff.objects.filter(
                user=request.user,
                status=ChurchStaff.STATUS_ACTIVE,
                role=ChurchStaff.ROLE_SECRETARY
            ).select_related('church').first()
            
            if staff_position:
                church = staff_position.church
            else:
                if request.user.is_superuser:
                    return redirect('core:super_admin_create_church')
                messages.info(
                    request,
                    "You don't have permission to manage any parishes. Please contact a Parish Manager."
                )
                return redirect('core:home')
    
    if request.method == 'POST':
        form = BookableServiceForm(request.POST, request.FILES, church=church)
        if form.is_valid():
            service = form.save()
            
            # Log staff activity for service creation
            from .models import StaffActivityLog
            log_staff_activity(
                user=request.user,
                church=church,
                action=StaffActivityLog.ACTION_CREATE,
                category=StaffActivityLog.CATEGORY_SERVICE,
                description=f"Created service '{service.name}' - {service.get_currency_display()}{service.price}",
                target_id=service.id,
                target_type='service',
                request=request
            )
            
            # Handle multiple image uploads
            images = request.FILES.getlist('images')
            if images:
                for i, image in enumerate(images):
                    if image:  # Check if file exists
                        ServiceImage.objects.create(
                            service=service,
                            image=image,
                            order=i,
                            is_primary=(i == 0)  # First image is primary
                        )
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True, 
                    'message': f'Service "{service.name}" has been created successfully!',
                    'service_id': service.id
                })
            messages.success(request, f'Service "{service.name}" has been created successfully!')
            return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=services')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'message': 'Please correct the errors below.',
                    'errors': form.errors
                })
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BookableServiceForm(church=church)
    
    # Return form HTML for modal
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'core/partials/service_form.html', {
            'form': form,
            'church': church,
            'action': 'create'
        })
    
    ctx = {
        'active': 'manage',
        'page_title': 'Create Service',
        'church': church,
        'form': form,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/create_service.html', ctx)


@login_required
def edit_service(request, service_id):
    """Edit a bookable service."""
    # Get the service first to determine its church
    service = get_object_or_404(BookableService, id=service_id)
    church = service.church
    
    # Check if user can manage services (Owner or Secretary)
    can_manage, role = user_can_manage_church(request.user, church, ['services'])
    if not can_manage:
        messages.error(request, "You don't have permission to edit services.")
        return redirect('core:select_church')
    
    if request.method == 'POST':
        form = BookableServiceForm(request.POST, request.FILES, instance=service, church=church)
        if form.is_valid():
            form.save()
            
            # Log staff activity for service update
            from .models import StaffActivityLog
            log_staff_activity(
                user=request.user,
                church=church,
                action=StaffActivityLog.ACTION_UPDATE,
                category=StaffActivityLog.CATEGORY_SERVICE,
                description=f"Updated service '{service.name}'",
                target_id=service.id,
                target_type='service',
                request=request
            )
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True, 
                    'message': f'Service "{service.name}" has been updated successfully!',
                    'service_id': service.id
                })
            messages.success(request, f'Service "{service.name}" has been updated successfully!')
            return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=services')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'message': 'Please correct the errors below.',
                    'errors': form.errors
                })
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BookableServiceForm(instance=service, church=church)
    
    # Return form HTML for modal
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'core/partials/service_form.html', {
            'form': form,
            'church': church,
            'service': service,
            'action': 'edit'
        })
    
    ctx = {
        'active': 'manage',
        'page_title': f'Edit {service.name}',
        'church': church,
        'service': service,
        'form': form,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/edit_service.html', ctx)


@login_required
def delete_service(request, service_id):
    """Delete a bookable service."""
    # Get the service to determine its church
    service = get_object_or_404(BookableService, id=service_id)
    church = service.church
    
    # Check if user can manage services (Owner or Secretary)
    can_manage, role = user_can_manage_church(request.user, church, ['services'])
    if not can_manage:
        messages.error(request, "You don't have permission to delete services.")
        return redirect('core:select_church')
    
    if request.method == 'POST':
        service_name = service.name
        service_id = service.id
        
        # Log staff activity for service deletion
        from .models import StaffActivityLog
        log_staff_activity(
            user=request.user,
            church=church,
            action=StaffActivityLog.ACTION_DELETE,
            category=StaffActivityLog.CATEGORY_SERVICE,
            description=f"Deleted service '{service_name}'",
            target_id=service_id,
            target_type='service',
            request=request
        )
        
        service.delete()
        messages.success(request, f'Service "{service_name}" has been deleted successfully!')
        return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=services')
    
    # If GET request, show confirmation page (keeping old logic below for compatibility)
    try:
        old_church = request.user.owned_churches.first()
        if not old_church:
            if request.user.is_superuser:
                return redirect('core:super_admin_create_church')
            messages.info(
                request,
                "You don't own any churches yet. Please contact a Super Admin to create one and assign you as manager."
            )
            return redirect('core:home')
    except Church.DoesNotExist:
        if request.user.is_superuser:
            return redirect('core:super_admin_create_church')
        messages.info(
            request,
            "You don't own any churches yet. Please contact a Super Admin to create one and assign you as manager."
        )
        return redirect('core:home')
    
    service = get_object_or_404(BookableService, id=service_id, church=church)
    
    if request.method == 'POST':
        service_name = service.name
        service.delete()
        messages.success(request, f'Service "{service_name}" has been deleted successfully!')
        return HttpResponseRedirect(reverse('core:manage_church') + '?tab=services')
    
    ctx = {
        'active': 'manage',
        'page_title': f'Delete {service.name}',
        'church': church,
        'service': service,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/delete_service.html', ctx)
@login_required
def manage_service_images(request, service_id):
    """Manage service images (add/delete)."""
    service = get_object_or_404(BookableService, id=service_id)
    church = service.church
    
    # Check if user can manage services (Owner or Secretary)
    can_manage, role = user_can_manage_church(request.user, church, ['services'])
    if not can_manage:
        messages.error(request, "You don't have permission to manage service images.")
        return redirect('core:select_church')
    
    if request.method == 'POST':
        form = ServiceImageForm(request.POST, request.FILES)
        if form.is_valid():
            service_image = form.save(commit=False)
            service_image.service = service
            service_image.save()
            messages.success(request, f'Image added to "{service.name}" successfully!')
            return HttpResponseRedirect(reverse('core:manage_service_images', args=[service.id]))
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ServiceImageForm()
    
    ctx = {
        'active': 'manage',
        'page_title': f'Manage Images - {service.name}',
        'church': church,
        'service': service,
        'form': form,
        'images': service.get_images(),
    }
    ctx.update(_app_context(request))
    return render(request, 'core/manage_service_images.html', ctx)


@login_required
def delete_service_image(request, image_id):
    """Delete a service image."""
    service_image = get_object_or_404(ServiceImage, id=image_id)
    service = service_image.service
    church = service.church
    
    # Check if user can manage services (Owner or Secretary)
    can_manage, role = user_can_manage_church(request.user, church, ['services'])
    if not can_manage:
        messages.error(request, "You don't have permission to manage services.")
        return redirect('core:select_church')
    
    if request.method == 'POST':
        service_image.delete()
        messages.success(request, 'Image deleted successfully!')
        return HttpResponseRedirect(reverse('core:manage_service_images', args=[service.id]))
    
    ctx = {
        'active': 'manage',
        'page_title': f'Delete Image - {service.name}',
        'church': church,
        'service': service,
        'service_image': service_image,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/delete_service_image.html', ctx)


# Availability Management
@login_required
def manage_availability(request):
    """Manage church availability and closed dates."""
    # Get church_id from GET parameters
    church_id = request.GET.get('church_id')
    
    try:
        if church_id:
            church = Church.objects.get(id=church_id)
        else:
            # Fallback to first church if no church_id provided
            church = request.user.owned_churches.first()
            if not church:
                # Check if user is a secretary of any church
                secretary_church = ChurchStaff.objects.filter(
                    user=request.user,
                    role__in=['secretary', 'volunteer'],
                    status='active'
                ).select_related('church').first()
                
                if secretary_church:
                    church = secretary_church.church
                elif request.user.is_superuser:
                    return redirect('core:super_admin_create_church')
                else:
                    messages.info(
                        request,
                        "You don't own any churches yet. Please contact a Super Admin to create one and assign you as manager."
                    )
                    return redirect('core:home')
        
        # Check if user can manage availability (Owner or Secretary with availability permission)
        can_manage, role = user_can_manage_church(request.user, church, ['availability'])
        if not can_manage:
            messages.error(request, "You don't have permission to manage this church.")
            return redirect('core:select_church')
            
    except Church.DoesNotExist:
        messages.error(request, "You don't have permission to manage this church.")
        return redirect('core:select_church')
    
    # Get availability entries for the next 3 months
    from datetime import date, timedelta
    start_date = date.today()
    end_date = start_date + timedelta(days=90)
    
    availability_entries = church.availability.filter(
        date__range=[start_date, end_date]
    ).order_by('date')
    
    ctx = {
        'active': 'manage',
        'page_title': 'Manage Availability',
        'church': church,
        'availability_entries': availability_entries,
        'start_date': start_date,
        'end_date': end_date,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/manage_availability.html', ctx)


@login_required
def api_service_images(request, service_id):
    """API endpoint to fetch service images for gallery modal."""
    try:
        service = get_object_or_404(BookableService, id=service_id)
        church = service.church
        
        # Check if user can manage services (Owner or Secretary)
        can_manage, role = user_can_manage_church(request.user, church, ['services'])
        if not can_manage:
            return JsonResponse({
                'success': False,
                'error': "You don't have permission to manage this church."
            }, status=403)
        
        images = service.get_images()
        
        images_data = []
        for image in images:
            images_data.append({
                'id': image.id,
                'url': image.image.url,
                'caption': image.caption,
                'is_primary': image.is_primary,
                'order': image.order
            })
        
        return JsonResponse({
            'success': True,
            'images': images_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def service_gallery(request, service_id):
    """Display full image gallery for a service."""
    try:
        service = get_object_or_404(BookableService, id=service_id)
        church = service.church
        
        # Check if user can manage services (Owner or Secretary)
        can_manage, role = user_can_manage_church(request.user, church, ['services'])
        if not can_manage:
            messages.error(request, "You don't have permission to manage this church.")
            return redirect('core:select_church')
        
        images = service.get_images()
        
        context = _app_context(request)
        context.update({
            'service': service,
            'images': images,
            'title': f'{service.name} - Image Gallery'
        })
        
        return render(request, 'core/service_gallery.html', context)
    except Exception as e:
        messages.error(request, f'Error loading gallery: {str(e)}')
        return HttpResponseRedirect(reverse('core:manage_church') + '?tab=services')


@login_required
def book_service(request, service_id):
    """Booking page for a specific service. Shows upcoming dates and allows submitting a request."""
    service = get_object_or_404(BookableService, id=service_id, is_active=True)
    church = service.church

    # Block booking when church is not verified
    if not church.is_verified:
        messages.error(request, 'This church is not yet verified and cannot accept appointment requests.')
        return redirect('core:church_detail', slug=church.slug)

    # Enforce essential profile completeness before allowing booking
    try:
        profile = request.user.profile
    except Exception:
        profile = None
    essential = get_essential_profile_status(request.user, profile)
    if not essential.get('is_complete'):
        missing = essential.get('missing', [])
        # Build a friendly missing summary (limit to first 4)
        if missing:
            summary = ', '.join(missing[:4])
            if len(missing) > 4:
                summary += f" (+{len(missing) - 4} more)"
        else:
            summary = 'required information'
        messages.error(request, f'Please complete your profile before requesting an appointment. Missing: {summary}.')
        return redirect('dashboard')

    # Calculate date range for calendar constraints
    from datetime import date, timedelta
    start_date = date.today()
    end_date = start_date + timedelta(days=service.advance_booking_days)

    form = BookingForm(request.POST or None, service=service, user=request.user)
    if request.method == 'POST':
        if form.is_valid():
            booking = form.save()
            
            # Log activity
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_BOOKING_CREATE,
                content_object=booking,
                metadata={
                    'service_name': service.name,
                    'church_name': church.name,
                    'booking_code': booking.code,
                    'requested_date': booking.date.isoformat()
                },
                request=request
            )
            
            messages.success(request, f'Appointment request submitted. Your ID is {booking.code}.')
            return HttpResponseRedirect(reverse('core:appointments'))
        else:
            messages.error(request, 'Please fix the errors below.')

    ctx = {
        'active': 'discover',
        'page_title': f'Request Appointment - {service.name}',
        'church': church,
        'service': service,
        'form': form,
        'today': start_date,
        'max_date': end_date,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/book_service.html', ctx)


@login_required
def create_booking_api(request):
    """API endpoint for creating booking requests via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        service_id = data.get('service_id')
        
        if not service_id:
            return JsonResponse({'error': 'Service ID is required'}, status=400)
        
        service = get_object_or_404(BookableService, id=service_id, is_active=True)
        church = service.church
        
        # Block booking when church is not verified
        if not church.is_verified:
            return JsonResponse({
                'error': 'This church is not yet verified and cannot accept appointment requests.'
            }, status=400)
        
        # Check profile completeness
        try:
            profile = request.user.profile
        except Exception:
            profile = None
        essential = get_essential_profile_status(request.user, profile)
        if not essential.get('is_complete'):
            missing = essential.get('missing', [])
            return JsonResponse({
                'error': 'Please complete your profile before requesting an appointment.',
                'missing_fields': missing
            }, status=400)
        
        # Create form with data
        form_data = {
            'date': data.get('date'),
            'time': data.get('time'),
            'notes': data.get('notes', '')
        }
        
        form = BookingForm(form_data, service=service, user=request.user)
        if form.is_valid():
            booking = form.save()
            
            # Log activity
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_BOOKING_CREATE,
                content_object=booking,
                metadata={
                    'service_name': service.name,
                    'church_name': church.name,
                    'booking_code': booking.code,
                    'requested_date': booking.date.isoformat(),
                    'requested_time': booking.start_time.isoformat() if booking.start_time else None
                },
                request=request
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Appointment request submitted successfully.',
                'booking_code': booking.code
            })
        else:
            return JsonResponse({
                'error': 'Please fix the form errors.',
                'form_errors': form.errors
            }, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def manage_booking(request, booking_id):
    """Owner view to review/update a single booking and handle conflicts."""
    booking = get_object_or_404(Booking.objects.select_related('service', 'service__category', 'church', 'user'), id=booking_id)
    
    # Check if user can manage appointments (Owner or Secretary)
    can_manage, role = user_can_manage_church(request.user, booking.church, ['appointments'])
    if not can_manage:
        messages.error(request, 'You do not have permission to manage this booking.')
        return redirect('core:manage_church', church_id=booking.church.id)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        valid_statuses = {s for s, _ in Booking.STATUS_CHOICES}
        if new_status in valid_statuses and new_status != booking.status:
            # If declining, ensure a reason is provided
            if new_status == Booking.STATUS_DECLINED:
                selected = (request.POST.get('decline_reason') or '').strip()
                # If user selected custom, read from custom field
                if selected == '__custom__':
                    reason = (request.POST.get('decline_reason_custom') or '').strip()
                else:
                    reason = selected
                if not reason:
                    messages.error(request, 'Please select a reason for declining this appointment.')
                    return HttpResponseRedirect(reverse('core:manage_booking', args=[booking.id]))
                booking.decline_reason = reason[:200]

            old_status = booking.status
            booking.status = new_status
            booking.status_changed_at = timezone.now()
            booking.handled_by = request.user  # Track who handled this booking
            # Clear cancel reason when status is changed away from canceled
            if new_status != Booking.STATUS_CANCELED:
                booking.cancel_reason = ''
            # If status changed away from declined, clear decline reason
            if new_status != Booking.STATUS_DECLINED:
                booking.decline_reason = ''
            booking.save(update_fields=['status', 'status_changed_at', 'updated_at', 'cancel_reason', 'decline_reason', 'handled_by'])
            
            # Log staff activity for booking management
            from .models import StaffActivityLog
            action_map = {
                Booking.STATUS_APPROVED: (StaffActivityLog.ACTION_APPROVE, f"Approved booking #{booking.code} for {booking.user.get_full_name()}"),
                Booking.STATUS_DECLINED: (StaffActivityLog.ACTION_REJECT, f"Declined booking #{booking.code} for {booking.user.get_full_name()}: {booking.decline_reason}"),
                Booking.STATUS_CANCELED: (StaffActivityLog.ACTION_CANCEL, f"Canceled booking #{booking.code} for {booking.user.get_full_name()}"),
                Booking.STATUS_REVIEWED: (StaffActivityLog.ACTION_UPDATE, f"Reviewed booking #{booking.code} for {booking.user.get_full_name()}"),
                Booking.STATUS_COMPLETED: (StaffActivityLog.ACTION_UPDATE, f"Marked booking #{booking.code} as completed"),
            }
            if new_status in action_map:
                action, desc = action_map[new_status]
                log_staff_activity(
                    user=request.user,
                    church=booking.church,
                    action=action,
                    category=StaffActivityLog.CATEGORY_BOOKING,
                    description=desc,
                    target_id=booking.id,
                    target_type='booking',
                    request=request
                )
            
            # Log activity for booking update
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=booking.user,  # Log for the booking owner, not the church owner
                activity_type=UserInteraction.ACTIVITY_BOOKING_UPDATE,
                content_object=booking,
                metadata={
                    'old_status': old_status,
                    'new_status': new_status,
                    'booking_code': booking.code,
                    'service_name': booking.service.name,
                    'church_name': booking.church.name,
                    'updated_by': request.user.username,
                    'decline_reason': booking.decline_reason if new_status == Booking.STATUS_DECLINED else None
                },
                request=request
            )

            # Auto-cancel conflicts when approving
            if new_status == Booking.STATUS_APPROVED:
                conflicts = booking.conflicts_qs().filter(status__in=[Booking.STATUS_REQUESTED, Booking.STATUS_REVIEWED])
                for other in conflicts:
                    other.status = Booking.STATUS_CANCELED
                    other.cancel_reason = 'Conflicted with another approved booking. Please reschedule.'
                    other.status_changed_at = timezone.now()
                    other.save(update_fields=['status', 'cancel_reason', 'status_changed_at', 'updated_at'])
                    # Notify conflicting requester about auto-cancel
                    try:
                        tmpl = NotificationTemplates.booking_canceled(other)
                        create_booking_notification(
                            booking=other,
                            notification_type=Notification.TYPE_BOOKING_CANCELED,
                            title=tmpl['title'],
                            message=tmpl['message'],
                            priority=tmpl['priority']
                        )
                    except Exception:
                        pass

            # Send notification to requester for this booking based on new status
            try:
                if new_status == Booking.STATUS_REVIEWED:
                    tmpl = NotificationTemplates.booking_reviewed(booking)
                    create_booking_notification(
                        booking=booking,
                        notification_type=Notification.TYPE_BOOKING_REVIEWED,
                        title=tmpl['title'],
                        message=tmpl['message'],
                        priority=tmpl['priority']
                    )
                elif new_status == Booking.STATUS_APPROVED:
                    tmpl = NotificationTemplates.booking_approved(booking)
                    create_booking_notification(
                        booking=booking,
                        notification_type=Notification.TYPE_BOOKING_APPROVED,
                        title=tmpl['title'],
                        message=tmpl['message'],
                        priority=tmpl['priority']
                    )
                elif new_status == Booking.STATUS_DECLINED:
                    tmpl = NotificationTemplates.booking_declined(booking)
                    create_booking_notification(
                        booking=booking,
                        notification_type=Notification.TYPE_BOOKING_DECLINED,
                        title=tmpl['title'],
                        message=tmpl['message'],
                        priority=tmpl['priority']
                    )
                elif new_status == Booking.STATUS_CANCELED:
                    tmpl = NotificationTemplates.booking_canceled(booking)
                    create_booking_notification(
                        booking=booking,
                        notification_type=Notification.TYPE_BOOKING_CANCELED,
                        title=tmpl['title'],
                        message=tmpl['message'],
                        priority=tmpl['priority']
                    )
                elif new_status == Booking.STATUS_COMPLETED:
                    tmpl = NotificationTemplates.booking_completed(booking)
                    create_booking_notification(
                        booking=booking,
                        notification_type=Notification.TYPE_BOOKING_COMPLETED,
                        title=tmpl['title'],
                        message=tmpl['message'],
                        priority=tmpl['priority']
                    )
            except Exception:
                pass

            messages.success(request, 'Booking status updated successfully.')
            return HttpResponseRedirect(reverse('core:manage_booking', args=[booking.id]))
        else:
            messages.error(request, 'Invalid status update.')

    ctx = {
        'page_title': f'Appointment {booking.code}',
        'booking': booking,
        'decline_reasons': booking.church.decline_reasons.filter(is_active=True).order_by('order', 'id'),
    }
    ctx.update(_app_context(request))
    return render(request, 'core/manage_booking_detail.html', ctx)


@login_required
@require_http_methods(["GET", "POST"])
def create_booking_manually(request):
    """Allow parish staff to create bookings on behalf of users."""
    import json
    
    # Get church_id from query parameters
    church_id = request.GET.get('church_id')
    
    if not church_id:
        return JsonResponse({'success': False, 'message': 'Church ID is required.'}, status=400)
    
    try:
        church = Church.objects.get(id=church_id)
    except Church.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Church not found.'}, status=404)
    
    # Check permissions - must be owner or secretary with appointments permission
    has_permission, role = user_can_manage_church(request.user, church, required_permissions=['appointments'])
    if not has_permission:
        return JsonResponse({'success': False, 'message': 'You do not have permission to create bookings for this church.'}, status=403)
    
    if request.method == 'GET':
        # Return form data (services, etc.)
        try:
            services = BookableService.objects.filter(church=church, is_active=True).select_related('category')
            services_data = []
            
            for s in services:
                service_data = {
                    'id': s.id,
                    'name': s.name,
                    'category': s.category.name if s.category else 'Uncategorized',
                    'price': float(s.price) if s.price else 0.0,
                    'is_free': s.is_free,
                    'duration_minutes': s.duration if s.duration else 0,
                }
                services_data.append(service_data)
            
            return JsonResponse({'success': True, 'services': services_data})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': f'Error loading services: {str(e)}'}, status=500)
    
    elif request.method == 'POST':
        # Create booking
        try:
            data = json.loads(request.body)
            
            # Get required fields
            mode = data.get('mode', 'search')
            user_id = data.get('user_id')
            service_id = data.get('service_id')
            date = data.get('date')
            start_time = data.get('start_time')
            end_time = data.get('end_time')
            notes = data.get('notes', '')
            auto_approve = data.get('auto_approve', False)
            payment_status = data.get('payment_status', 'pending')
            payment_method = data.get('payment_method')
            
            # Handle user based on mode
            if mode == 'create':
                # Create new user with contact info
                contact_name = data.get('contact_name', '').strip()
                contact_email = data.get('contact_email', '').strip()
                contact_phone = data.get('contact_phone', '').strip()
                
                if not contact_name:
                    return JsonResponse({'success': False, 'message': 'Contact name is required.'}, status=400)
                
                if not contact_email and not contact_phone:
                    return JsonResponse({'success': False, 'message': 'At least email or phone number is required.'}, status=400)
                
                # Generate unique username
                from django.utils.text import slugify
                import random
                base_username = slugify(contact_name.lower().replace(' ', '_'))
                username = base_username
                counter = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}_{counter}"
                    counter += 1
                
                # Create user
                user = User.objects.create_user(
                    username=username,
                    email=contact_email if contact_email else f"{username}@temporary.local",
                    first_name=contact_name.split()[0] if contact_name else '',
                    last_name=' '.join(contact_name.split()[1:]) if len(contact_name.split()) > 1 else '',
                    password=User.objects.make_random_password(length=32)
                )
                
                # Update profile with phone if provided
                if contact_phone:
                    from accounts.models import Profile
                    Profile.objects.update_or_create(
                        user=user,
                        defaults={'phone': contact_phone}
                    )
            else:
                # Search mode - get existing user
                if not user_id:
                    return JsonResponse({'success': False, 'message': 'User ID is required in search mode.'}, status=400)
                
                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'User not found.'}, status=404)
            
            # Validate service and date
            if not service_id or not date:
                return JsonResponse({'success': False, 'message': 'Service and date are required.'}, status=400)
            
            # Get service
            try:
                service = BookableService.objects.get(id=service_id, church=church)
            except BookableService.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Service not found.'}, status=404)
            
            # Parse date and times
            from datetime import datetime
            from django.utils import timezone
            
            booking_date = datetime.strptime(date, '%Y-%m-%d').date()
            booking_start_time = datetime.strptime(start_time, '%H:%M').time() if start_time else None
            booking_end_time = datetime.strptime(end_time, '%H:%M').time() if end_time else None
            
            # Create the booking
            booking = Booking.objects.create(
                user=user,
                church=church,
                service=service,
                date=booking_date,
                start_time=booking_start_time,
                end_time=booking_end_time,
                notes=notes,
                status=Booking.STATUS_APPROVED if auto_approve else Booking.STATUS_REQUESTED,
                created_by=request.user,  # Track who created it
                handled_by=request.user if auto_approve else None,
                payment_status=payment_status,
                payment_method=payment_method if payment_method else None,
                payment_amount=service.price if payment_status == 'paid' else None,
                payment_date=timezone.now() if payment_status == 'paid' else None,
            )
            
            # TODO: Send notification to user (notifications system to be implemented)
            # Booking is created successfully without notification for now
            
            return JsonResponse({
                'success': True,
                'message': 'Booking created successfully.',
                'booking': {
                    'id': booking.id,
                    'code': booking.code,
                    'status': booking.get_status_display(),
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def search_users_api(request):
    """Search for users by name, email, phone, or username for booking creation."""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'users': []})
    
    # Search users by username, email, phone, first name, or last name
    from django.db.models import Q
    users = User.objects.filter(
        Q(username__icontains=query) |
        Q(email__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(profile__phone__icontains=query)
    ).select_related('profile').distinct()[:10]
    
    users_data = []
    for user in users:
        # Get display name
        display_name = user.username
        if user.get_full_name():
            display_name = user.get_full_name()
        elif hasattr(user, 'profile') and user.profile and user.profile.display_name:
            display_name = user.profile.display_name
        
        # Get avatar
        avatar = None
        if hasattr(user, 'profile') and user.profile and user.profile.profile_image:
            avatar = user.profile.profile_image.url
        
        # Get phone number
        phone = None
        if hasattr(user, 'profile') and user.profile and user.profile.phone:
            phone = user.profile.phone
        
        # Build contact info (show email or phone or both)
        contact_info = []
        if user.email:
            contact_info.append(user.email)
        if phone:
            contact_info.append(phone)
        
        users_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email if user.email else '',
            'phone': phone if phone else '',
            'contact_info': ' | '.join(contact_info) if contact_info else 'No contact info',
            'display_name': display_name,
            'avatar': avatar,
        })
    
    return JsonResponse({'users': users_data})


@login_required
def create_decline_reason(request):
    """Create a new decline reason for the owner's church."""
    # Get church_id from POST or JSON data
    is_ajax = request.headers.get('Content-Type') == 'application/json'
    
    if is_ajax:
        import json
        data = json.loads(request.body)
        church_id = data.get('church_id')
        # Convert to int if it's a string
        if church_id:
            try:
                church_id = int(church_id)
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'message': 'Invalid church ID.'}, status=400)
    else:
        church_id = request.POST.get('church_id')
        if church_id:
            try:
                church_id = int(church_id)
            except (ValueError, TypeError):
                messages.error(request, "Invalid church ID.")
                return redirect('core:select_church')
    
    try:
        if church_id:
            church = request.user.owned_churches.get(id=church_id)
        else:
            # Log warning when falling back to first church
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"create_decline_reason: No church_id provided for user {request.user.id}, falling back to first church")
            church = request.user.owned_churches.first()
            if not church:
                if is_ajax:
                    return JsonResponse({'success': False, 'message': 'You don\'t own any churches.'}, status=403)
                if request.user.is_superuser:
                    return redirect('core:super_admin_create_church')
                messages.info(request, "You don't own any churches yet. Please contact a Super Admin to create one and assign you as manager.")
                return redirect('core:home')
    except Church.DoesNotExist:
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'You don\'t have permission to manage this church.'}, status=403)
        messages.error(request, "You don't have permission to manage this church.")
        return redirect('core:select_church')

    if request.method != 'POST':
        return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')
    
    if is_ajax:
        import json
        data = json.loads(request.body)
        label = (data.get('label') or '').strip()
        is_active = bool(data.get('is_active'))
    else:
        label = (request.POST.get('label') or '').strip()
        is_active = bool(request.POST.get('is_active'))
    
    if not label:
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'Reason label is required.'}, status=400)
        messages.error(request, 'Reason label is required.')
        return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')

    # Determine next order
    next_order = (church.decline_reasons.aggregate(c=Count('id'))['c'] or 0)
    try:
        reason = DeclineReason.objects.create(church=church, label=label[:200], is_active=is_active, order=next_order)
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': 'Decline reason added.',
                'reason': {
                    'id': reason.id,
                    'label': reason.label,
                    'is_active': reason.is_active
                }
            })
        messages.success(request, 'Decline reason added.')
    except IntegrityError:
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'A reason with this label already exists.'}, status=400)
        messages.info(request, 'A reason with this label already exists.')
    return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')


@login_required
def delete_decline_reason(request, reason_id):
    """Delete a decline reason belonging to the owner's church."""
    # Get the reason first to determine its church
    reason = get_object_or_404(DeclineReason, id=reason_id)
    church = reason.church
    
    # Only owner can manage decline reasons
    can_manage, role = user_can_manage_church(request.user, church)
    if not can_manage or role != 'owner':
        if request.headers.get('Content-Type') == 'application/json':
            return JsonResponse({'success': False, 'message': 'Only the parish manager can manage decline reasons.'}, status=403)
        messages.error(request, "Only the parish manager can manage decline reasons.")
        return redirect('core:select_church')
    if request.method == 'POST':
        reason.delete()
        if request.headers.get('Content-Type') == 'application/json':
            return JsonResponse({'success': True, 'message': 'Decline reason removed.'})
        messages.success(request, 'Decline reason removed.')
    return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')


@login_required
def toggle_decline_reason(request, reason_id):
    """Toggle active state of a decline reason belonging to the owner's church."""
    # Get the reason first to determine its church
    reason = get_object_or_404(DeclineReason, id=reason_id)
    church = reason.church
    
    # Only owner can manage decline reasons
    can_manage, role = user_can_manage_church(request.user, church)
    if not can_manage or role != 'owner':
        if request.headers.get('Content-Type') == 'application/json':
            return JsonResponse({'success': False, 'message': 'Only the parish manager can manage decline reasons.'}, status=403)
        messages.error(request, "Only the parish manager can manage decline reasons.")
        return redirect('core:select_church')
    if request.method == 'POST':
        # Check if JSON request to parse body
        if request.headers.get('Content-Type') == 'application/json':
            import json
            data = json.loads(request.body)
            reason.is_active = bool(data.get('is_active'))
        else:
            reason.is_active = not reason.is_active
        reason.save(update_fields=['is_active', 'updated_at'])
        
        if request.headers.get('Content-Type') == 'application/json':
            return JsonResponse({
                'success': True,
                'message': f'Decline reason set to {"active" if reason.is_active else "inactive"}.',
                'is_active': reason.is_active
            })
        messages.success(request, f'Decline reason "{reason.label}" set to {"active" if reason.is_active else "inactive"}.')
    return HttpResponseRedirect(reverse('core:manage_church') + '?tab=settings')


@login_required
def booking_invoice(request, booking_id):
    """Printable invoice page for a booking. Accessible by the booking user or the church owner."""
    booking = get_object_or_404(Booking.objects.select_related('service', 'service__category', 'church', 'user'), id=booking_id)
    is_owner = request.user == booking.user
    is_church_owner = request.user == booking.church.owner
    if not (is_owner or is_church_owner):
        messages.error(request, 'You do not have permission to view this invoice.')
        return redirect('core:appointments')

    # Do not allow invoice viewing for canceled/declined bookings
    if booking.status in (Booking.STATUS_CANCELED, Booking.STATUS_DECLINED):
        messages.error(request, 'Invoice is not available for canceled or declined bookings.')
        return redirect('core:appointments' if is_owner else 'core:manage_church')

    can_print = booking.status == Booking.STATUS_REVIEWED or is_church_owner
    ctx = {
        'active': 'appointments' if is_owner else 'manage',
        'page_title': f'Invoice {booking.code}',
        'booking': booking,
        'can_print': can_print,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/booking_invoice.html', ctx)


# Payment Processing Views
@login_required
def gcash_payment(request, booking_id):
    """GCash payment processing page"""
    booking = get_object_or_404(Booking.objects.select_related('service', 'service__category', 'church', 'user'), id=booking_id)
    
    # Check if user owns this booking
    if request.user != booking.user:
        messages.error(request, 'You do not have permission to pay for this booking.')
        return redirect('core:appointments')
    
    # Check if booking is in requested status
    if booking.status != Booking.STATUS_REQUESTED:
        messages.info(request, 'This booking has already been processed.')
        return redirect('core:appointments')
    
    ctx = {
        'active': 'appointments',
        'page_title': 'GCash Payment',
        'booking': booking,
        'payment_method': 'GCash',
    }
    ctx.update(_app_context(request))
    return render(request, 'core/payment_processing.html', ctx)


@login_required
def paypal_payment(request, booking_id):
    """PayPal payment processing page"""
    booking = get_object_or_404(Booking.objects.select_related('service', 'service__category', 'church', 'user'), id=booking_id)
    
    # Check if user owns this booking
    if request.user != booking.user:
        messages.error(request, 'You do not have permission to pay for this booking.')
        return redirect('core:appointments')
    
    # Check if booking is in requested status
    if booking.status != Booking.STATUS_REQUESTED:
        messages.info(request, 'This booking has already been processed.')
        return redirect('core:appointments')
    
    ctx = {
        'active': 'appointments',
        'page_title': 'PayPal Payment',
        'booking': booking,
        'payment_method': 'PayPal',
    }
    ctx.update(_app_context(request))
    return render(request, 'core/payment_processing.html', ctx)


# Church Verification Flow
@login_required
def request_verification(request):
    """Owner submits legal documents for church verification (requires at least 2)."""
    # Get church_id from POST parameters
    church_id = request.POST.get('church_id')
    
    try:
        if church_id:
            church = request.user.owned_churches.get(id=church_id)
        else:
            church = request.user.owned_churches.first()
            if not church:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': "You don't own any churches yet."}, status=403)
                if request.user.is_superuser:
                    return redirect('core:super_admin_create_church')
                messages.info(request, "You don't own any churches yet. Please contact a Super Admin to create one and assign you as manager.")
                return redirect('core:home')
    except Church.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': "You don't have permission to manage this church."}, status=403)
        messages.error(request, "You don't have permission to manage this church.")
        return redirect('core:select_church')

    if request.method != 'POST':
        # Redirect to settings tab where the form lives
        return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')

    # Disallow multiple pending requests
    existing_pending = church.verification_requests.filter(status=ChurchVerificationRequest.STATUS_PENDING).first()
    if existing_pending:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'You already have a pending verification request.'}, status=400)
        messages.info(request, 'You already have a pending verification request. Our team will review it soon.')
        return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')

    form = ChurchVerificationUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Flatten form errors
            errors = {field: [str(e) for e in errs] for field, errs in form.errors.items()}
            return JsonResponse({'success': False, 'message': 'Invalid submission', 'errors': errors}, status=400)
        for field, errs in form.errors.items():
            for err in errs:
                messages.error(request, f"{field}: {err}")
        return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')

    # Create request and documents
    ver_req = ChurchVerificationRequest.objects.create(
        church=church,
        submitted_by=request.user,
        status=ChurchVerificationRequest.STATUS_PENDING,
    )
    for f in form.cleaned_data['documents']:
        ChurchVerificationDocument.objects.create(
            request=ver_req,
            file=f,
            title=getattr(f, 'name', '')[:200],
        )

    # AJAX success response
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': 'Verification request submitted', 'request_id': ver_req.id})

    messages.success(request, 'Your verification request has been submitted. We will notify you once it is reviewed.')
    return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=settings')


@login_required
def super_admin_church_detail(request, church_id):
    """Detailed view of a church for super admins."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    from datetime import timedelta
    from django.db.models import Count, Q, Avg
    import json
    
    church = get_object_or_404(Church.objects.select_related('owner'), id=church_id, is_active=True)
    
    # Get church services
    services = church.bookable_services.filter(is_active=True).prefetch_related('service_images').order_by('name')
    
    # Get recent posts
    posts = church.posts.filter(is_active=True).select_related('church').order_by('-created_at')[:10]
    
    # Get followers count
    followers_count = ChurchFollow.objects.filter(church=church).count()
    
    # Get parish administrators (staff members)
    staff_members = ChurchStaff.objects.filter(
        church=church
    ).select_related('user', 'user__profile', 'added_by').order_by('role', '-added_at')
    
    # Get verification requests history with documents
    verification_requests = ChurchVerificationRequest.objects.filter(
        church=church
    ).select_related('submitted_by', 'reviewed_by').prefetch_related('documents').order_by('-created_at')
    
    # === CHART DATA CALCULATIONS ===
    
    # 1. Booking Trends (Last 30 days)
    booking_trends_labels = []
    booking_trends_data = []
    for i in range(29, -1, -1):
        day = timezone.now() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        count = Booking.objects.filter(
            church=church,
            created_at__gte=day_start,
            created_at__lt=day_end
        ).count()
        
        booking_trends_labels.append(day.strftime('%b %d'))
        booking_trends_data.append(count)
    
    booking_trends = json.dumps({
        'labels': booking_trends_labels,
        'data': booking_trends_data
    })
    
    # 2. Most Popular Services (Top 5)
    popular_services_data = Booking.objects.filter(
        church=church
    ).values('service__name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    popular_services = json.dumps({
        'labels': [item['service__name'] for item in popular_services_data],
        'data': [item['count'] for item in popular_services_data]
    })
    
    # 3. Service Bookings Trend by Category (Last 30 days)
    from core.models import ServiceCategory
    categories = ServiceCategory.objects.filter(is_active=True)[:3]  # Top 3 categories
    
    service_trend_labels = []
    service_trend_datasets = []
    
    # Generate labels
    for i in range(29, -1, -1):
        day = timezone.now() - timedelta(days=i)
        service_trend_labels.append(day.strftime('%b %d'))
    
    # Generate datasets for each category
    colors = [
        'rgba(30, 144, 255, 1)',
        'rgba(65, 105, 225, 1)',
        'rgba(100, 149, 237, 1)'
    ]
    
    for idx, category in enumerate(categories):
        category_data = []
        for i in range(29, -1, -1):
            day = timezone.now() - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            
            count = Booking.objects.filter(
                church=church,
                service__category=category,
                created_at__gte=day_start,
                created_at__lt=day_end
            ).count()
            
            category_data.append(count)
        
        service_trend_datasets.append({
            'label': category.name,
            'data': category_data,
            'borderColor': colors[idx % len(colors)],
            'backgroundColor': colors[idx % len(colors)].replace('1)', '0.1)'),
            'borderWidth': 2,
            'fill': False,
            'tension': 0.4
        })
    
    service_bookings_trend = json.dumps({
        'labels': service_trend_labels,
        'datasets': service_trend_datasets
    })
    
    # 4. Service Performance (Completion Rate)
    total_bookings = Booking.objects.filter(church=church).count()
    completed_bookings = Booking.objects.filter(church=church, status=Booking.STATUS_COMPLETED).count()
    pending_bookings = Booking.objects.filter(church=church, status=Booking.STATUS_REQUESTED).count()
    cancelled_bookings = Booking.objects.filter(
        church=church
    ).filter(
        Q(status=Booking.STATUS_CANCELED) | Q(status=Booking.STATUS_DECLINED)
    ).count()
    
    service_performance = json.dumps({
        'labels': ['Completed', 'Pending', 'Cancelled'],
        'data': [completed_bookings, pending_bookings, cancelled_bookings]
    })
    
    # 5. Engagement Trends (Last 30 days)
    engagement_labels = []
    engagement_likes = []
    engagement_comments = []
    engagement_bookmarks = []
    
    for i in range(29, -1, -1):
        day = timezone.now() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        likes = PostLike.objects.filter(
            post__church=church,
            created_at__gte=day_start,
            created_at__lt=day_end
        ).count()
        
        comments = PostComment.objects.filter(
            post__church=church,
            created_at__gte=day_start,
            created_at__lt=day_end
        ).count()
        
        bookmarks = PostBookmark.objects.filter(
            post__church=church,
            created_at__gte=day_start,
            created_at__lt=day_end
        ).count()
        
        engagement_labels.append(day.strftime('%b %d'))
        engagement_likes.append(likes)
        engagement_comments.append(comments)
        engagement_bookmarks.append(bookmarks)
    
    engagement_trends = json.dumps({
        'labels': engagement_labels,
        'likes': engagement_likes,
        'comments': engagement_comments,
        'bookmarks': engagement_bookmarks
    })
    
    ctx = {
        'active': 'super_admin_verifications',
        'page_title': f'{church.name} - Church Details',
        'church': church,
        'services': services,
        'posts': posts,
        'followers_count': followers_count,
        'staff_members': staff_members,
        'verification_requests': verification_requests,
        # Chart data
        'booking_trends': booking_trends,
        'popular_services': popular_services,
        'service_bookings_trend': service_bookings_trend,
        'service_performance': service_performance,
        'engagement_trends': engagement_trends,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_church_detail.html', ctx)


@login_required
def super_admin_verifications(request):
    """List and filter church verification requests for super-admins."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    from django.utils import timezone
    from datetime import timedelta

    # Default to showing all statuses when no filter is applied
    status = request.GET.get('status', 'all')
    qs = ChurchVerificationRequest.objects.select_related('church', 'submitted_by', 'reviewed_by')
    if status and status != 'all':
        qs = qs.filter(status=status)
    qs = qs.order_by('-created_at')

    # Calculate church statistics for sidebar
    total_churches = Church.objects.count()
    verified_churches = Church.objects.filter(is_verified=True).count()
    pending_verifications = ChurchVerificationRequest.objects.filter(
        status=ChurchVerificationRequest.STATUS_PENDING
    ).count()

    # Church creation over time (last 30 days)
    church_creation_data = []
    church_creation_labels = []
    for i in range(29, -1, -1):  # Last 30 days
        day = timezone.now() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        count = Church.objects.filter(
            created_at__gte=day_start,
            created_at__lt=day_end
        ).count()
        
        church_creation_data.append(count)
        church_creation_labels.append(day.strftime('%b %d'))

    # Denomination distribution (all time) — include ALL choices, even if count is zero
    denom_counts = {
        row['denomination']: row['count']
        for row in Church.objects.values('denomination').annotate(count=Count('id'))
    }
    # Preserve the order defined in Church.DENOMINATION_CHOICES
    church_denom_labels = [label for key, label in Church.DENOMINATION_CHOICES]
    church_denom_data = [denom_counts.get(key, 0) for key, _ in Church.DENOMINATION_CHOICES]

    # Iligan-specific faith categories for clearer grouping
    catholic_codes = {'catholic'}
    protestant_evangelical_codes = {
        'protestant', 'uccp', 'baptist', 'methodist', 'presbyterian',
        'lutheran', 'pentecostal', 'evangelical'
    }
    other_christian_codes = {'iglesia_ni_cristo', 'adventist', 'mormon', 'jehovah', 'orthodox'}
    islam_codes = {'islam'}
    indigenous_codes = {'indigenous'}
    other_faith_codes = {'buddhism', 'hinduism'}

    group_def = [
        ('Roman Catholic', catholic_codes),
        ('Protestant & Evangelical', protestant_evangelical_codes),
        ('Other Christian', other_christian_codes),
        ('Islam', islam_codes),
        ('Indigenous Beliefs', indigenous_codes),
        ('Other Faiths', other_faith_codes),
    ]

    # Any remaining codes not explicitly mapped go to "Other Faiths"
    mapped_codes = set().union(*(codes for _, codes in group_def))
    remaining = set(denom_counts.keys()) - mapped_codes
    if remaining:
        # Accumulate their counts into Other Faiths
        extra_other = sum(denom_counts.get(code, 0) for code in remaining)
        if extra_other:
            # If Other Faiths already exists, we'll add to it later
            pass

    church_denom_group_labels = []
    church_denom_group_data = []
    other_index = None
    for label, codes in group_def:
        count = sum(denom_counts.get(code, 0) for code in codes)
        if label == 'Other Faiths':
            other_index = len(church_denom_group_labels)
        if count > 0:
            church_denom_group_labels.append(label)
            church_denom_group_data.append(count)

    # Add remaining unmapped counts into Other Faiths if present, else create it
    if remaining:
        extra_other = sum(denom_counts.get(code, 0) for code in remaining)
        if extra_other:
            if other_index is not None and other_index < len(church_denom_group_data):
                church_denom_group_data[other_index] += extra_other
            else:
                church_denom_group_labels.append('Other Faiths')
                church_denom_group_data.append(extra_other)

    # Calculate additional insights
    # Verification rate
    verification_rate = round((verified_churches / total_churches * 100) if total_churches > 0 else 0, 1)
    
    # Average churches created per day (last 30 days)
    total_created_30_days = sum(church_creation_data)
    avg_daily_growth = round(total_created_30_days / 30, 1)
    
    # Most popular category
    most_popular_category = 'N/A'
    if church_denom_group_data:
        max_count = max(church_denom_group_data)
        max_index = church_denom_group_data.index(max_count)
        most_popular_category = church_denom_group_labels[max_index]

    # Fetch churches for management table with filters
    search_query = request.GET.get('search', '')
    verified_filter = request.GET.get('verified', '')
    denom_filter = request.GET.get('denomination', '')
    
    churches_qs = Church.objects.select_related('owner').annotate(
        follower_total=Count('followers', distinct=True)
    )
    
    if search_query:
        churches_qs = churches_qs.filter(
            Q(name__icontains=search_query) | 
            Q(city__icontains=search_query) |
            Q(owner__email__icontains=search_query)
        )
    
    if verified_filter == 'verified':
        churches_qs = churches_qs.filter(is_verified=True)
    elif verified_filter == 'unverified':
        churches_qs = churches_qs.filter(is_verified=False)
    
    if denom_filter:
        churches_qs = churches_qs.filter(denomination=denom_filter)
    
    churches_qs = churches_qs.order_by('-created_at')
    
    # Paginate churches (5 per page)
    paginator = Paginator(churches_qs, 5)
    page_number = request.GET.get('page', 1)
    churches_page = paginator.get_page(page_number)

    ctx = {
        'active': 'super_admin_verifications',
        'page_title': 'Church Management',
        'requests': qs,
        'status_filter': status,
        'total_churches': total_churches,
        'verified_churches': verified_churches,
        'pending_verifications': pending_verifications,
        'church_creation_data': church_creation_data,
        'church_creation_labels': church_creation_labels,
        'church_denom_labels': church_denom_labels,
        'church_denom_data': church_denom_data,
        'church_denom_group_labels': church_denom_group_labels,
        'church_denom_group_data': church_denom_group_data,
        'verification_rate': verification_rate,
        'avg_daily_growth': avg_daily_growth,
        'most_popular_category': most_popular_category,
        'churches_page': churches_page,
        'search_query': search_query,
        'verified_filter': verified_filter,
        'denom_filter': denom_filter,
        'denomination_choices': Church.DENOMINATION_CHOICES,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_verifications.html', ctx)


@login_required
def approve_verification(request, request_id):
    """Approve a verification request and mark the church as verified."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    ver_req = get_object_or_404(ChurchVerificationRequest, id=request_id)
    if request.method != 'POST':
        return redirect('core:super_admin_verifications')
    if ver_req.status != ChurchVerificationRequest.STATUS_PENDING:
        messages.info(request, 'This request has already been processed.')
        return redirect('core:super_admin_verifications')

    ver_req.status = ChurchVerificationRequest.STATUS_APPROVED
    ver_req.reviewed_by = request.user
    ver_req.notes = request.POST.get('notes', '')
    ver_req.save(update_fields=['status', 'reviewed_by', 'notes', 'updated_at'])

    # Mark church as verified
    church = ver_req.church
    if not church.is_verified:
        church.is_verified = True
        church.save(update_fields=['is_verified', 'updated_at'])
    messages.success(request, f'Verification for "{church.name}" approved.')
    return redirect('core:super_admin_verifications')


@login_required
def reject_verification(request, request_id):
    """Reject a verification request and keep the church unverified."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    ver_req = get_object_or_404(ChurchVerificationRequest, id=request_id)
    if request.method != 'POST':
        return redirect('core:super_admin_verifications')
    if ver_req.status != ChurchVerificationRequest.STATUS_PENDING:
        messages.info(request, 'This request has already been processed.')
        return redirect('core:super_admin_verifications')

    ver_req.status = ChurchVerificationRequest.STATUS_REJECTED
    ver_req.reviewed_by = request.user
    ver_req.notes = request.POST.get('notes', '') or 'Rejected'
    ver_req.save(update_fields=['status', 'reviewed_by', 'notes', 'updated_at'])

    # Ensure church remains unverified
    church = ver_req.church
    if church.is_verified:
        church.is_verified = False
        church.save(update_fields=['is_verified', 'updated_at'])
    messages.warning(request, f'Verification for "{church.name}" rejected.')
    return redirect('core:super_admin_verifications')

@login_required
def create_availability(request):
    """Create a new availability entry."""
    # Get church_id from GET or POST parameters
    church_id = request.GET.get('church_id') or request.POST.get('church_id')
    
    try:
        if church_id:
            church = Church.objects.get(id=church_id)
        else:
            church = request.user.owned_churches.first()
            if not church:
                # Check if user is a secretary of any church
                secretary_church = ChurchStaff.objects.filter(
                    user=request.user,
                    role__in=['secretary', 'volunteer'],
                    status='active'
                ).select_related('church').first()
                
                if secretary_church:
                    church = secretary_church.church
                elif request.user.is_superuser:
                    return redirect('core:super_admin_create_church')
                else:
                    messages.info(request, "You don't own any churches yet. Please contact a Super Admin to create one and assign you as manager.")
                    return redirect('core:home')
        
        # Check if user can manage availability (Owner or Secretary with availability permission)
        can_manage, role = user_can_manage_church(request.user, church, ['availability'])
        if not can_manage:
            messages.error(request, "You don't have permission to manage this church.")
            return redirect('core:select_church')
            
    except Church.DoesNotExist:
        messages.error(request, "You don't have permission to manage this church.")
        return redirect('core:select_church')
    
    if request.method == 'POST':
        form = AvailabilityForm(request.POST, church=church)
        if form.is_valid():
            try:
                availability = form.save()
                
                # Log staff activity for availability creation
                from .models import StaffActivityLog
                status_desc = "closed" if availability.is_closed else "special hours"
                log_staff_activity(
                    user=request.user,
                    church=church,
                    action=StaffActivityLog.ACTION_CREATE,
                    category=StaffActivityLog.CATEGORY_AVAILABILITY,
                    description=f"Created {status_desc} entry for {availability.date}",
                    target_id=availability.id,
                    target_type='availability',
                    request=request
                )
                
                messages.success(request, f'Availability entry for {availability.date} has been created successfully!')
                return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=availability')
            except IntegrityError:
                # Handle duplicate entry - redirect to edit existing entry
                existing_availability = Availability.objects.get(church=church, date=form.cleaned_data['date'])
                messages.info(request, f'An availability entry for {form.cleaned_data["date"]} already exists. Redirecting to edit...')
                return HttpResponseRedirect(reverse('core:edit_availability', args=[existing_availability.id]))
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Check if date parameter is provided in the URL
        initial_data = {}
        date_param = request.GET.get('date')
        if date_param and date_param != '0':
            try:
                # Validate the date format
                from datetime import datetime, date
                parsed_date = datetime.strptime(date_param, '%Y-%m-%d').date()
                
                # Check if the date is in the past
                if parsed_date < date.today():
                    messages.error(request, 'Cannot create availability entries for past dates.')
                    return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=availability')
                
                initial_data['date'] = parsed_date
            except ValueError:
                pass  # Invalid date format, ignore
        
        form = AvailabilityForm(church=church, initial=initial_data)
    
    ctx = {
        'active': 'manage',
        'page_title': 'Create Availability Entry',
        'church': church,
        'form': form,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/create_availability.html', ctx)


@login_required
def edit_availability(request, availability_id):
    """Edit an availability entry."""
    # Get the availability first to determine its church
    availability = get_object_or_404(Availability, id=availability_id)
    church = availability.church
    
    # Check if user can manage availability (Owner or Secretary)
    can_manage, role = user_can_manage_church(request.user, church, ['availability'])
    if not can_manage:
        messages.error(request, "You don't have permission to edit availability.")
        return redirect('core:select_church')
    
    if request.method == 'POST':
        form = AvailabilityForm(request.POST, instance=availability, church=church)
        if form.is_valid():
            form.save()
            
            # Log staff activity for availability update
            from .models import StaffActivityLog
            status_desc = "closed" if availability.is_closed else "special hours"
            log_staff_activity(
                user=request.user,
                church=church,
                action=StaffActivityLog.ACTION_UPDATE,
                category=StaffActivityLog.CATEGORY_AVAILABILITY,
                description=f"Updated {status_desc} entry for {availability.date}",
                target_id=availability.id,
                target_type='availability',
                request=request
            )
            
            messages.success(request, f'Availability entry for {availability.date} has been updated successfully!')
            return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=availability')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AvailabilityForm(instance=availability, church=church)
    
    ctx = {
        'active': 'manage',
        'page_title': f'Edit Availability for {availability.date}',
        'church': church,
        'availability': availability,
        'form': form,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/edit_availability.html', ctx)


@login_required
def delete_availability(request, availability_id):
    """Delete an availability entry."""
    # Get the availability first to determine its church
    availability = get_object_or_404(Availability, id=availability_id)
    church = availability.church
    
    # Check if user can manage availability (Owner or Secretary)
    can_manage, role = user_can_manage_church(request.user, church, ['availability'])
    if not can_manage:
        messages.error(request, "You don't have permission to delete availability.")
        return redirect('core:select_church')
    
    if request.method == 'POST':
        availability_date = availability.date
        availability.delete()
        messages.success(request, f'Availability entry for {availability_date} has been deleted successfully!')
        return HttpResponseRedirect(reverse('core:manage_church', kwargs={'church_id': church.id}) + '?tab=availability')
    
    ctx = {
        'active': 'manage',
        'page_title': f'Delete Availability for {availability.date}',
        'church': church,
        'availability': availability,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/delete_availability.html', ctx)


@login_required
def bulk_availability(request):
    """Handle bulk availability management."""
    # Get church_id from POST parameters
    church_id = request.POST.get('church_id')
    
    try:
        if church_id:
            church = Church.objects.get(id=church_id)
        else:
            church = request.user.owned_churches.first()
            if not church:
                # Check if user is a secretary of any church
                secretary_church = ChurchStaff.objects.filter(
                    user=request.user,
                    role__in=['secretary', 'volunteer'],
                    status='active'
                ).select_related('church').first()
                
                if secretary_church:
                    church = secretary_church.church
                elif request.user.is_superuser:
                    return redirect('core:super_admin_create_church')
                else:
                    messages.info(request, "You don't own any churches yet. Please contact a Super Admin to create one and assign you as manager.")
                    return redirect('core:home')
        
        # Check if user can manage availability (Owner or Secretary with availability permission)
        can_manage, role = user_can_manage_church(request.user, church, ['availability'])
        if not can_manage:
            messages.error(request, "You don't have permission to manage this church.")
            return redirect('core:select_church')
            
    except Church.DoesNotExist:
        messages.error(request, "You don't have permission to manage this church.")
        return redirect('core:select_church')
    
    if request.method == 'POST':
        form = AvailabilityBulkForm(request.POST)
        if form.is_valid():
            dates_str = form.cleaned_data['dates']
            action = form.cleaned_data['action']
            start_time = form.cleaned_data.get('start_time')
            end_time = form.cleaned_data.get('end_time')
            reason = form.cleaned_data.get('reason')
            notes = form.cleaned_data.get('notes')
            
            # Parse dates
            from datetime import datetime
            dates = [datetime.strptime(d.strip(), '%Y-%m-%d').date() for d in dates_str.split(',') if d.strip()]
            
            created_count = 0
            for date in dates:
                availability, created = Availability.objects.get_or_create(
                    church=church,
                    date=date,
                    defaults={
                        'type': 'closed_date' if action == 'close' else 'special_hours',
                        'is_closed': action == 'close',
                        'start_time': start_time if action == 'special_hours' else None,
                        'end_time': end_time if action == 'special_hours' else None,
                        'reason': reason,
                        'notes': notes,
                    }
                )
                if created:
                    created_count += 1
                else:
                    # Update existing entry
                    availability.type = 'closed_date' if action == 'close' else 'special_hours'
                    availability.is_closed = action == 'close'
                    if action == 'special_hours':
                        availability.start_time = start_time
                        availability.end_time = end_time
                    else:
                        availability.start_time = None
                        availability.end_time = None
                    availability.reason = reason
                    availability.notes = notes
                    availability.save()
            
            messages.success(request, f'Successfully updated availability for {created_count} dates!')
            return HttpResponseRedirect(reverse('core:manage_church') + '?tab=availability')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AvailabilityBulkForm()
    
    ctx = {
        'active': 'manage',
        'page_title': 'Bulk Availability Management',
        'church': church,
        'form': form,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/bulk_availability.html', ctx)


@login_required
def create_post(request, church_slug):
    """Create a new post for a church."""
    church = get_object_or_404(Church, slug=church_slug, is_active=True)
    
    # Check if user can manage content (Owner or Ministry Leader)
    can_manage, role = user_can_manage_church(request.user, church, ['content'])
    if not can_manage:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'You do not have permission to create posts.'}, status=403)
        messages.error(request, 'You do not have permission to create posts.')
        return redirect('core:church_detail', slug=church.slug)
    
    if request.method == 'POST':
        form = PostForm(request.POST, request.FILES, church=church)
        if form.is_valid():
            try:
                post = form.save(commit=False)
                post.church = church
                post.save()
                
                # Handle multiple images
                images = request.FILES.getlist('images')
                if images:
                    from .models import PostImage
                    for index, image_file in enumerate(images):
                        # Validate each image
                        if image_file.size > 10 * 1024 * 1024:  # 10MB limit per image
                            continue  # Skip large images
                        
                        # Create PostImage record
                        PostImage.objects.create(
                            post=post,
                            image=image_file,
                            order=index
                        )
                
                # Log activity if staff member
                log_staff_activity(
                    user=request.user,
                    church=church,
                    action='create',
                    category='post',
                    description=f'Created {post.get_post_type_display()}: {post.content[:50]}...',
                    target_id=post.id,
                    target_type='post',
                    request=request
                )
                
                # Return JSON for AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Post created successfully!',
                        'post': {
                            'id': post.id,
                            'content': post.content,
                            'post_type': post.post_type,
                            'image_url': post.image.url if post.image else None,
                            'images_count': post.images.count(),
                            'created_at': post.created_at.strftime('%B %d, %Y at %I:%M %p')
                        }
                    })
                
                messages.success(request, 'Post created successfully!')
                return redirect('core:church_detail', slug=church.slug)
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'}, status=500)
                messages.error(request, f'An error occurred: {str(e)}')
        else:
            # Return form errors for AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = []
                for field, field_errors in form.errors.items():
                    for error in field_errors:
                        errors.append(f"{field}: {error}")
                return JsonResponse({'success': False, 'message': ' '.join(errors)}, status=400)
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PostForm(church=church)
    
    ctx = {
        'page_title': f'Create Post - {church.name}',
        'church': church,
        'form': form,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/create_post.html', ctx)


@login_required
def edit_post(request, post_id):
    """Edit an existing post."""
    post = get_object_or_404(Post, id=post_id, is_active=True)
    
    # Check if user can manage content (Owner or Ministry Leader)
    can_manage, role = user_can_manage_church(request.user, post.church, ['content'])
    if not can_manage:
        return JsonResponse({'success': False, 'message': 'You do not have permission to edit this post.'}, status=403)
    
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        remove_image = request.POST.get('remove_image') == 'true'
        
        if not content:
            return JsonResponse({'success': False, 'message': 'Post content cannot be empty.'}, status=400)
        
        if len(content) > 1000:
            return JsonResponse({'success': False, 'message': 'Post content is too long (max 1000 characters).'}, status=400)
        
        try:
            # Update content
            post.content = content
            
            # Handle image updates
            if remove_image and post.image:
                # Remove current image
                post.image.delete()
                post.image = None
            
            # Handle new image upload
            if 'image' in request.FILES:
                new_image = request.FILES['image']
                # Validate image
                if new_image.size > 10 * 1024 * 1024:  # 10MB limit
                    return JsonResponse({'success': False, 'message': 'Image size must be less than 10MB.'}, status=400)
                
                # Remove old image if exists
                if post.image:
                    post.image.delete()
                
                post.image = new_image
            
            post.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Post updated successfully!',
                'post': {
                    'id': post.id,
                    'content': post.content,
                    'image_url': post.image.url if post.image else None,
                    'updated_at': post.updated_at.strftime('%B %d, %Y at %I:%M %p')
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': 'An error occurred while updating the post.'}, status=500)
    
    elif request.method == 'GET':
        # Return post data for editing
        return JsonResponse({
            'success': True,
            'post': {
                'id': post.id,
                'content': post.content,
                'image_url': post.image.url if post.image else None,
                'created_at': post.created_at.strftime('%B %d, %Y at %I:%M %p')
            }
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)


# Notification Views
@login_required
def notifications(request):
    """Display user's notifications page."""
    
    # Get filter parameters (category-based)
    filter_type = request.GET.get('type', 'all')
    # Backward compatibility: also respect unread=true query param
    unread_only = request.GET.get('unread', 'false').lower() == 'true' or filter_type == 'unread'

    # Category groupings for common filters
    category_map = {
        'bookings': [
            Notification.TYPE_BOOKING_REQUESTED,
            Notification.TYPE_BOOKING_REVIEWED,
            Notification.TYPE_BOOKING_APPROVED,
            Notification.TYPE_BOOKING_DECLINED,
            Notification.TYPE_BOOKING_CANCELED,
            Notification.TYPE_BOOKING_COMPLETED,
        ],
        'church': [
            Notification.TYPE_CHURCH_APPROVED,
            Notification.TYPE_CHURCH_DECLINED,
        ],
        'follows': [
            Notification.TYPE_FOLLOW_REQUEST,
            Notification.TYPE_FOLLOW_ACCEPTED,
        ],
    }

    # Base queryset
    base_queryset = Notification.objects.filter(user=request.user).select_related('booking', 'church')

    # Apply filters
    if unread_only:
        base_queryset = base_queryset.filter(is_read=False)

    if filter_type in category_map:
        base_queryset = base_queryset.filter(notification_type__in=category_map[filter_type])

    # Get notifications with limit
    notifications_qs = base_queryset[:50]

    # Get counts for categories (fresh queryset)
    all_notifications = Notification.objects.filter(user=request.user)
    counts = {
        'all': all_notifications.count(),
        'unread': all_notifications.filter(is_read=False).count(),
        'bookings': all_notifications.filter(notification_type__in=category_map['bookings']).count(),
        'church': all_notifications.filter(notification_type__in=category_map['church']).count(),
        'follows': all_notifications.filter(notification_type__in=category_map['follows']).count(),
    }

    ctx = {
        'active': 'notifications',
        'page_title': 'Notifications',
        'notifications': notifications_qs,
        'counts': counts,
        'filter_type': filter_type,
        'unread_only': unread_only,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/notifications.html', ctx)


@login_required
def notification_dropdown(request):
    """AJAX endpoint for notification dropdown content."""
    from .notifications import get_user_notifications, get_user_unread_count
    
    # Get recent notifications (both read and unread) for display
    all_notifications = get_user_notifications(request.user, limit=20, unread_only=False)
    
    # Separate message notifications from regular notifications
    message_notifications = [n for n in all_notifications if n.notification_type == Notification.TYPE_MESSAGE_RECEIVED]
    regular_notifications = [n for n in all_notifications if n.notification_type != Notification.TYPE_MESSAGE_RECEIVED]
    
    unread_count = get_user_unread_count(request.user)
    
    # Compute a target URL for each notification
    def _target_url(n):
        try:
            # Message notifications - go to church page to open chat
            if n.notification_type == Notification.TYPE_MESSAGE_RECEIVED:
                if n.church and getattr(n.church, 'slug', None):
                    return reverse('core:church_detail', args=[n.church.slug])
            # Booking-related
            elif n.notification_type == Notification.TYPE_BOOKING_REQUESTED:
                # If current user is the church owner -> manage appointments tab
                if n.church and n.church.owner_id == request.user.id:
                    return reverse('core:manage_church') + '?tab=appointments'
                # Fallback for any other case
                return reverse('core:appointments')
            elif n.notification_type in (
                Notification.TYPE_BOOKING_REVIEWED,
                Notification.TYPE_BOOKING_APPROVED,
                Notification.TYPE_BOOKING_DECLINED,
                Notification.TYPE_BOOKING_CANCELED,
                Notification.TYPE_BOOKING_COMPLETED,
            ):
                # Requester views their appointments
                return reverse('core:appointments')
            # Church verification/admin notifications
            elif n.notification_type in (
                Notification.TYPE_CHURCH_APPROVED,
                Notification.TYPE_CHURCH_DECLINED,
            ):
                return reverse('core:manage_church') + '?tab=settings'
            # Fallbacks
            if n.church and getattr(n.church, 'slug', None):
                return reverse('core:church_detail', args=[n.church.slug])
        except Exception:
            pass
        return reverse('core:notifications')

    regular_items = [
        {'notification': n, 'url': _target_url(n)}
        for n in regular_notifications
    ]
    
    message_items = [
        {'notification': n, 'url': _target_url(n)}
        for n in message_notifications
    ]
    
    return render(request, 'core/partials/notification_dropdown.html', {
        'dropdown_items': regular_items,
        'message_items': message_items,
        'unread_count': unread_count,
    })


@login_required
def mark_notification_read(request, notification_id):
    """Mark a specific notification as read."""
    from .notifications import mark_notification_as_read
    
    success = mark_notification_as_read(notification_id, request.user)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': success})
    else:
        if success:
            messages.success(request, 'Notification marked as read.')
        else:
            messages.error(request, 'Notification not found.')
        return redirect('core:notifications')


@login_required
def mark_all_notifications_read(request):
    """Mark all notifications as read for the current user."""
    from .notifications import mark_all_notifications_as_read
    
    count = mark_all_notifications_as_read(request.user)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'count': count,
            'message': f'{count} notifications marked as read.'
        })
    else:
        messages.success(request, f'{count} notifications marked as read.')
        return redirect('core:notifications')


def notification_count(request):
    """AJAX endpoint to get unread notification count."""
    if not request.user.is_authenticated:
        return JsonResponse({'count': 0, 'authenticated': False})
    
    try:
        from .notifications import get_user_unread_count
        count = get_user_unread_count(request.user)
        return JsonResponse({'count': count, 'authenticated': True})
    except Exception as e:
        # Log the error and return a safe response
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error getting notification count for user {request.user}: {str(e)}")
        return JsonResponse({'count': 0, 'authenticated': True, 'error': 'Failed to fetch count'})


# Service API - Test endpoint
def api_test(request):
    """Simple test endpoint to verify API routing works."""
    return JsonResponse({
        'success': True,
        'message': 'API is working!',
        'user': str(request.user),
        'authenticated': request.user.is_authenticated
    })

def api_get_service(request, service_id):
    """API endpoint to fetch service data for booking modal."""
    # Add CORS headers for debugging
    from django.http import JsonResponse
    
    # Always return JSON, even for errors
    try:
        # Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({
                'success': False,
                'error': 'Authentication required.'
            }, status=401)
            
        service = get_object_or_404(BookableService, id=service_id, is_active=True)
        church = service.church
        
        # Check if church is verified (required for bookings)
        if not church.is_verified:
            return JsonResponse({
                'success': False,
                'error': 'This church is not yet verified and cannot accept appointment requests.'
            }, status=400)
        
        # Get service image
        service_images = service.get_images()
        image_url = service_images[0].image.url if service_images else (service.image.url if service.image else None)
        
        # Prepare service data
        service_data = {
            'id': service.id,
            'name': service.name,
            'description': service.description or 'No description provided.',
            'image': image_url,
            'duration': service.duration_display,
            'price': service.price_display,
            'advance_booking_days': service.advance_booking_days,
            'is_free': service.is_free,
            'currency': service.currency,
            'raw_price': float(service.price) if service.price else 0.0,
            'duration_minutes': service.duration,
            'category': {
                'id': service.category.id,
                'name': service.category.name,
                'icon': service.category.icon,
                'color': service.category.color
            } if service.category else None
        }
        
        # Prepare church data
        church_data = {
            'id': church.id,
            'name': church.name,
            'slug': church.slug,
            'is_verified': church.is_verified
        }
        
        return JsonResponse({
            'success': True,
            'service': service_data,
            'church': church_data
        }, status=200)
        
    except BookableService.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Service not found or is not active.'
        }, status=404)
        
    except Exception as e:
        logger.error(f'API Service fetch error: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while fetching service data.'
        }, status=500)


def api_get_church_services(request, church_id):
    """API endpoint to fetch all bookable services for a specific church."""
    try:
        church = get_object_or_404(Church, id=church_id, is_active=True)
        
        # Get all active bookable services for this church
        services = church.bookable_services.filter(is_active=True).prefetch_related('service_images').order_by('name')
        
        services_data = []
        for service in services:
            # Get the first image if available
            images = service.get_images()
            image_url = images[0].image.url if images else None
            
            services_data.append({
                'id': service.id,
                'name': service.name,
                'description': service.description or 'No description provided.',
                'duration': service.duration_display,
                'price': service.price_display,
                'is_free': service.is_free,
                'currency': service.currency,
                'image_url': image_url,
                'advance_booking_days': service.advance_booking_days,
            })
        
        return JsonResponse({
            'success': True,
            'church_id': church.id,
            'church_name': church.name,
            'services': services_data,
            'total': len(services_data)
        }, status=200)
        
    except Church.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Church not found.'
        }, status=404)
        
    except Exception as e:
        logger.error(f'API Church Services fetch error: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while fetching church services.'
        }, status=500)


@login_required
def toggle_post_like(request, post_id):
    """AJAX endpoint to toggle like on a post."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        post = get_object_or_404(Post, id=post_id, is_active=True)
        
        # Check if user already liked the post
        like, created = PostLike.objects.get_or_create(
            user=request.user,
            post=post
        )
        
        if created:
            # User liked the post
            action = 'liked'
            message = 'Post liked'
            # Log activity
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_POST_LIKE,
                content_object=post,
                request=request
            )
        else:
            # User already liked, so unlike it
            like.delete()
            action = 'unliked'
            message = 'Post unliked'
            # Log activity
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_POST_UNLIKE,
                content_object=post,
                request=request
            )
        
        return JsonResponse({
            'success': True,
            'action': action,
            'message': message,
            'like_count': post.like_count,
            'is_liked': post.is_liked_by(request.user)
        })
        
    except Post.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Post not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': 'An error occurred'}, status=500)


@login_required
def add_post_comment(request, post_id):
    """AJAX endpoint to add a comment to a post."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        post = get_object_or_404(Post, id=post_id, is_active=True)
        content = request.POST.get('content', '').strip()
        parent_id = request.POST.get('parent_id')
        
        if not content:
            return JsonResponse({'success': False, 'message': 'Comment content is required'}, status=400)
        
        if len(content) > 500:
            return JsonResponse({'success': False, 'message': 'Comment is too long (max 500 characters)'}, status=400)
        
        # Handle reply to another comment
        parent_comment = None
        if parent_id:
            try:
                parent_comment = PostComment.objects.get(id=parent_id, post=post, is_active=True)
            except PostComment.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Parent comment not found'}, status=404)
        
        # Create the comment
        comment = PostComment.objects.create(
            user=request.user,
            post=post,
            content=content,
            parent=parent_comment
        )
        
        # Log activity
        from .models import UserInteraction
        UserInteraction.log_activity(
            user=request.user,
            activity_type=UserInteraction.ACTIVITY_POST_COMMENT,
            content_object=post,
            metadata={
                'comment_id': comment.id,
                'is_reply': parent_comment is not None,
                'parent_comment_id': parent_comment.id if parent_comment else None
            },
            request=request
        )
        
        # Get user display data
        user_display_name, user_initial = get_user_display_data(request.user, getattr(request.user, 'profile', None))
        
        # Get user profile picture
        user_profile = getattr(request.user, 'profile', None)
        user_profile_picture = None
        if user_profile:
            try:
                if user_profile.profile_image:
                    user_profile_picture = user_profile.profile_image.url
            except:
                pass
        
        # Get donation rank
        from accounts.donation_utils import get_user_donation_rank
        user_rank = get_user_donation_rank(request.user)
        
        return JsonResponse({
            'success': True,
            'message': 'Comment added successfully',
            'comment': {
                'id': comment.id,
                'content': comment.content,
                'user_name': user_display_name,
                'user_initial': user_initial,
                'user_profile_picture': user_profile_picture,
                'created_at': comment.created_at.isoformat(),
                'is_reply': comment.is_reply,
                'parent_id': parent_comment.id if parent_comment else None,
                'donation_rank': user_rank
            },
            'comment_count': post.comment_count
        })
        
    except Post.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Post not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': 'An error occurred'}, status=500)


@login_required
def get_post_comments(request, post_id):
    """AJAX endpoint to get comments for a post."""
    try:
        post = get_object_or_404(Post, id=post_id, is_active=True)
        
        # Get all top-level comments (not replies)
        comments = PostComment.objects.filter(
            post=post, 
            is_active=True, 
            parent__isnull=True
        ).select_related('user', 'user__profile').order_by('created_at')
        
        # Import donation rank utility
        from accounts.donation_utils import get_user_donation_rank
        
        comments_data = []
        for comment in comments:
            user_display_name, user_initial = get_user_display_data(comment.user, getattr(comment.user, 'profile', None))
            user_rank = get_user_donation_rank(comment.user)
            
            # Get user profile picture
            user_profile = getattr(comment.user, 'profile', None)
            user_profile_picture = None
            if user_profile:
                try:
                    if user_profile.profile_image:
                        user_profile_picture = user_profile.profile_image.url
                except:
                    pass
            
            # Get replies for this comment
            replies = comment.replies.filter(is_active=True).select_related('user', 'user__profile').order_by('created_at')
            replies_data = []
            
            for reply in replies:
                reply_user_display_name, reply_user_initial = get_user_display_data(reply.user, getattr(reply.user, 'profile', None))
                reply_rank = get_user_donation_rank(reply.user)
                
                # Get reply user profile picture
                reply_profile = getattr(reply.user, 'profile', None)
                reply_profile_picture = None
                if reply_profile:
                    try:
                        if reply_profile.profile_image:
                            reply_profile_picture = reply_profile.profile_image.url
                    except:
                        pass
                
                replies_data.append({
                    'id': reply.id,
                    'content': reply.content,
                    'user_name': reply_user_display_name,
                    'user_initial': reply_user_initial,
                    'user_profile_picture': reply_profile_picture,
                    'created_at': reply.created_at.isoformat(),
                    'is_reply': True,
                    'parent_id': comment.id,
                    'donation_rank': reply_rank
                })
            
            comments_data.append({
                'id': comment.id,
                'content': comment.content,
                'user_name': user_display_name,
                'user_initial': user_initial,
                'user_profile_picture': user_profile_picture,
                'created_at': comment.created_at.isoformat(),
                'is_reply': False,
                'replies': replies_data,
                'reply_count': len(replies_data),
                'donation_rank': user_rank
            })
        
        return JsonResponse({
            'success': True,
            'comments': comments_data,
            'comment_count': len(comments_data)
        })
        
    except Post.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Post not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': 'An error occurred'}, status=500)


@login_required 
def share_post(request, post_id):
    """AJAX endpoint to handle post sharing (future implementation for social sharing)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        post = get_object_or_404(Post, id=post_id, is_active=True)
        
        # For now, just return the share URL
        # In the future, this could track share counts or create notifications
        share_url = request.build_absolute_uri(f'/posts/{post.id}/')
        
        return JsonResponse({
            'success': True,
            'message': 'Share link generated',
            'share_url': share_url,
            'post_title': f"{post.church.name}'s post"
        })
        
    except Post.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Post not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': 'An error occurred'}, status=500)


def track_post_view(request, post_id):
    """AJAX endpoint to track post views."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        post = get_object_or_404(Post, id=post_id, is_active=True)
        
        # Get client info
        ip_address = get_client_ip(request)
        user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
        user = request.user if request.user.is_authenticated else None
        
        # Check if this view should be counted (prevent spam)
        should_count = True
        
        # For authenticated users, check if they viewed this post in the last hour
        if user:
            recent_view = PostView.objects.filter(
                post=post,
                user=user,
                viewed_at__gte=timezone.now() - timezone.timedelta(hours=1)
            ).exists()
            should_count = not recent_view
        else:
            # For anonymous users, check by IP in the last hour
            recent_view = PostView.objects.filter(
                post=post,
                ip_address=ip_address,
                viewed_at__gte=timezone.now() - timezone.timedelta(hours=1)
            ).exists()
            should_count = not recent_view
        
        if should_count:
            # Create view record
            PostView.objects.create(
                post=post,
                user=user,
                ip_address=ip_address,
                user_agent=user_agent
            )
            
            # Increment post view count
            post.view_count = F('view_count') + 1
            post.save(update_fields=['view_count'])
            
            # Refresh to get updated count
            post.refresh_from_db()
            
            # Log activity for authenticated users
            if user:
                from .models import UserInteraction
                UserInteraction.log_activity(
                    user=user,
                    activity_type=UserInteraction.ACTIVITY_POST_VIEW,
                    content_object=post,
                    request=request
                )
        
        return JsonResponse({
            'success': True,
            'view_count': post.view_count,
            'counted': should_count
        })
        
    except Post.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Post not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def report_comment(request):
    """AJAX endpoint to report a comment."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        comment_id = request.POST.get('comment_id')
        reason = request.POST.get('reason')
        description = request.POST.get('description', '').strip()
        
        if not comment_id or not reason or not description:
            return JsonResponse({
                'success': False,
                'message': 'Comment ID, reason, and description are required'
            }, status=400)
        
        # Get the comment
        comment = get_object_or_404(PostComment, id=comment_id, is_active=True)
        
        # Check if user already reported this comment
        from core.models import CommentReport
        existing_report = CommentReport.objects.filter(
            user=request.user,
            comment=comment
        ).first()
        
        if existing_report:
            return JsonResponse({
                'success': False,
                'message': 'You have already reported this comment'
            }, status=400)
        
        # Create the report
        report = CommentReport.objects.create(
            user=request.user,
            comment=comment,
            reason=reason,
            description=description,
            status='pending'
        )
        
        # TODO: Send notification to admins about new report
        
        return JsonResponse({
            'success': True,
            'message': 'Report submitted successfully',
            'report_id': report.id
        })
        
    except PostComment.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Comment not found'}, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error reporting comment: {str(e)}")
        return JsonResponse({'success': False, 'message': 'An error occurred while submitting your report'}, status=500)


@login_required
def delete_reported_comment(request, comment_id):
    """Delete a reported comment (super admin only)."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        comment = get_object_or_404(PostComment, id=comment_id)
        
        # Mark comment as inactive instead of deleting
        comment.is_active = False
        comment.save()
        
        # Update all reports for this comment to 'action_taken'
        from core.models import CommentReport
        CommentReport.objects.filter(comment=comment, status='pending').update(
            status='action_taken',
            reviewed_at=timezone.now(),
            reviewed_by=request.user,
            admin_notes='Comment deleted by admin'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Comment deleted successfully'
        })
        
    except PostComment.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Comment not found'}, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error deleting comment: {str(e)}")
        return JsonResponse({'success': False, 'message': 'An error occurred'}, status=500)


@login_required
def dismiss_comment_report(request, report_id):
    """Dismiss a comment report (super admin only)."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        from core.models import CommentReport
        report = get_object_or_404(CommentReport, id=report_id)
        
        report.status = 'dismissed'
        report.reviewed_at = timezone.now()
        report.reviewed_by = request.user
        report.admin_notes = request.POST.get('notes', 'Report dismissed by admin')
        report.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Report dismissed successfully'
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error dismissing report: {str(e)}")
        return JsonResponse({'success': False, 'message': 'An error occurred'}, status=500)


def get_client_ip(request):
    """Helper function to get client IP address."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@login_required
@require_POST
def toggle_post_bookmark(request, post_id):
    """Toggle bookmark status for a post via AJAX."""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user already bookmarked this post
        bookmark = PostBookmark.objects.filter(user=request.user, post=post).first()
        
        if bookmark:
            # Remove bookmark
            bookmark.delete()
            is_bookmarked = False
            action = 'unbookmarked'
            message = 'Post removed from saved posts'
            # Log activity
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_POST_UNBOOKMARK,
                content_object=post,
                request=request
            )
        else:
            # Add bookmark
            PostBookmark.objects.create(user=request.user, post=post)
            is_bookmarked = True
            action = 'bookmarked'
            message = 'Post saved successfully!'
            # Log activity
            from .models import UserInteraction
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_POST_BOOKMARK,
                content_object=post,
                request=request
            )
        
        # Debug logging
        print(f"User {request.user.username} {action} post {post.id}")
        
        return JsonResponse({
            'success': True,
            'is_bookmarked': is_bookmarked,
            'bookmark_count': post.bookmark_count,
            'action': action,
            'message': message
        })
        
    except Post.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Post not found'}, status=404)
    except Exception as e:
        print(f"Bookmark error: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def super_admin_user_activities(request):
    """Super Admin page to display user activity registration including verification codes, IP addresses, and device info."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('dashboard')

    # Import the models from accounts
    from accounts.models import UserActivity, EmailVerification, PasswordReset, LoginCode

    # Get filter parameters
    activity_type_filter = request.GET.get('activity_type', '')
    success_filter = request.GET.get('success', '')
    days_filter = request.GET.get('days', '30')  # Default to 30 days
    search_query = request.GET.get('search', '')
    
    # Convert days filter to integer
    try:
        days = int(days_filter) if days_filter else 30
    except ValueError:
        days = 30

    # Base queryset for user activities
    activities_qs = UserActivity.objects.select_related('user').all()
    
    # Apply date filter
    if days > 0:
        from datetime import timedelta
        cutoff_date = timezone.now() - timedelta(days=days)
        activities_qs = activities_qs.filter(created_at__gte=cutoff_date)
    
    # Apply activity type filter
    if activity_type_filter:
        activities_qs = activities_qs.filter(activity_type=activity_type_filter)
    
    # Apply success filter
    if success_filter:
        if success_filter == 'success':
            activities_qs = activities_qs.filter(success=True)
        elif success_filter == 'failed':
            activities_qs = activities_qs.filter(success=False)
    
    # Apply search filter
    if search_query:
        activities_qs = activities_qs.filter(
            Q(email__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(ip_address__icontains=search_query) |
            Q(verification_code__icontains=search_query)
        )

    # Get email verifications with additional tracking info
    verifications_qs = EmailVerification.objects.all()
    # Get password resets and login codes
    password_resets_qs = PasswordReset.objects.all()
    login_codes_qs = LoginCode.objects.all()
    
    # Apply date filter to verifications and code datasets
    if days > 0:
        verifications_qs = verifications_qs.filter(created_at__gte=cutoff_date)
        password_resets_qs = password_resets_qs.filter(created_at__gte=cutoff_date)
        login_codes_qs = login_codes_qs.filter(created_at__gte=cutoff_date)
    
    # Apply search filter to verifications
    if search_query:
        verifications_qs = verifications_qs.filter(
            Q(email__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )
        password_resets_qs = password_resets_qs.filter(
            Q(email__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )
        login_codes_qs = login_codes_qs.filter(
            Q(email__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )

    # Pagination for activities
    activities_paginator = Paginator(activities_qs, 5)
    activities_page = request.GET.get('activities_page', 1)
    try:
        activities = activities_paginator.page(activities_page)
    except:
        activities = activities_paginator.page(1)

    # Pagination for verifications  
    verifications_paginator = Paginator(verifications_qs, 5)
    verifications_page = request.GET.get('verifications_page', 1)
    try:
        verifications = verifications_paginator.page(verifications_page)
    except:
        verifications = verifications_paginator.page(1)

    # Pagination for password resets
    password_resets_paginator = Paginator(password_resets_qs, 5)
    password_resets_page = request.GET.get('password_resets_page', 1)
    try:
        password_resets = password_resets_paginator.page(password_resets_page)
    except:
        password_resets = password_resets_paginator.page(1)

    # Pagination for login codes
    login_codes_paginator = Paginator(login_codes_qs, 5)
    login_codes_page = request.GET.get('login_codes_page', 1)
    try:
        login_codes = login_codes_paginator.page(login_codes_page)
    except:
        login_codes = login_codes_paginator.page(1)

    # Activity type choices for filter dropdown
    activity_type_choices = UserActivity.ACTIVITY_TYPE_CHOICES

    # Statistics
    total_activities = activities_qs.count()
    total_verifications = verifications_qs.count()
    
    # Success/failure stats for activities
    success_count = activities_qs.filter(success=True).count()
    failed_count = activities_qs.filter(success=False).count()
    
    # Verification stats
    used_verifications = verifications_qs.filter(is_used=True).count()
    unused_verifications = verifications_qs.filter(is_used=False).count()
    expired_verifications = verifications_qs.filter(expires_at__lt=timezone.now()).count()

    # Password reset stats
    total_password_resets = password_resets_qs.count()
    used_password_resets = password_resets_qs.filter(is_used=True).count()
    expired_password_resets = password_resets_qs.filter(expires_at__lt=timezone.now()).count()
    unused_password_resets = max(total_password_resets - used_password_resets - expired_password_resets, 0)

    # Login code stats
    total_login_codes = login_codes_qs.count()
    used_login_codes = login_codes_qs.filter(is_used=True).count()
    expired_login_codes = login_codes_qs.filter(expires_at__lt=timezone.now()).count()
    unused_login_codes = max(total_login_codes - used_login_codes - expired_login_codes, 0)

    # Recent activity breakdown
    activity_breakdown = {}
    for activity_type, display_name in activity_type_choices:
        count = activities_qs.filter(activity_type=activity_type).count()
        activity_breakdown[display_name] = count

    from datetime import timedelta
    from django.db.models.functions import TruncDate
    today = timezone.now().date()
    chart_days = 14
    daily_labels = []
    daily_activity_counts = []
    daily_counts_qs = (
        activities_qs
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(c=Count('id'))
    )
    daily_count_map = {row['day']: row['c'] for row in daily_counts_qs}
    for i in range(chart_days - 1, -1, -1):
        day = today - timedelta(days=i)
        daily_labels.append(day.strftime('%b %d'))
        daily_activity_counts.append(daily_count_map.get(day, 0))

    type_counts_qs = activities_qs.values('activity_type').annotate(c=Count('id'))
    type_counts_map = {row['activity_type']: row['c'] for row in type_counts_qs}
    activity_type_labels = [label for _, label in activity_type_choices]
    activity_type_counts = [type_counts_map.get(code, 0) for code, label in activity_type_choices]

    codes_total_all = total_verifications + total_password_resets + total_login_codes
    codes_used_all = used_verifications + used_password_resets + used_login_codes
    codes_expired_all = expired_verifications + expired_password_resets + expired_login_codes
    codes_unused_all = max(codes_total_all - codes_used_all - codes_expired_all, 0)

    ctx = {
        'active': 'super_admin_activities',
        'page_title': 'User Activity Registration',
        'activities': activities,
        'verifications': verifications,
        'password_resets': password_resets,
        'login_codes': login_codes,
        'activity_type_choices': activity_type_choices,
        'now': timezone.now(),
        'current_filters': {
            'activity_type': activity_type_filter,
            'success': success_filter,
            'days': days_filter,
            'search': search_query,
        },
        'stats': {
            'total_activities': total_activities,
            'total_verifications': total_verifications,
            'success_count': success_count,
            'failed_count': failed_count,
            'used_verifications': used_verifications,
            'unused_verifications': unused_verifications,
            'expired_verifications': expired_verifications,
            'total_password_resets': total_password_resets,
            'used_password_resets': used_password_resets,
            'unused_password_resets': unused_password_resets,
            'expired_password_resets': expired_password_resets,
            'total_login_codes': total_login_codes,
            'used_login_codes': used_login_codes,
            'unused_login_codes': unused_login_codes,
            'expired_login_codes': expired_login_codes,
        },
        'activity_breakdown': activity_breakdown,
        'charts': {
            'daily_labels': daily_labels,
            'daily_activity_counts': daily_activity_counts,
            'activity_type_labels': activity_type_labels,
            'activity_type_counts': activity_type_counts,
            'codes_mix': {
                'total': codes_total_all,
                'used': codes_used_all,
                'expired': codes_expired_all,
                'unused': codes_unused_all,
            },
        },
    }
    ctx.update(_app_context(request))

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        partial = (request.GET.get('partial') or '').strip()
        if partial == 'activities':
            return render(request, 'core/partials/super_admin_activities_list.html', ctx)
        if partial == 'verifications':
            return render(request, 'core/partials/super_admin_verifications_list.html', ctx)
        if partial == 'login_codes':
            return render(request, 'core/partials/super_admin_login_codes_list.html', ctx)
        if partial == 'password_resets':
            return render(request, 'core/partials/super_admin_password_resets_list.html', ctx)

    return render(request, 'core/super_admin_user_activities.html', ctx)


# Super Admin - Posts Management
@login_required
def super_admin_posts(request):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    # Filters
    search = (request.GET.get('search') or '').strip()
    post_type = (request.GET.get('type') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()  # active|inactive|''
    reported_filter = (request.GET.get('reported') or '').strip()  # pending|any|''
    order = (request.GET.get('order') or '-created_at').strip()

    # Base queryset with annotations to avoid N+1 issues
    posts_qs = (
        Post.objects.select_related('church')
        .annotate(
            likes_count=Count('likes', distinct=True),
            comments_count=Count('comments', filter=Q(comments__is_active=True), distinct=True),
            bookmarks_count=Count('bookmarks', distinct=True),
            pending_reports_count=Count('reports', filter=Q(reports__status='pending'), distinct=True),
        )
    )

    if search:
        posts_qs = posts_qs.filter(Q(content__icontains=search) | Q(church__name__icontains=search))
    if post_type:
        posts_qs = posts_qs.filter(post_type=post_type)
    if status_filter == 'active':
        posts_qs = posts_qs.filter(is_active=True)
    elif status_filter == 'inactive':
        posts_qs = posts_qs.filter(is_active=False)
    if reported_filter == 'pending':
        posts_qs = posts_qs.filter(reports__status='pending')
    elif reported_filter == 'any':
        posts_qs = posts_qs.filter(reports__isnull=False)

    posts_qs = posts_qs.distinct()

    allowed_orders = {
        '-created_at', 'created_at',
        '-likes_count', 'likes_count',
        '-comments_count', 'comments_count',
        '-view_count', 'view_count',
    }
    if order not in allowed_orders:
        order = '-created_at'
    posts_qs = posts_qs.order_by(order)

    # Pagination
    paginator = Paginator(posts_qs, 20)
    page_number = request.GET.get('page')
    posts_page = paginator.get_page(page_number)

    # Analytics
    from datetime import timedelta
    from django.db.models import Sum
    today = timezone.now().date()
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)

    total_posts = Post.objects.count()
    posts_last_30_days = Post.objects.filter(created_at__date__gte=last_30_days).count()
    posts_last_7_days = Post.objects.filter(created_at__date__gte=last_7_days).count()
    new_posts_this_week = Post.objects.filter(created_at__date__gte=last_7_days).count()

    type_dist = Post.objects.values('post_type').annotate(count=Count('id'))
    type_map = {row['post_type']: row['count'] for row in type_dist}
    type_counts = {
        'general': type_map.get('general', 0),
        'photo': type_map.get('photo', 0),
        'event': type_map.get('event', 0),
        'prayer': type_map.get('prayer', 0),
    }

    total_likes = PostLike.objects.count()
    total_comments = PostComment.objects.filter(is_active=True).count()
    total_views = Post.objects.aggregate(total_views=Sum('view_count'))['total_views'] or 0
    total_shares = total_views  # Using views as shares for now
    
    # Calculate averages
    avg_likes_per_post = round(total_likes / total_posts) if total_posts > 0 else 0
    avg_comments_per_post = round(total_comments / total_posts) if total_posts > 0 else 0
    avg_shares_per_post = round(total_shares / total_posts) if total_posts > 0 else 0
    
    # Weekly engagement data (last 7 days)
    import json
    engagement_comments = []
    engagement_likes = []
    engagement_shares = []
    
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_comments = PostComment.objects.filter(created_at__date=day, is_active=True).count()
        day_likes = PostLike.objects.filter(created_at__date=day).count()
        day_shares = PostView.objects.filter(viewed_at__date=day).count()
        
        engagement_comments.append(day_comments)
        engagement_likes.append(day_likes)
        engagement_shares.append(day_shares)
    
    # Type distribution data
    type_data = [
        type_counts.get('general', 0),
        type_counts.get('event', 0),
        type_counts.get('photo', 0),
        type_counts.get('prayer', 0),
    ]

    likes_last_30_days = PostLike.objects.filter(created_at__date__gte=last_30_days).count()
    comments_last_30_days = PostComment.objects.filter(is_active=True, created_at__date__gte=last_30_days).count()
    views_last_30_days = PostView.objects.filter(viewed_at__date__gte=last_30_days).count()

    # Donations summary
    from decimal import Decimal
    donations_qs = Donation.objects.filter(payment_status='completed')
    donations_total_count = donations_qs.count()
    donations_total_amount = donations_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # Reports summary
    from .models import PostReport
    total_reports = PostReport.objects.count()
    pending_reports = PostReport.objects.filter(status='pending').count()
    reviewed_reports = PostReport.objects.filter(status='reviewed').count()
    dismissed_reports = PostReport.objects.filter(status='dismissed').count()
    action_taken_reports = PostReport.objects.filter(status='action_taken').count()
    top_report_reasons = list(
        PostReport.objects.values('reason').annotate(count=Count('id')).order_by('-count')[:5]
    )

    # Top posts by engagement
    analytics_posts_qs = (
        Post.objects.filter(is_active=True)
        .annotate(
            likes_count=Count('likes', distinct=True),
            comments_count=Count('comments', filter=Q(comments__is_active=True), distinct=True),
        )
        .annotate(
            total_engagement=ExpressionWrapper(
                F('view_count') + F('likes_count') + F('comments_count'),
                output_field=IntegerField(),
            )
        )
    )
    top_5_posts = list(analytics_posts_qs.select_related('church').order_by('-total_engagement', '-created_at')[:5])
    max_engagement = top_5_posts[0].total_engagement if top_5_posts else 0

    ctx = {
        'active': 'super_admin_posts',
        'page_title': 'Posts Management',
        'posts': posts_page,
        'search_query': search,
        'post_type_filter': post_type,
        'status_filter': status_filter,
        'reported_filter': reported_filter,
        'order': order,
        'stats': {
            'total_posts': total_posts,
            'new_posts_this_week': new_posts_this_week,
            'total_likes': total_likes,
            'avg_likes_per_post': avg_likes_per_post,
            'total_comments': total_comments,
            'avg_comments_per_post': avg_comments_per_post,
            'total_shares': total_shares,
            'avg_shares_per_post': avg_shares_per_post,
        },
        'engagement_data': {
            'comments': json.dumps(engagement_comments),
            'likes': json.dumps(engagement_likes),
            'shares': json.dumps(engagement_shares),
        },
        'type_data': json.dumps(type_data),
        'analytics': {
            'total_posts': total_posts,
            'posts_last_30_days': posts_last_30_days,
            'posts_last_7_days': posts_last_7_days,
            'type_counts': type_counts,
            'engagement': {
                'total_likes': total_likes,
                'total_comments': total_comments,
                'total_views': total_views,
                'likes_last_30_days': likes_last_30_days,
                'comments_last_30_days': comments_last_30_days,
                'views_last_30_days': views_last_30_days,
            },
            'donations': {
                'total_donations': donations_total_count,
                'total_amount': donations_total_amount,
            },
            'top_5_posts': top_5_posts,
            'max_engagement': max_engagement,
        },
        'reports_stats': {
            'total_reports': total_reports,
            'pending': pending_reports,
            'reviewed': reviewed_reports,
            'dismissed': dismissed_reports,
            'action_taken': action_taken_reports,
            'top_reasons': top_report_reasons,
        },
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_posts.html', ctx)


# Super Admin - Services Management
@login_required
def super_admin_services(request):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    from datetime import timedelta
    from django.db.models import Sum, Avg, Count, Q
    import json
    
    today = timezone.now().date()
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)
    
    # Get all services and bookings with annotations
    services = BookableService.objects.select_related('church', 'category').annotate(
        booking_count=Count('bookings', distinct=True)
    ).all()
    bookings = Booking.objects.select_related('service', 'user').all()
    
    # Calculate stats
    total_services = services.count()
    total_bookings = bookings.count()
    bookings_this_month = bookings.filter(created_at__date__gte=last_30_days).count()
    
    # Calculate revenue (paid bookings)
    total_revenue = bookings.filter(payment_status='paid').aggregate(
        total=Sum('service__price')
    )['total'] or 0
    revenue_this_month = bookings.filter(
        payment_status='paid',
        created_at__date__gte=last_30_days
    ).aggregate(total=Sum('service__price'))['total'] or 0
    
    # Calculate average rating
    from .models import ServiceReview
    avg_rating = ServiceReview.objects.aggregate(avg=Avg('rating'))['avg'] or 0
    total_reviews = ServiceReview.objects.count()
    
    # Booking trends (last 7 days)
    booking_trends_labels = []
    booking_trends_data = []
    revenue_trends_data = []
    
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_bookings = bookings.filter(created_at__date=day).count()
        day_revenue = bookings.filter(
            created_at__date=day,
            payment_status='paid'
        ).aggregate(total=Sum('service__price'))['total'] or 0
        
        booking_trends_labels.append(day.strftime('%a'))
        booking_trends_data.append(day_bookings)
        revenue_trends_data.append(float(day_revenue))
    
    # Services by category
    category_stats = services.values('category__name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    category_labels = []
    category_data = []
    for stat in category_stats:
        category_labels.append(stat['category__name'] or 'Uncategorized')
        category_data.append(stat['count'])
    
    # Recent reviews
    recent_reviews = ServiceReview.objects.select_related(
        'service', 'service__church', 'user'
    ).order_by('-created_at')[:10]
    
    # Paginate services
    from django.core.paginator import Paginator
    paginator = Paginator(services.order_by('-created_at'), 20)
    page_number = request.GET.get('page')
    services_page = paginator.get_page(page_number)

    ctx = {
        'active': 'super_admin_services',
        'page_title': 'Services Management',
        'stats': {
            'total_services': total_services,
            'total_bookings': total_bookings,
            'bookings_this_month': bookings_this_month,
            'total_revenue': total_revenue,
            'revenue_this_month': revenue_this_month,
            'avg_rating': round(avg_rating, 1),
            'total_reviews': total_reviews,
        },
        'booking_trends': {
            'labels': json.dumps(booking_trends_labels),
            'bookings': json.dumps(booking_trends_data),
            'revenue': json.dumps(revenue_trends_data),
        },
        'category_data': {
            'labels': json.dumps(category_labels),
            'data': json.dumps(category_data),
        },
        'services': services_page,
        'recent_reviews': recent_reviews,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_services.html', ctx)


@login_required
def super_admin_services_booking_data(request):
    """API endpoint to fetch booking data for different time ranges."""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    from datetime import timedelta
    
    days = int(request.GET.get('days', 7))
    today = timezone.now().date()
    
    labels = []
    bookings_data = []
    revenue_data = []
    
    for i in range(days - 1, -1, -1):
        day = today - timedelta(days=i)
        day_bookings = Booking.objects.filter(created_at__date=day).count()
        day_revenue = Booking.objects.filter(
            created_at__date=day,
            payment_status='paid'
        ).aggregate(total=Sum('service__price'))['total'] or 0
        
        bookings_data.append(day_bookings)
        revenue_data.append(float(day_revenue))
        
        # Format label based on time range
        if days <= 7:
            labels.append(day.strftime('%a'))
        elif days <= 30:
            labels.append(day.strftime('%m/%d'))
        else:
            labels.append(day.strftime('%m/%d'))
    
    return JsonResponse({
        'labels': labels,
        'bookings': bookings_data,
        'revenue': revenue_data
    })


@login_required
def super_admin_services_stats_data(request):
    """API endpoint to fetch services statistics for different time periods."""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    from datetime import timedelta
    from django.db.models import Sum, Avg
    
    period = request.GET.get('period', 'all')
    today = timezone.now().date()
    
    # Determine date filter
    if period == '7':
        start_date = today - timedelta(days=7)
        period_text = 'last 7 days'
    elif period == '30':
        start_date = today - timedelta(days=30)
        period_text = 'last 30 days'
    elif period == '90':
        start_date = today - timedelta(days=90)
        period_text = 'last 90 days'
    else:  # 'all'
        start_date = None
        period_text = 'all time'
    
    # Filter data
    if start_date:
        services_qs = BookableService.objects.filter(created_at__date__gte=start_date)
        bookings_qs = Booking.objects.filter(created_at__date__gte=start_date)
        reviews_qs = ServiceReview.objects.filter(created_at__date__gte=start_date)
    else:
        services_qs = BookableService.objects.all()
        bookings_qs = Booking.objects.all()
        reviews_qs = ServiceReview.objects.all()
    
    # Calculate stats
    total_services = services_qs.count()
    total_bookings = bookings_qs.count()
    total_revenue = bookings_qs.filter(payment_status='paid').aggregate(
        total=Sum('service__price')
    )['total'] or 0
    avg_rating = reviews_qs.aggregate(avg=Avg('rating'))['avg'] or 0
    total_reviews = reviews_qs.count()
    
    # Subtitles
    if period == 'all':
        services_subtitle = 'All services'
        bookings_subtitle = f'{total_bookings} total bookings'
        revenue_subtitle = f'₱{total_revenue:,.0f} total'
        rating_subtitle = f'From {total_reviews} reviews'
    else:
        services_subtitle = f'{total_services} in {period_text}'
        bookings_subtitle = f'+{total_bookings} in {period_text}'
        revenue_subtitle = f'+₱{total_revenue:,.0f} in {period_text}'
        rating_subtitle = f'From {total_reviews} reviews'
    
    return JsonResponse({
        'total_services': total_services,
        'services_subtitle': services_subtitle,
        'total_bookings': total_bookings,
        'bookings_subtitle': bookings_subtitle,
        'total_revenue': float(total_revenue),
        'revenue_subtitle': revenue_subtitle,
        'avg_rating': round(avg_rating, 1),
        'rating_subtitle': rating_subtitle
    })


# Super Admin - Service Categories Management
@login_required
def super_admin_categories(request):
    """Manage service categories."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    from .forms import ServiceCategoryForm
    
    categories = ServiceCategory.objects.all().order_by('order', 'name')
    
    ctx = {
        'active': 'super_admin_categories',
        'page_title': 'Service Categories',
        'categories': categories,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_categories.html', ctx)


@login_required
def super_admin_create_category(request):
    """Create a new service category."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    from .forms import ServiceCategoryForm
    
    if request.method == 'POST':
        form = ServiceCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" created successfully!')
            return JsonResponse({
                'success': True,
                'message': f'Category "{category.name}" created successfully!',
                'category_id': category.id
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Please correct the errors below.',
                'errors': form.errors
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required
def super_admin_edit_category(request, category_id):
    """Edit a service category."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    from .forms import ServiceCategoryForm
    
    category = get_object_or_404(ServiceCategory, id=category_id)
    
    if request.method == 'POST':
        form = ServiceCategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" updated successfully!')
            return JsonResponse({
                'success': True,
                'message': f'Category "{category.name}" updated successfully!',
                'category_id': category.id
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Please correct the errors below.',
                'errors': form.errors
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required
def super_admin_delete_category(request, category_id):
    """Delete a service category."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    category = get_object_or_404(ServiceCategory, id=category_id)
    
    if request.method == 'POST':
        # Check if category has services
        service_count = category.services.count()
        if service_count > 0:
            return JsonResponse({
                'success': False,
                'message': f'Cannot delete category "{category.name}" because it has {service_count} service(s) associated with it. Please reassign or delete those services first.'
            }, status=400)
        
        category_name = category.name
        category.delete()
        messages.success(request, f'Category "{category_name}" deleted successfully!')
        return JsonResponse({
            'success': True,
            'message': f'Category "{category_name}" deleted successfully!'
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required
def super_admin_toggle_category(request, category_id):
    """Toggle category active status."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    category = get_object_or_404(ServiceCategory, id=category_id)
    
    if request.method == 'POST':
        category.is_active = not category.is_active
        category.save()
        
        status_text = 'activated' if category.is_active else 'deactivated'
        messages.success(request, f'Category "{category.name}" {status_text} successfully!')
        return JsonResponse({
            'success': True,
            'message': f'Category "{category.name}" {status_text} successfully!',
            'is_active': category.is_active
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required
def super_admin_category_services(request, category_id):
    """Get all services for a specific category across all parishes."""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    try:
        from .models import BookableService
        
        category = get_object_or_404(ServiceCategory, id=category_id)
        
        # Get all services with this category
        services = BookableService.objects.filter(category=category).select_related(
            'church', 'category'
        ).order_by('church__name', 'name')
        
        services_data = []
        for service in services:
            services_data.append({
                'id': service.id,
                'name': service.name,
                'church_name': service.church.name,
                'church_id': service.church.id,
                'price': str(service.price),
                'duration': service.duration,
                'is_active': service.is_active,
                'description': service.description or '',
            })
        
        return JsonResponse({
            'success': True,
            'services': services_data,
            'category_name': category.name,
            'total_count': len(services_data)
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error fetching category services: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while fetching services'
        }, status=500)


# Super Admin - Bookings Management
@login_required
def super_admin_bookings(request):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    
    # Get time period filter from request
    period = request.GET.get('period', 'all')  # all, 7, 30, 90
    trend_days = int(request.GET.get('trend_days', 30))
    service_days = request.GET.get('service_days', 'all')
    
    # Get all bookings with related data
    bookings = Booking.objects.select_related(
        'user', 'church', 'service', 'service__category'
    ).order_by('-created_at')
    
    # Apply period filter for statistics
    if period != 'all':
        period_days = int(period)
        period_start = timezone.now() - timedelta(days=period_days)
        bookings_filtered = bookings.filter(created_at__gte=period_start)
    else:
        bookings_filtered = bookings
    
    # Calculate statistics (use filtered bookings)
    total_bookings = bookings_filtered.count()
    completed_bookings = bookings_filtered.filter(status=Booking.STATUS_COMPLETED).count()
    pending_bookings = bookings_filtered.filter(status=Booking.STATUS_REQUESTED).count()
    reviewed_bookings = bookings_filtered.filter(status=Booking.STATUS_REVIEWED).count()
    confirmed_bookings = bookings_filtered.filter(status=Booking.STATUS_APPROVED).count()
    cancelled_bookings = bookings_filtered.filter(
        Q(status=Booking.STATUS_CANCELED) | Q(status=Booking.STATUS_DECLINED)
    ).count()
    
    # Calculate online paid bookings
    online_paid_bookings = bookings_filtered.filter(payment_status='paid').count()
    
    # Calculate this week's bookings
    week_ago = timezone.now() - timedelta(days=7)
    this_week_bookings = bookings.filter(created_at__gte=week_ago).count()
    
    # Calculate this week's online paid bookings
    this_week_paid_bookings = bookings.filter(
        payment_status='paid',
        payment_date__gte=week_ago
    ).count()
    
    # Calculate completion rate
    completion_rate = (completed_bookings / total_bookings * 100) if total_bookings > 0 else 0
    
    # Calculate reviewed percentage
    reviewed_percentage = (reviewed_bookings / total_bookings * 100) if total_bookings > 0 else 0
    
    # Calculate cancellation rate
    cancellation_rate = (cancelled_bookings / total_bookings * 100) if total_bookings > 0 else 0
    
    # Calculate online payment rate
    online_payment_rate = (online_paid_bookings / total_bookings * 100) if total_bookings > 0 else 0
    
    # Get booking trends (based on trend_days filter)
    daily_trends = []
    for i in range(trend_days - 1, -1, -1):
        day = timezone.now() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        completed = bookings.filter(
            status=Booking.STATUS_COMPLETED,
            updated_at__gte=day_start,
            updated_at__lt=day_end
        ).count()
        
        cancelled = bookings.filter(
            Q(status=Booking.STATUS_CANCELED) | Q(status=Booking.STATUS_DECLINED),
            updated_at__gte=day_start,
            updated_at__lt=day_end
        ).count()
        
        pending = bookings.filter(
            status=Booking.STATUS_REQUESTED,
            created_at__gte=day_start,
            created_at__lt=day_end
        ).count()
        
        # Format label based on days
        if trend_days <= 7:
            label = day.strftime('%a')  # Mon, Tue, etc.
        elif trend_days <= 30:
            label = day.strftime('%b %d')  # Jan 15
        else:
            label = day.strftime('%m/%d')  # 01/15
        
        daily_trends.append({
            'month': label,
            'completed': completed,
            'cancelled': cancelled,
            'pending': pending
        })
    
    monthly_trends = daily_trends
    
    # Get revenue trend over time (based on trend_days filter)
    from django.db.models import Sum
    revenue_trends = []
    for i in range(trend_days - 1, -1, -1):
        day = timezone.now() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        revenue = bookings.filter(
            payment_status='paid',
            payment_date__gte=day_start,
            payment_date__lt=day_end
        ).aggregate(total=Sum('payment_amount'))['total'] or 0
        
        # Format label based on days
        if trend_days <= 7:
            label = day.strftime('%a')  # Mon, Tue, etc.
        elif trend_days <= 30:
            label = day.strftime('%b %d')  # Jan 15
        else:
            label = day.strftime('%m/%d')  # 01/15
        
        revenue_trends.append({
            'date': label,
            'revenue': float(revenue)
        })
    
    # Get bookings by service category (based on service_days filter)
    from core.models import ServiceCategory
    category_stats = []
    categories = ServiceCategory.objects.filter(is_active=True)
    
    # Apply service days filter
    if service_days != 'all':
        service_period_days = int(service_days)
        service_period_start = timezone.now() - timedelta(days=service_period_days)
        bookings_for_categories = bookings.filter(created_at__gte=service_period_start)
    else:
        bookings_for_categories = bookings
    
    for category in categories:
        count = bookings_for_categories.filter(service__category=category).count()
        if count > 0:
            category_stats.append({
                'name': category.name,
                'count': count,
                'color': category.color
            })
    
    # Add uncategorized services
    uncategorized_count = bookings_for_categories.filter(service__category__isnull=True).count()
    if uncategorized_count > 0:
        category_stats.append({
            'name': 'Uncategorized',
            'count': uncategorized_count,
            'color': '#9CA3AF'
        })
    
    # Sort by count descending
    category_stats = sorted(category_stats, key=lambda x: x['count'], reverse=True)[:5]
    
    # Get online paid bookings by parish (church) and revenue
    from core.models import Church
    from django.db.models import Sum
    parish_paid_stats = []
    parish_revenue_stats = []
    churches = Church.objects.filter(is_active=True).order_by('name')
    
    for church in churches:
        paid_count = bookings_filtered.filter(
            church=church,
            payment_status='paid'
        ).count()
        if paid_count > 0:
            parish_paid_stats.append({
                'name': church.name,
                'count': paid_count,
                'color': '#3B82F6'  # Default blue color
            })
        
        # Calculate revenue for this church
        revenue = bookings_filtered.filter(
            church=church,
            payment_status='paid'
        ).aggregate(total=Sum('payment_amount'))['total'] or 0
        
        if revenue > 0:
            parish_revenue_stats.append({
                'name': church.name,
                'revenue': float(revenue),
                'color': '#10B981'  # Green color for revenue
            })
    
    # Sort by count/revenue descending, then reverse for horizontal bar charts
    # This makes highest values appear at the bottom (best visual practice for horizontal bars)
    parish_paid_stats = sorted(parish_paid_stats, key=lambda x: x['count'], reverse=True)[:10]
    parish_paid_stats = list(reversed(parish_paid_stats))
    
    parish_revenue_stats = sorted(parish_revenue_stats, key=lambda x: x['revenue'], reverse=True)[:10]
    parish_revenue_stats = list(reversed(parish_revenue_stats))
    
    # Paginate bookings (20 per page)
    from django.core.paginator import Paginator
    paginator = Paginator(bookings, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Convert data to JSON for JavaScript
    import json
    monthly_trends_json = json.dumps(monthly_trends)
    category_stats_json = json.dumps(category_stats)
    parish_paid_stats_json = json.dumps(parish_paid_stats)
    parish_revenue_stats_json = json.dumps(parish_revenue_stats)
    revenue_trends_json = json.dumps(revenue_trends)
    
    ctx = {
        'active': 'super_admin_bookings',
        'page_title': 'Bookings Management',
        'bookings': page_obj,
        'total_bookings': total_bookings,
        'completed_bookings': completed_bookings,
        'pending_bookings': pending_bookings,
        'reviewed_bookings': reviewed_bookings,
        'confirmed_bookings': confirmed_bookings,
        'cancelled_bookings': cancelled_bookings,
        'online_paid_bookings': online_paid_bookings,
        'this_week_bookings': this_week_bookings,
        'this_week_paid_bookings': this_week_paid_bookings,
        'completion_rate': round(completion_rate, 1),
        'reviewed_percentage': round(reviewed_percentage, 1),
        'cancellation_rate': round(cancellation_rate, 1),
        'online_payment_rate': round(online_payment_rate, 1),
        'monthly_trends': monthly_trends_json,
        'category_stats': category_stats_json,
        'parish_paid_stats': parish_paid_stats_json,
        'parish_revenue_stats': parish_revenue_stats_json,
        'revenue_trends': revenue_trends_json,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_bookings.html', ctx)


# Super Admin - Bookings Chart Data API
@login_required
def super_admin_bookings_chart_data(request):
    """API endpoint to fetch chart data for bookings page without page reload."""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    from django.db.models import Count, Q, Sum
    from datetime import datetime, timedelta
    import json
    
    # Get parameters
    chart = request.GET.get('chart')  # 'trends', 'serviceType', 'revenueTrend' (default)
    trend_days = int(request.GET.get('trend_days', 30))
    service_days = request.GET.get('service_days', 'all')
    
    # Get all bookings with related data
    bookings = Booking.objects.select_related(
        'user', 'church', 'service', 'service__category'
    ).order_by('-created_at')
    
    # Booking Trends (daily aggregates)
    if chart == 'trends':
        monthly_trends = []
        for i in range(trend_days - 1, -1, -1):
            day = timezone.now() - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            
            completed = bookings.filter(
                status=Booking.STATUS_COMPLETED,
                updated_at__gte=day_start,
                updated_at__lt=day_end
            ).count()
            cancelled = bookings.filter(
                Q(status=Booking.STATUS_CANCELED) | Q(status=Booking.STATUS_DECLINED),
                updated_at__gte=day_start,
                updated_at__lt=day_end
            ).count()
            pending = bookings.filter(
                status=Booking.STATUS_REQUESTED,
                created_at__gte=day_start,
                created_at__lt=day_end
            ).count()
            
            if trend_days <= 7:
                label = day.strftime('%a')  # Mon, Tue, etc.
            elif trend_days <= 30:
                label = day.strftime('%b %d')  # Jan 15
            else:
                label = day.strftime('%m/%d')  # 01/15
            
            monthly_trends.append({
                'month': label,
                'completed': completed,
                'cancelled': cancelled,
                'pending': pending,
            })
        return JsonResponse({'monthly_trends': monthly_trends})
    
    # Bookings by Service Type
    if chart == 'serviceType':
        from core.models import ServiceCategory
        category_stats = []
        categories = ServiceCategory.objects.filter(is_active=True)
        
        if service_days != 'all':
            service_period_days = int(service_days)
            service_period_start = timezone.now() - timedelta(days=service_period_days)
            bookings_for_categories = bookings.filter(created_at__gte=service_period_start)
        else:
            bookings_for_categories = bookings
        
        for category in categories:
            count = bookings_for_categories.filter(service__category=category).count()
            if count > 0:
                category_stats.append({
                    'name': category.name,
                    'count': count,
                    'color': category.color,
                })
        # Add uncategorized services
        uncategorized_count = bookings_for_categories.filter(service__category__isnull=True).count()
        if uncategorized_count > 0:
            category_stats.append({
                'name': 'Uncategorized',
                'count': uncategorized_count,
                'color': '#9CA3AF',
            })
        category_stats = sorted(category_stats, key=lambda x: x['count'], reverse=True)[:5]
        return JsonResponse({'category_stats': category_stats})
    
    # Default: Revenue Trend Over Time (kept for backward compatibility)
    revenue_trends = []
    for i in range(trend_days - 1, -1, -1):
        day = timezone.now() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        revenue = bookings.filter(
            payment_status='paid',
            payment_date__gte=day_start,
            payment_date__lt=day_end
        ).aggregate(total=Sum('payment_amount'))['total'] or 0
        
        if trend_days <= 7:
            label = day.strftime('%a')  # Mon, Tue, etc.
        elif trend_days <= 30:
            label = day.strftime('%b %d')  # Jan 15
        else:
            label = day.strftime('%m/%d')  # 01/15
        
        revenue_trends.append({'date': label, 'revenue': float(revenue)})
    return JsonResponse({'revenue_trends': revenue_trends})


# Super Admin - Donation Rankings
@login_required
def super_admin_donations(request):
    """Super Admin donations and rankings page."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    from django.db.models import Sum, Count, Q
    from django.contrib.auth.models import User
    from accounts.donation_utils import RANK_TIERS, get_user_donation_rank
    from datetime import datetime, timedelta
    
    # Get filter parameters
    rank_filter = request.GET.get('rank', 'all')
    time_filter = request.GET.get('time', 'all')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'amount')  # amount, date, name
    
    # Get all users with completed donations
    donors = User.objects.filter(
        donations_made__payment_status='completed'
    ).annotate(
        total_donated=Sum('donations_made__amount', filter=Q(donations_made__payment_status='completed')),
        donation_count=Count('donations_made', filter=Q(donations_made__payment_status='completed'))
    ).filter(total_donated__gte=50).select_related('profile')
    
    # Apply time filter
    if time_filter != 'all':
        now = timezone.now()
        if time_filter == '30days':
            start_date = now - timedelta(days=30)
        elif time_filter == '90days':
            start_date = now - timedelta(days=90)
        elif time_filter == 'year':
            start_date = now - timedelta(days=365)
        else:
            start_date = None
        
        if start_date:
            donors = donors.filter(donations_made__created_at__gte=start_date)
    
    # Apply search filter
    if search_query:
        donors = donors.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(profile__display_name__icontains=search_query)
        )
    
    # Build donor data with ranks
    donor_data = []
    for donor in donors:
        rank = get_user_donation_rank(donor)
        
        # Apply rank filter
        if rank_filter != 'all' and (not rank or rank['icon'] != rank_filter):
            continue
        
        # Get user display name
        display_name = donor.username
        if hasattr(donor, 'profile') and donor.profile and donor.profile.display_name:
            display_name = donor.profile.display_name
        elif donor.get_full_name():
            display_name = donor.get_full_name()
        
        donor_data.append({
            'user': donor,
            'display_name': display_name,
            'email': donor.email,
            'total_donated': donor.total_donated or 0,
            'donation_count': donor.donation_count or 0,
            'rank': rank,
            'show_rank': donor.profile.show_donation_rank if hasattr(donor, 'profile') and donor.profile else True
        })
    
    # Sort donors
    if sort_by == 'amount':
        donor_data.sort(key=lambda x: x['total_donated'], reverse=True)
    elif sort_by == 'date':
        # Get latest donation date for each donor
        for data in donor_data:
            latest = data['user'].donations_made.filter(
                payment_status='completed'
            ).order_by('-created_at').first()
            data['latest_donation'] = latest.created_at if latest else timezone.now()
        donor_data.sort(key=lambda x: x['latest_donation'], reverse=True)
    elif sort_by == 'name':
        donor_data.sort(key=lambda x: x['display_name'].lower())
    
    # Calculate statistics
    total_donors = len(donor_data)
    total_donations_amount = sum(d['total_donated'] for d in donor_data)
    
    # Rank distribution
    rank_distribution = {}
    for tier in RANK_TIERS:
        rank_distribution[tier['icon']] = {
            'name': tier['name'],
            'color': tier['color'],
            'count': 0,
            'total_amount': 0
        }
    
    for data in donor_data:
        if data['rank']:
            icon = data['rank']['icon']
            rank_distribution[icon]['count'] += 1
            rank_distribution[icon]['total_amount'] += data['total_donated']
    
    # Top donors (top 10)
    top_donors = donor_data[:10]
    
    # Serialize data for JavaScript charts
    import json
    from decimal import Decimal
    
    # Convert rank_distribution to JSON-safe format
    rank_distribution_json = {}
    for icon, data in rank_distribution.items():
        rank_distribution_json[icon] = {
            'name': data['name'],
            'color': data['color'],
            'count': data['count'],
            'total_amount': float(data['total_amount']) if data['total_amount'] else 0
        }
    
    # Convert donor_data to JSON-safe format
    donor_data_json = []
    for data in donor_data:
        donor_data_json.append({
            'display_name': data['display_name'],
            'email': data['email'],
            'total_donated': float(data['total_donated']) if data['total_donated'] else 0,
            'donation_count': data['donation_count'],
            'rank': {
                'name': data['rank']['name'],
                'color': data['rank']['color'],
                'icon': data['rank']['icon']
            } if data['rank'] else None,
            'show_rank': data['show_rank']
        })
    
    # Convert rank_tiers to JSON-safe format
    rank_tiers_json = []
    for tier in RANK_TIERS:
        rank_tiers_json.append({
            'name': tier['name'],
            'color': tier['color'],
            'icon': tier['icon']
        })
    
    context = {
        'donor_data': donor_data,
        'total_donors': total_donors,
        'total_donations_amount': total_donations_amount,
        'rank_distribution': rank_distribution,
        'top_donors': top_donors,
        'rank_tiers': RANK_TIERS,
        'rank_filter': rank_filter,
        'time_filter': time_filter,
        'search_query': search_query,
        'sort_by': sort_by,
        'active': 'super_admin_donations',
        # JSON data for charts
        'rank_distribution_json': json.dumps(rank_distribution_json),
        'donor_data_json': json.dumps(donor_data_json),
        'rank_tiers_json': json.dumps(rank_tiers_json),
    }
    context.update(_app_context(request))
    
    return render(request, 'core/super_admin_donations.html', context)


@login_required
def super_admin_donations_filter_data(request):
    """AJAX endpoint for filtering donation data by time period."""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    from django.db.models import Sum, Count, Q
    from django.contrib.auth.models import User
    from accounts.donation_utils import RANK_TIERS, get_user_donation_rank
    from datetime import timedelta
    import json
    
    period = request.GET.get('period', 'all')
    
    # Build query for donations based on period
    donors_query = User.objects.filter(
        donations_made__payment_status='completed'
    )
    
    # Apply time filter
    donation_filter = Q(donations_made__payment_status='completed')
    if period != 'all':
        now = timezone.now()
        if period == '7':
            start_date = now - timedelta(days=7)
        elif period == '30':
            start_date = now - timedelta(days=30)
        elif period == '90':
            start_date = now - timedelta(days=90)
        else:
            start_date = None
        
        if start_date:
            donation_filter &= Q(donations_made__created_at__gte=start_date)
    
    # Get donors with filtered donations
    donors = donors_query.annotate(
        total_donated=Sum('donations_made__amount', filter=donation_filter),
        donation_count=Count('donations_made', filter=donation_filter)
    ).filter(total_donated__gte=50).select_related('profile')
    
    # Build donor data with ranks
    donor_data = []
    for donor in donors:
        rank = get_user_donation_rank(donor)
        
        display_name = donor.username
        if hasattr(donor, 'profile') and donor.profile and donor.profile.display_name:
            display_name = donor.profile.display_name
        elif donor.get_full_name():
            display_name = donor.get_full_name()
        
        donor_data.append({
            'display_name': display_name,
            'total_donated': float(donor.total_donated or 0),
            'donation_count': donor.donation_count or 0,
            'rank': rank
        })
    
    # Sort by amount
    donor_data.sort(key=lambda x: x['total_donated'], reverse=True)
    
    # Calculate statistics
    total_donors = len(donor_data)
    total_donations = sum(d['total_donated'] for d in donor_data)
    avg_donation = total_donations / total_donors if total_donors > 0 else 0
    
    # Rank distribution
    rank_distribution = {}
    for tier in RANK_TIERS:
        rank_distribution[tier['icon']] = {
            'name': tier['name'],
            'color': tier['color'],
            'count': 0,
            'total_amount': 0
        }
    
    for data in donor_data:
        if data['rank']:
            icon = data['rank']['icon']
            rank_distribution[icon]['count'] += 1
            rank_distribution[icon]['total_amount'] += data['total_donated']
    
    # Prepare chart data for rank distribution
    rank_chart_labels = []
    rank_chart_data = []
    rank_chart_colors = []
    for tier in RANK_TIERS:
        tier_data = rank_distribution[tier['icon']]
        if tier_data['count'] > 0:
            rank_chart_labels.append(tier['name'])
            rank_chart_data.append(tier_data['count'])
            rank_chart_colors.append(tier['color'])
    
    # Prepare chart data for top donors
    top_10 = donor_data[:10]
    top_donors_labels = [d['display_name'] for d in top_10]
    top_donors_amounts = [d['total_donated'] for d in top_10]
    top_donors_colors = [d['rank']['color'] if d['rank'] else '#9ca3af' for d in top_10]
    
    # Period labels
    period_labels = {
        'all': 'All-time',
        '7': 'Last 7 days',
        '30': 'Last 30 days',
        '90': 'Last 90 days'
    }
    period_label = period_labels.get(period, 'All-time')
    
    response_data = {
        'total_donors': total_donors,
        'donors_subtitle': f'{period_label} contributors',
        'total_donations': total_donations,
        'donations_subtitle': f'{period_label} contributions',
        'avg_donation': avg_donation,
        'avg_subtitle': f'{period_label} per donor',
        'rank_distribution': rank_distribution,
        'rank_distribution_chart': {
            'labels': rank_chart_labels,
            'data': rank_chart_data,
            'colors': rank_chart_colors
        },
        'top_donors_chart': {
            'labels': top_donors_labels,
            'amounts': top_donors_amounts,
            'colors': top_donors_colors
        }
    }
    
    return JsonResponse(response_data)


# Super Admin - Parish Donations Analytics
@login_required
def super_admin_parish_donations(request):
    """Super Admin parish donations analytics page."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    from django.db.models import Sum, Count, Avg, Q
    from datetime import timedelta
    
    # Get all completed donations grouped by parish
    donations = Donation.objects.filter(
        payment_status='completed'
    ).select_related('post', 'post__church', 'donor')
    
    # Calculate overall statistics
    stats = {
        'total_donations': donations.aggregate(total=Sum('amount'))['total'] or 0,
        'total_count': donations.count(),
        'avg_donation': donations.aggregate(avg=Avg('amount'))['avg'] or 0,
        'total_parishes': donations.values('post__church').distinct().count(),
    }
    
    # Get donations by parish
    from django.db.models import F
    parish_donations = donations.values(
        'post__church__id',
        'post__church__name'
    ).annotate(
        total_amount=Sum('amount'),
        donation_count=Count('id'),
        avg_amount=Avg('amount')
    ).order_by('-total_amount')[:10]
    
    # Get recent donations (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_stats = donations.filter(created_at__gte=thirty_days_ago).aggregate(
        total=Sum('amount'),
        count=Count('id')
    )
    stats['recent_donations'] = recent_stats['total'] or 0
    stats['recent_count'] = recent_stats['count'] or 0
    
    # Get donation trends (last 7 days)
    trend_data = []
    trend_labels = []
    for i in range(6, -1, -1):
        date = timezone.now() - timedelta(days=i)
        day_donations = donations.filter(
            created_at__date=date.date()
        ).aggregate(total=Sum('amount'))
        trend_data.append(float(day_donations['total'] or 0))
        trend_labels.append(date.strftime('%b %d'))
    
    # Get payment method distribution
    payment_methods = donations.values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-total')
    
    # Prepare chart data
    parish_chart_labels = [p['post__church__name'] for p in parish_donations]
    parish_chart_data = [float(p['total_amount']) for p in parish_donations]
    
    payment_method_labels = []
    payment_method_data = []
    payment_method_dict = dict(Donation.PAYMENT_METHOD_CHOICES)
    for pm in payment_methods:
        payment_method_labels.append(payment_method_dict.get(pm['payment_method'], pm['payment_method']))
        payment_method_data.append(pm['count'])
    
    context = {
        'stats': stats,
        'parish_donations': parish_donations,
        'trend_labels': trend_labels,
        'trend_data': trend_data,
        'parish_chart_labels': parish_chart_labels,
        'parish_chart_data': parish_chart_data,
        'payment_method_labels': payment_method_labels,
        'payment_method_data': payment_method_data,
    }
    
    return render(request, 'core/super_admin_parish_donations.html', context)


# Super Admin - Moderation Management
@login_required
def super_admin_moderation(request):
    """Super Admin moderation page for managing reports and verifications."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')
    
    from datetime import timedelta
    from core.models import CommentReport
    
    # Get post reports
    post_reports = PostReport.objects.select_related(
        'post', 'post__church', 'user'
    ).order_by('-created_at')[:20]
    
    # Get comment reports
    comment_reports = CommentReport.objects.select_related(
        'comment', 'comment__post', 'comment__post__church', 'comment__user', 'user'
    ).order_by('-created_at')[:20]
    
    # Get church verification requests
    verifications = ChurchVerificationRequest.objects.select_related(
        'church', 'submitted_by'
    ).prefetch_related('documents').filter(
        status=ChurchVerificationRequest.STATUS_PENDING
    ).order_by('created_at')[:20]
    
    # Calculate statistics
    pending_post_reports = PostReport.objects.filter(status='pending').count()
    pending_comment_reports = CommentReport.objects.filter(status='pending').count()
    pending_reports = pending_post_reports + pending_comment_reports
    
    high_severity_reports = pending_reports  # You can add severity field later
    
    church_verifications = ChurchVerificationRequest.objects.filter(
        status=ChurchVerificationRequest.STATUS_PENDING
    ).count()
    
    # Resolved this week
    seven_days_ago = timezone.now() - timedelta(days=7)
    resolved_post_reports = PostReport.objects.filter(
        status__in=['reviewed', 'dismissed', 'action_taken'],
        reviewed_at__isnull=False,
        reviewed_at__gte=seven_days_ago
    ).count()
    resolved_comment_reports = CommentReport.objects.filter(
        status__in=['reviewed', 'dismissed', 'action_taken'],
        reviewed_at__isnull=False,
        reviewed_at__gte=seven_days_ago
    ).count()
    resolved_this_week = resolved_post_reports + resolved_comment_reports
    
    # Calculate resolution rate
    total_post_reports = PostReport.objects.count()
    total_comment_reports = CommentReport.objects.count()
    total_reports = total_post_reports + total_comment_reports
    
    resolved_post_reports_all = PostReport.objects.filter(
        status__in=['reviewed', 'dismissed', 'action_taken']
    ).count()
    resolved_comment_reports_all = CommentReport.objects.filter(
        status__in=['reviewed', 'dismissed', 'action_taken']
    ).count()
    resolved_reports = resolved_post_reports_all + resolved_comment_reports_all
    
    resolution_rate = int((resolved_reports / total_reports * 100)) if total_reports > 0 else 0
    
    # Calculate weekly activity data for chart (last 4 weeks)
    from django.db.models import Count, Q
    from django.db.models.functions import TruncWeek
    
    weekly_activity = []
    for i in range(3, -1, -1):  # Last 4 weeks
        week_start = timezone.now() - timedelta(days=(i+1)*7)
        week_end = timezone.now() - timedelta(days=i*7)
        
        # New reports this week
        new_post_reports = PostReport.objects.filter(
            created_at__gte=week_start,
            created_at__lt=week_end
        ).count()
        new_comment_reports = CommentReport.objects.filter(
            created_at__gte=week_start,
            created_at__lt=week_end
        ).count()
        new_reports = new_post_reports + new_comment_reports
        
        # Resolved this week
        resolved_post = PostReport.objects.filter(
            reviewed_at__gte=week_start,
            reviewed_at__lt=week_end,
            status__in=['reviewed', 'action_taken']
        ).count()
        resolved_comment = CommentReport.objects.filter(
            reviewed_at__gte=week_start,
            reviewed_at__lt=week_end,
            status__in=['reviewed', 'action_taken']
        ).count()
        resolved = resolved_post + resolved_comment
        
        # Dismissed this week
        dismissed_post = PostReport.objects.filter(
            reviewed_at__gte=week_start,
            reviewed_at__lt=week_end,
            status='dismissed'
        ).count()
        dismissed_comment = CommentReport.objects.filter(
            reviewed_at__gte=week_start,
            reviewed_at__lt=week_end,
            status='dismissed'
        ).count()
        dismissed = dismissed_post + dismissed_comment
        
        weekly_activity.append({
            'new': new_reports,
            'resolved': resolved,
            'dismissed': dismissed
        })
    
    # Calculate report reasons distribution
    post_reasons = PostReport.objects.values('reason').annotate(count=Count('id'))
    comment_reasons = CommentReport.objects.values('reason').annotate(count=Count('id'))
    
    # Combine and aggregate reasons
    reason_counts = {}
    for report in post_reasons:
        reason = report['reason']
        reason_counts[reason] = reason_counts.get(reason, 0) + report['count']
    
    for report in comment_reasons:
        reason = report['reason']
        reason_counts[reason] = reason_counts.get(reason, 0) + report['count']
    
    # Map reason codes to display names
    reason_labels = {
        'spam': 'Spam',
        'inappropriate': 'Inappropriate Content',
        'harassment': 'Harassment',
        'violence': 'Violence',
        'false_info': 'Misinformation',
        'offensive': 'Suspicious Activity',
        'other': 'Other'
    }
    
    report_reasons = {
        'labels': [reason_labels.get(k, k.title()) for k in reason_counts.keys()],
        'data': list(reason_counts.values())
    }
    
    stats = {
        'pending_reports': pending_reports,
        'pending_post_reports': pending_post_reports,
        'pending_comment_reports': pending_comment_reports,
        'high_severity_reports': high_severity_reports,
        'church_verifications': church_verifications,
        'resolved_this_week': resolved_this_week,
        'resolution_rate': resolution_rate,
        'avg_response_time': '2.3',  # You can calculate this based on your data
    }
    
    import json
    ctx = {
        'active': 'super_admin_moderation',
        'page_title': 'Moderation',
        'post_reports': post_reports,
        'comment_reports': comment_reports,
        'verifications': verifications,
        'stats': stats,
        'weekly_activity_json': json.dumps(weekly_activity),
        'report_reasons_json': json.dumps(report_reasons),
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_moderation.html', ctx)


@login_required
def super_admin_post_detail(request, post_id):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to access Super Admin.')
        return redirect('core:home')

    post = get_object_or_404(Post.objects.select_related('church'), id=post_id)

    likes_count = PostLike.objects.filter(post=post).count()
    comments_count = PostComment.objects.filter(post=post, is_active=True).count()
    bookmarks_count = PostBookmark.objects.filter(post=post).count()
    views_count = getattr(post, 'view_count', 0)

    from datetime import timedelta
    today = timezone.now().date()
    daily_labels = []
    daily_views = []
    daily_likes = []
    daily_comments = []
    for i in range(13, -1, -1):
        day = today - timedelta(days=i)
        daily_labels.append(day.strftime('%b %d'))
        daily_views.append(PostView.objects.filter(post=post, viewed_at__date=day).count())
        daily_likes.append(PostLike.objects.filter(post=post, created_at__date=day).count())
        daily_comments.append(PostComment.objects.filter(post=post, is_active=True, created_at__date=day).count())

    donations = Donation.objects.filter(post=post, payment_status='completed').select_related('donor', 'donor__profile')
    from decimal import Decimal
    from django.db.models import Sum
    donations_total_amount = donations.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    donations_total_count = donations.count()
    top_donors_data = {}
    for d in donations:
        if d.donor_id:
            top_donors_data.setdefault(d.donor_id, Decimal('0.00'))
            top_donors_data[d.donor_id] += d.amount
    top_donor_ids = sorted(top_donors_data, key=top_donors_data.get, reverse=True)[:5]
    top_donors = []
    if top_donor_ids:
        users = {u.id: u for u in User.objects.select_related('profile').filter(id__in=top_donor_ids)}
        for uid in top_donor_ids:
            u = users.get(uid)
            if u:
                setattr(u, 'total_donated', top_donors_data.get(uid))
                top_donors.append(u)

    reports = list(PostReport.objects.filter(post=post).select_related('user').order_by('-created_at')[:20])
    reports_status_counts = {
        'pending': PostReport.objects.filter(post=post, status='pending').count(),
        'reviewed': PostReport.objects.filter(post=post, status='reviewed').count(),
        'dismissed': PostReport.objects.filter(post=post, status='dismissed').count(),
        'action_taken': PostReport.objects.filter(post=post, status='action_taken').count(),
    }
    reasons_dist_qs = PostReport.objects.filter(post=post).values('reason').annotate(c=Count('id')).order_by('-c')
    reasons_distribution = list(reasons_dist_qs[:5])

    recent_comments = list(
        PostComment.objects.filter(post=post, is_active=True).select_related('user', 'user__profile').order_by('-created_at')[:10]
    )

    ctx = {
        'active': 'super_admin_posts',
        'page_title': 'Post Details',
        'post': post,
        'counts': {
            'likes': likes_count,
            'comments': comments_count,
            'bookmarks': bookmarks_count,
            'views': views_count,
        },
        'daily': {
            'labels': daily_labels,
            'views': daily_views,
            'likes': daily_likes,
            'comments': daily_comments,
        },
        'donations': {
            'total_amount': donations_total_amount,
            'total_count': donations_total_count,
            'top_donors': top_donors,
        },
        'reports': {
            'list': reports,
            'status_counts': reports_status_counts,
            'reasons': reasons_distribution,
        },
        'recent_comments': recent_comments,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/super_admin_post_detail.html', ctx)


@login_required
@require_POST
def super_admin_toggle_post_active(request, post_id):
    if not request.user.is_superuser:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Forbidden'}, status=403)
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('core:super_admin_posts')

    try:
        post = get_object_or_404(Post, id=post_id)
        post.is_active = not post.is_active
        post.save(update_fields=['is_active', 'updated_at'])
        msg = 'Post activated.' if post.is_active else 'Post deactivated.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': msg, 'is_active': post.is_active})
        messages.success(request, msg)
        return redirect(request.META.get('HTTP_REFERER') or 'core:super_admin_posts')
    except Exception:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Failed to toggle post.'}, status=500)
        messages.error(request, 'Failed to toggle post.')
        return redirect('core:super_admin_posts')


@login_required
@require_POST
def super_admin_delete_post(request, post_id):
    if not request.user.is_superuser:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Forbidden'}, status=403)
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('core:super_admin_posts')

    try:
        post = get_object_or_404(Post, id=post_id)
        post_type = post.post_type
        post.delete()
        msg = f'{"Event" if post_type == "event" else "Post"} deleted successfully.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': msg})
        messages.success(request, msg)
        return redirect(request.META.get('HTTP_REFERER') or 'core:super_admin_posts')
    except Exception:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Failed to delete post.'}, status=500)
        messages.error(request, 'Failed to delete post.')
        return redirect('core:super_admin_posts')


# Service Review Views

@login_required
def create_service_review(request, service_id):
    """Create a review for a service (only available to users with completed bookings)."""
    from .models import ServiceReview, BookableService, UserInteraction, Booking
    from django.http import JsonResponse
    
    service = get_object_or_404(BookableService, id=service_id, is_active=True)
    church = service.church
    
    # Check if user has any completed bookings for this service
    completed_bookings = Booking.objects.filter(
        user=request.user,
        service=service,
        status=Booking.STATUS_COMPLETED
    ).order_by('-updated_at')
    
    if not completed_bookings.exists():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'You can only review services after completing a booking.'
            })
        messages.error(request, 'You can only review services after completing a booking.')
        return redirect('core:church_detail', slug=church.slug)
    
    # Check if user has already reviewed this service
    if service.has_user_reviewed(request.user):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'You have already reviewed this service.'
            })
        messages.warning(request, 'You have already reviewed this service.')
        return redirect('core:church_detail', slug=church.slug)
    
    if request.method == 'POST':
        try:
            # Get form data
            rating = int(request.POST.get('rating', 0))
            title = request.POST.get('title', '').strip()
            comment = request.POST.get('comment', '').strip()
            
            # Optional ratings
            staff_rating = request.POST.get('staff_rating')
            facility_rating = request.POST.get('facility_rating')
            value_rating = request.POST.get('value_rating')
            
            is_anonymous = request.POST.get('is_anonymous') == 'on'
            
            # Basic validation
            if not rating or rating < 1 or rating > 5:
                error_msg = 'Please provide a rating between 1 and 5 stars.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
                return redirect('core:church_detail', slug=church.slug)
            
            if not title or len(title) < 5:
                error_msg = 'Please provide a review title with at least 5 characters.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
                return redirect('core:church_detail', slug=church.slug)
            
            if not comment or len(comment) < 10:
                error_msg = 'Please provide a review comment with at least 10 characters.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
                return redirect('core:church_detail', slug=church.slug)
            
            # Use the most recent completed booking for this review
            booking = completed_bookings.first()
            
            # Create review
            review_data = {
                'user': request.user,
                'service': service,
                'church': church,
                'booking': booking,
                'rating': rating,
                'title': title,
                'comment': comment,
                'is_anonymous': is_anonymous,
            }
            
            # Add optional ratings if provided
            if staff_rating:
                try:
                    review_data['staff_rating'] = int(staff_rating)
                except ValueError:
                    pass
            
            if facility_rating:
                try:
                    review_data['facility_rating'] = int(facility_rating)
                except ValueError:
                    pass
            
            if value_rating:
                try:
                    review_data['value_rating'] = int(value_rating)
                except ValueError:
                    pass
            
            review = ServiceReview.objects.create(**review_data)
            
            # Log activity
            UserInteraction.log_activity(
                user=request.user,
                activity_type=UserInteraction.ACTIVITY_SERVICE_REVIEW,
                content_object=review,
                metadata={
                    'service_name': service.name,
                    'church_name': church.name,
                    'rating': rating,
                    'title': title,
                },
                request=request
            )
            
            # Handle AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
                return JsonResponse({
                    'success': True,
                    'message': f'Thank you for your review of {service.name}!'
                })
            
            messages.success(request, f'Thank you for your review of {service.name}!')
            return redirect('core:service_reviews', service_id=service.id)
            
        except Exception as e:
            error_message = f'Error creating review: {str(e)}'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': error_message
                })
            messages.error(request, error_message)
            return redirect('core:church_detail', slug=church.slug)
    
    return redirect('core:church_detail', slug=church.slug)


@login_required
def service_reviews(request, service_id):
    """View all reviews for a specific service."""
    from .models import ServiceReview, BookableService
    
    service = get_object_or_404(BookableService, id=service_id, is_active=True)
    church = service.church
    
    # Get all reviews for this service
    reviews = service.reviews.filter(is_active=True).select_related(
        'user', 'booking'
    ).order_by('-created_at')
    
    # Check if current user has reviewed this service and can review
    user_review = None
    can_review = False
    completed_bookings = None
    if request.user.is_authenticated:
        try:
            user_review = reviews.filter(user=request.user).first()
            can_review = service.can_user_review(request.user)
            if can_review:
                completed_bookings = service.get_user_completed_bookings(request.user)
        except:
            pass
    
    # Rating statistics
    total_reviews = reviews.count()
    average_rating = service.average_rating
    rating_distribution = service.rating_distribution
    
    ctx = {
        'page_title': f'{service.name} Reviews',
        'service': service,
        'church': church,
        'reviews': reviews,
        'user_review': user_review,
        'can_review': can_review,
        'completed_bookings': completed_bookings,
        'total_reviews': total_reviews,
        'average_rating': average_rating,
        'rating_distribution': rating_distribution,
    }
    ctx.update(_app_context(request))
    return render(request, 'core/service_reviews.html', ctx)


@login_required
def toggle_review_helpful(request, review_id):
    """Toggle helpful vote for a review (AJAX)."""
    from .models import ServiceReview, ServiceReviewHelpful
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})
    
    try:
        review = get_object_or_404(ServiceReview, id=review_id, is_active=True)
        
        # Check if user has already voted
        helpful_vote, created = ServiceReviewHelpful.objects.get_or_create(
            user=request.user,
            review=review
        )
        
        if not created:
            # Remove vote
            helpful_vote.delete()
            is_helpful = False
            action = 'removed'
        else:
            # Add vote
            is_helpful = True
            action = 'added'
        
        # Update helpful votes count
        review.helpful_votes = review.helpful_votes_records.count()
        review.save(update_fields=['helpful_votes'])
        
        return JsonResponse({
            'success': True,
            'is_helpful': is_helpful,
            'helpful_count': review.helpful_votes,
            'action': action,
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def service_images_api(request, service_id):
    """API endpoint to fetch all images for a service"""
    try:
        service = get_object_or_404(BookableService, id=service_id)
        
        # Get all images for the service
        images = []
        
        # Add primary image if exists
        if service.get_primary_image():
            images.append({
                'url': service.get_primary_image().image.url,
                'is_primary': True
            })
        
        # Add all other service images
        for service_image in service.service_images.all().order_by('order', 'created_at'):
            # Skip if this is already the primary image
            if service.get_primary_image() and service_image == service.get_primary_image():
                continue
                
            images.append({
                'url': service_image.image.url,
                'is_primary': False
            })
        
        # Fallback to single service image if no ServiceImage objects exist
        if not images and service.image:
            images.append({
                'url': service.image.url,
                'is_primary': True
            })
        
        return JsonResponse({
            'success': True,
            'service_name': service.name,
            'images': images,
            'total_count': len(images)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


def api_get_church_availability(request, church_id):
    """API endpoint to fetch church availability (closed dates) for booking modal."""
    from django.http import JsonResponse
    
    try:
        # Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({
                'success': False,
                'error': 'Authentication required.'
            }, status=401)
            
        church = get_object_or_404(Church, id=church_id, is_verified=True)
        
        # Get all availability entries for the church
        # Only get future dates (today and onwards)
        from datetime import date
        today = date.today()
        availability_entries = church.availability.filter(date__gte=today).order_by('date')
        
        # Prepare availability data
        closed_dates = []
        special_hours = []
        
        for entry in availability_entries:
            date_str = entry.date.strftime('%Y-%m-%d')
            entry_data = {
                'date': date_str,
                'reason': entry.reason or '',
                'notes': entry.notes or ''
            }
            
            if entry.is_closed:
                closed_dates.append(entry_data)
            else:
                entry_data.update({
                    'start_time': entry.start_time.strftime('%H:%M') if entry.start_time else '',
                    'end_time': entry.end_time.strftime('%H:%M') if entry.end_time else ''
                })
                special_hours.append(entry_data)
        
        return JsonResponse({
            'success': True,
            'church_id': church.id,
            'church_name': church.name,
            'closed_dates': closed_dates,
            'special_hours': special_hours
        }, status=200)
        
    except Church.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Church not found or is not verified.'
        }, status=404)
        
    except Exception as e:
        logger.error(f'API Church availability fetch error: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while fetching church availability data.'
        }, status=500)


def api_get_pending_booking_dates(request, church_id):
    """API endpoint to fetch dates with pending bookings for the current user at a specific church."""
    from django.http import JsonResponse
    from datetime import date, timedelta
    
    try:
        # Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({
                'success': False,
                'error': 'Authentication required.',
                'pending_dates': []
            }, status=401)
            
        church = get_object_or_404(Church, id=church_id)
        
        # Get pending bookings for this user at this church
        # Only get bookings from today onwards
        today = date.today()
        future_date = today + timedelta(days=30)  # Look ahead 30 days
        
        pending_bookings = Booking.objects.filter(
            user=request.user,
            service__church=church,
            status__in=['pending', 'requested'],  # Handle both status values
            date__gte=today,
            date__lte=future_date
        ).values_list('date', flat=True).distinct()
        
        # Convert dates to string format
        pending_dates = [d.strftime('%Y-%m-%d') for d in pending_bookings]
        
        return JsonResponse({
            'success': True,
            'church_id': church.id,
            'pending_dates': pending_dates
        }, status=200)
        
    except Church.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Church not found.',
            'pending_dates': []
        }, status=404)
        
    except Exception as e:
        logger.error(f'API Pending booking dates fetch error: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while fetching pending booking dates.',
            'pending_dates': []
        }, status=500)


@login_required
@require_POST
def dashboard_create_post(request):
    """Create a new post from the dashboard for one of user's owned churches."""
    import json
    
    try:
        # Get form data
        church_id = request.POST.get('church_id')
        content = request.POST.get('content', '').strip()
        image = request.FILES.get('image')  # Legacy single image
        images = request.FILES.getlist('images')  # Multiple images
        post_type = request.POST.get('post_type', 'general')
        
        # Get event-specific fields
        event_title = request.POST.get('event_title', '').strip()
        event_start_date = request.POST.get('event_start_date')
        event_end_date = request.POST.get('event_end_date')
        event_location = request.POST.get('event_location', '').strip()
        max_participants = request.POST.get('max_participants')
        
        # Get donation fields
        enable_donation = request.POST.get('enable_donation') == 'on' or request.POST.get('enable_donation') == 'true'
        donation_goal = request.POST.get('donation_goal', '').strip()
        
        if not church_id:
            return JsonResponse({'success': False, 'message': 'Church selection is required.'}, status=400)
        
        if not content:
            return JsonResponse({'success': False, 'message': 'Post content cannot be empty.'}, status=400)
        
        # Validate content length
        if len(content) > 1000:
            return JsonResponse({'success': False, 'message': 'Post content is too long (max 1000 characters).'}, status=400)
        
        # Validate event post fields
        if post_type == 'event':
            if not event_title:
                return JsonResponse({'success': False, 'message': 'Event title is required for event posts.'}, status=400)
            if not event_start_date:
                return JsonResponse({'success': False, 'message': 'Event start date is required for event posts.'}, status=400)
            if not event_end_date:
                return JsonResponse({'success': False, 'message': 'Event end date is required for event posts.'}, status=400)
        
        # Get the church and verify ownership
        try:
            church = Church.objects.get(id=church_id, is_active=True)
        except Church.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Church not found.'}, status=404)
        
        # Check if user can manage content (Owner or Ministry Leader)
        can_manage, role = user_can_manage_church(request.user, church, ['content'])
        if not can_manage:
            return JsonResponse({'success': False, 'message': 'You do not have permission to create posts.'}, status=403)
        
        # Validate PayPal email is set if donations are enabled
        if enable_donation and not church.paypal_email:
            return JsonResponse({
                'success': False,
                'message': 'Please set up your PayPal email in Church Profile settings before enabling donations. This is required to receive donation payments directly.'
            }, status=400)
        
        # Create the post with all fields
        post_data = {
            'church': church,
            'content': content,
            'post_type': post_type,
        }
        
        # Add image if provided
        if image:
            post_data['image'] = image
        
        # Add event-specific fields if it's an event post
        if post_type == 'event':
            post_data['event_title'] = event_title
            post_data['event_start_date'] = event_start_date
            post_data['event_end_date'] = event_end_date
            if event_location:
                post_data['event_location'] = event_location
            if max_participants:
                try:
                    post_data['max_participants'] = int(max_participants)
                except (ValueError, TypeError):
                    pass
        
        # Add donation fields
        post_data['enable_donation'] = enable_donation
        if enable_donation and donation_goal:
            try:
                post_data['donation_goal'] = float(donation_goal)
            except (ValueError, TypeError):
                pass
        
        post = Post.objects.create(**post_data)
        
        # Handle multiple images
        if images:
            from .models import PostImage
            for index, image_file in enumerate(images):
                # Validate each image
                if image_file.size > 10 * 1024 * 1024:  # 10MB limit per image
                    continue  # Skip large images
                
                # Create PostImage record
                PostImage.objects.create(
                    post=post,
                    image=image_file,
                    order=index
                )
        
        # Log staff activity for post creation
        from .models import StaffActivityLog
        log_staff_activity(
            user=request.user,
            church=church,
            action=StaffActivityLog.ACTION_CREATE,
            category=StaffActivityLog.CATEGORY_POST,
            description=f"Created {post.get_post_type_display()} post" + (f": '{event_title}'" if post_type == 'event' else ""),
            target_id=post.id,
            target_type='post',
            request=request
        )
        
        # Return success response with post data
        return JsonResponse({
            'success': True,
            'message': 'Post created successfully!',
            'post': {
                'id': post.id,
                'content': post.content,
                'post_type': post.post_type,
                'church_name': church.name,
                'time_ago': post.time_ago,
                'created_at': post.created_at.isoformat(),
            }
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Dashboard post creation error: {str(e)}')
        return JsonResponse({'success': False, 'message': 'An error occurred while creating the post.'}, status=500)


@login_required
@require_POST
def report_post(request, post_id):
    """Report a post for inappropriate content."""
    from .models import PostReport
    
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user already reported this post
        if PostReport.objects.filter(user=request.user, post=post).exists():
            return JsonResponse({
                'success': False,
                'message': 'You have already reported this post.'
            }, status=400)
        
        # Get report data
        reason = request.POST.get('reason')
        description = request.POST.get('description', '').strip()
        
        if not reason:
            return JsonResponse({
                'success': False,
                'message': 'Please select a reason for reporting.'
            }, status=400)
        
        if not description:
            return JsonResponse({
                'success': False,
                'message': 'Please provide additional details.'
            }, status=400)
        
        # Create the report
        report = PostReport.objects.create(
            user=request.user,
            post=post,
            reason=reason,
            description=description
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Thank you for your report. We will review it shortly.'
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Post report error: {str(e)}')
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while submitting your report.'
        }, status=500)


@login_required
@require_POST
def delete_post(request, post_id):
    """Delete a post."""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user can manage content (Owner or Ministry Leader)
        can_manage, role = user_can_manage_church(request.user, post.church, ['content'])
        if not can_manage:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to delete this post.'
            }, status=403)
        
        # Store post info before deletion
        post_type = post.post_type
        
        # Delete the post
        post.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'{"Event" if post_type == "event" else "Post"} deleted successfully.'
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Post deletion error: {str(e)}')
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while deleting the post.'
        }, status=500)


@login_required
def get_post_data(request, post_id):
    """Get post data for editing."""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user can manage content (Owner or Ministry Leader)
        can_manage, role = user_can_manage_church(request.user, post.church, ['content'])
        if not can_manage:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to edit this post.'
            }, status=403)
        
        return JsonResponse({
            'success': True,
            'post': {
                'id': post.id,
                'content': post.content,
                'post_type': post.post_type,
                'event_title': post.event_title or '',
                'event_start_date': post.event_start_date.isoformat() if post.event_start_date else '',
                'event_end_date': post.event_end_date.isoformat() if post.event_end_date else '',
                'event_start_time': post.event_start_time.strftime('%H:%M') if post.event_start_time else '',
                'event_end_time': post.event_end_time.strftime('%H:%M') if post.event_end_time else '',
                'event_location': post.event_location or '',
                'has_image': bool(post.image),
                'image_url': post.image.url if post.image else None,
            }
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Get post data error: {str(e)}')
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while fetching post data.'
        }, status=500)


@login_required
@require_POST
def update_post(request, post_id):
    """Update an existing post."""
    try:
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user can manage content (Owner or Ministry Leader)
        can_manage, role = user_can_manage_church(request.user, post.church, ['content'])
        if not can_manage:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to edit this post.'
            }, status=403)
        
        # Get form data
        content = request.POST.get('content', '').strip()
        post_type = request.POST.get('post_type', 'general')
        
        if not content:
            return JsonResponse({
                'success': False,
                'message': 'Post content cannot be empty.'
            }, status=400)
        
        # Validate content length
        if len(content) > 1000:
            return JsonResponse({
                'success': False,
                'message': 'Post content is too long (max 1000 characters).'
            }, status=400)
        
        # Update basic fields
        post.content = content
        post.post_type = post_type
        
        # Handle event fields
        if post_type == 'event':
            event_title = request.POST.get('event_title', '').strip()
            event_start_date = request.POST.get('event_start_date')
            
            if not event_title:
                return JsonResponse({
                    'success': False,
                    'message': 'Event title is required for event posts.'
                }, status=400)
            
            if not event_start_date:
                return JsonResponse({
                    'success': False,
                    'message': 'Event start date is required for event posts.'
                }, status=400)
            
            post.event_title = event_title
            post.event_start_date = event_start_date
            post.event_end_date = request.POST.get('event_end_date') or None
            post.event_start_time = request.POST.get('event_start_time') or None
            post.event_end_time = request.POST.get('event_end_time') or None
            post.event_location = request.POST.get('event_location', '').strip() or None
        else:
            # Clear event fields if not an event post
            post.event_title = None
            post.event_start_date = None
            post.event_end_date = None
            post.event_start_time = None
            post.event_end_time = None
            post.event_location = None
        
        # Handle image update
        if 'image' in request.FILES:
            post.image = request.FILES['image']
        elif request.POST.get('remove_image') == 'true':
            post.image = None
        
        post.save()
        
        # Log staff activity for post update
        from .models import StaffActivityLog
        log_staff_activity(
            user=request.user,
            church=post.church,
            action=StaffActivityLog.ACTION_UPDATE,
            category=StaffActivityLog.CATEGORY_POST,
            description=f"Updated {post.get_post_type_display()} post" + (f": '{post.event_title}'" if post_type == 'event' and post.event_title else ""),
            target_id=post.id,
            target_type='post',
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Post updated successfully!'
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Post update error: {str(e)}')
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while updating the post.'
        }, status=500)


@login_required
def get_post_analytics(request, post_id):
    """Get detailed analytics for a specific post."""
    try:
        from django.db.models import Count, Q, F
        from django.utils import timezone
        from datetime import timedelta
        
        post = get_object_or_404(Post, id=post_id)
        
        # Check if user has permission to manage content
        has_permission, role = user_can_manage_church(request.user, post.church, required_permissions=['content'])
        
        if not request.user.is_superuser and not has_permission:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to view this post analytics.'
            }, status=403)
        
        # Get basic stats
        views_count = post.view_count
        likes_count = post.likes.count()
        comments_count = post.comments.filter(is_active=True).count()
        bookmarks_count = post.bookmarks.count()
        
        # Calculate engagement rate
        total_interactions = likes_count + comments_count + bookmarks_count
        engagement_rate = (total_interactions / views_count * 100) if views_count > 0 else 0
        
        # Get hourly views and interactions data (last 24 hours)
        now = timezone.now()
        hours_ago_24 = now - timedelta(hours=24)
        
        # Get views over time (hourly)
        views_over_time = []
        interactions_over_time = []
        
        for i in range(24):
            hour_start = hours_ago_24 + timedelta(hours=i)
            hour_end = hour_start + timedelta(hours=1)
            
            # Count views in this hour
            hour_views = PostView.objects.filter(
                post=post,
                viewed_at__gte=hour_start,
                viewed_at__lt=hour_end
            ).count()
            
            # Count interactions (likes + comments + bookmarks) in this hour
            hour_likes = PostLike.objects.filter(
                post=post,
                created_at__gte=hour_start,
                created_at__lt=hour_end
            ).count()
            
            hour_comments = PostComment.objects.filter(
                post=post,
                created_at__gte=hour_start,
                created_at__lt=hour_end,
                is_active=True
            ).count()
            
            hour_bookmarks = PostBookmark.objects.filter(
                post=post,
                created_at__gte=hour_start,
                created_at__lt=hour_end
            ).count()
            
            views_over_time.append({
                'hour': hour_start.strftime('%H:%M'),
                'count': hour_views
            })
            
            interactions_over_time.append({
                'hour': hour_start.strftime('%H:%M'),
                'count': hour_likes + hour_comments + hour_bookmarks
            })
        
        # Get engagement breakdown
        engagement_breakdown = {
            'likes': likes_count,
            'comments': comments_count,
            'bookmarks': bookmarks_count,
            'shares': 0  # Placeholder for future share tracking
        }
        
        # Get audience demographics (followers vs non-followers)
        followers_views = PostView.objects.filter(
            post=post,
            user__in=post.church.followers.values_list('user', flat=True)
        ).count()
        
        total_followers = post.church.follower_count
        non_followers_views = views_count - followers_views
        
        # Calculate active vs new followers
        active_followers_views = followers_views
        new_followers_views = 0  # Simplified for now
        
        audience_demographics = {
            'active_followers': {
                'count': active_followers_views,
                'percentage': round((active_followers_views / views_count * 100) if views_count > 0 else 0, 1)
            },
            'new_followers': {
                'count': new_followers_views,
                'percentage': round((new_followers_views / views_count * 100) if views_count > 0 else 0, 1)
            },
            'non_followers': {
                'count': non_followers_views,
                'percentage': round((non_followers_views / views_count * 100) if views_count > 0 else 0, 1)
            }
        }
        
        # Get top comments (most liked)
        top_comments = []
        comments = post.comments.filter(is_active=True).annotate(
            likes_count=Count('comment_likes')
        ).order_by('-likes_count', '-created_at')[:5]
        
        for comment in comments:
            # Generate initials safely
            if comment.user.first_name and comment.user.last_name and len(comment.user.first_name) > 0 and len(comment.user.last_name) > 0:
                initials = f"{comment.user.first_name[0]}{comment.user.last_name[0]}".upper()
            elif comment.user.username and len(comment.user.username) > 0:
                initials = comment.user.username[0].upper()
            else:
                initials = "U"
            
            # Get profile picture safely
            profile_picture = None
            try:
                if hasattr(comment.user, 'profile') and comment.user.profile and comment.user.profile.profile_image:
                    profile_picture = comment.user.profile.profile_image.url
            except Exception:
                profile_picture = None
            
            top_comments.append({
                'user': {
                    'name': comment.user.get_full_name() or comment.user.username,
                    'initials': initials,
                    'profile_picture': profile_picture
                },
                'content': comment.content,
                'likes': comment.likes_count,
                'time_ago': comment.time_ago
            })
        
        # Calculate performance insights
        # Get average stats for all posts from this church
        all_posts = Post.objects.filter(church=post.church, is_active=True).exclude(id=post.id)
        
        if all_posts.exists():
            avg_views = all_posts.aggregate(avg=Count('post_views'))['avg'] or 0
            avg_engagement = 0
            
            for p in all_posts:
                p_interactions = p.likes.count() + p.comments.filter(is_active=True).count() + p.bookmarks.count()
                p_views = p.view_count
                if p_views > 0:
                    avg_engagement += (p_interactions / p_views * 100)
            
            avg_engagement = avg_engagement / all_posts.count() if all_posts.count() > 0 else 0
            
            # Compare current post with average
            performance_vs_average = ((engagement_rate - avg_engagement) / avg_engagement * 100) if avg_engagement > 0 else 0
        else:
            performance_vs_average = 0
            avg_engagement = 0
        
        # Find peak engagement time
        peak_hour = max(interactions_over_time, key=lambda x: x['count'])
        
        performance_insights = {
            'above_average': {
                'percentage': round(performance_vs_average, 1),
                'message': f"This post is performing {abs(round(performance_vs_average))}% {'better' if performance_vs_average > 0 else 'worse'} than your average post"
            },
            'peak_engagement': {
                'time': peak_hour['hour'],
                'message': f"Most interactions occurred around {peak_hour['hour']}"
            }
        }
        
        # Donation analytics (if enabled)
        donation_analytics = None
        if post.enable_donation:
            from django.db.models import Sum, Avg
            from decimal import Decimal
            
            # Get completed donations
            completed_donations = post.donations.filter(payment_status='completed')
            
            # Calculate donation stats
            donation_stats = completed_donations.aggregate(
                total_raised=Sum('amount'),
                donor_count=Count('donor', distinct=True),
                avg_donation=Avg('amount')
            )
            
            total_raised = donation_stats['total_raised'] or Decimal('0.00')
            donor_count = donation_stats['donor_count'] or 0
            avg_donation = donation_stats['avg_donation'] or Decimal('0.00')
            
            # Calculate progress
            goal = post.donation_goal or Decimal('0.00')
            progress_percentage = 0
            if goal > 0:
                progress_percentage = min((float(total_raised) / float(goal)) * 100, 100)
            
            # Get donations over time (last 24 hours)
            donations_over_time = []
            for i in range(24):
                hour_start = hours_ago_24 + timedelta(hours=i)
                hour_end = hour_start + timedelta(hours=1)
                
                hour_donations = completed_donations.filter(
                    completed_at__gte=hour_start,
                    completed_at__lt=hour_end
                ).aggregate(
                    count=Count('id'),
                    amount=Sum('amount')
                )
                
                donations_over_time.append({
                    'hour': hour_start.strftime('%H:%M'),
                    'count': hour_donations['count'] or 0,
                    'amount': float(hour_donations['amount'] or 0)
                })
            
            # Get recent donations (last 10)
            recent_donations = []
            for donation in completed_donations.order_by('-completed_at')[:10]:
                donor_name = 'Anonymous'
                profile_picture = None
                initials = 'A'
                
                if donation.donor:
                    if donation.is_anonymous:
                        donor_name = 'Anonymous Donor'
                        initials = 'A'
                    else:
                        donor_name = donation.donor.get_full_name() or donation.donor.username
                        # Get profile image safely
                        try:
                            if hasattr(donation.donor, 'profile') and donation.donor.profile and donation.donor.profile.profile_image:
                                profile_picture = donation.donor.profile.profile_image.url
                        except Exception:
                            profile_picture = None
                        # Generate initials
                        if donation.donor.first_name and donation.donor.last_name and len(donation.donor.first_name) > 0 and len(donation.donor.last_name) > 0:
                            initials = f"{donation.donor.first_name[0]}{donation.donor.last_name[0]}".upper()
                        elif donation.donor.username and len(donation.donor.username) > 0:
                            initials = donation.donor.username[0].upper()
                        else:
                            initials = "U"
                elif donation.donor_name:
                    donor_name = donation.donor_name if not donation.is_anonymous else 'Anonymous Donor'
                    # Generate initials from donor_name
                    if not donation.is_anonymous and donation.donor_name:
                        name_parts = donation.donor_name.split()
                        if len(name_parts) >= 2:
                            initials = f"{name_parts[0][0]}{name_parts[-1][0]}".upper()
                        elif len(name_parts) == 1 and len(name_parts[0]) > 0:
                            initials = name_parts[0][0].upper()
                
                # Get time display
                if hasattr(donation, 'time_ago') and donation.time_ago:
                    time_display = donation.time_ago
                elif donation.completed_at:
                    time_display = donation.completed_at.strftime('%Y-%m-%d %H:%M')
                else:
                    time_display = 'Unknown'
                
                recent_donations.append({
                    'donor_name': donor_name,
                    'amount': float(donation.amount),
                    'time_ago': time_display,
                    'is_anonymous': donation.is_anonymous,
                    'profile_picture': profile_picture,
                    'initials': initials
                })
            
            # Donor demographics (followers vs non-followers)
            follower_donations = completed_donations.filter(
                donor__in=post.church.followers.values_list('user', flat=True)
            ).count()
            non_follower_donations = donor_count - follower_donations
            
            donation_analytics = {
                'enabled': True,
                'total_raised': float(total_raised),
                'goal': float(goal) if goal else None,
                'progress_percentage': round(progress_percentage, 1),
                'donor_count': donor_count,
                'avg_donation': float(avg_donation),
                'donations_over_time': donations_over_time,
                'recent_donations': recent_donations,
                'donor_demographics': {
                    'followers': {
                        'count': follower_donations,
                        'percentage': round((follower_donations / donor_count * 100) if donor_count > 0 else 0, 1)
                    },
                    'non_followers': {
                        'count': non_follower_donations,
                        'percentage': round((non_follower_donations / donor_count * 100) if donor_count > 0 else 0, 1)
                    }
                }
            }
        
        # Get multiple images
        post_images = []
        for img in post.images.all().order_by('order', 'created_at'):
            post_images.append({
                'id': img.id,
                'image_url': img.image.url,
                'order': img.order
            })
        
        return JsonResponse({
            'success': True,
            'analytics': {
                'post_id': post.id,
                'post_type': 'Fundraiser' if post.enable_donation else post.get_post_type_display(),
                'post_type_value': 'fundraiser' if post.enable_donation else post.post_type,
                'created_at': post.created_at.strftime('%Y-%m-%d %H:%M'),
                'updated_at': post.updated_at.strftime('%Y-%m-%d %H:%M'),
                'content': post.content[:200],  # First 200 chars
                'image_url': post.image.url if post.image else None,  # Legacy single image
                'has_image': bool(post.image),  # Legacy single image check
                'images': post_images,  # Multiple images
                'status': 'published' if post.is_active else 'inactive',
                'stats': {
                    'views': views_count,
                    'likes': likes_count,
                    'comments': comments_count,
                    'bookmarks': bookmarks_count,
                    'engagement_rate': round(engagement_rate, 2)
                },
                'views_over_time': views_over_time,
                'interactions_over_time': interactions_over_time,
                'engagement_breakdown': engagement_breakdown,
                'audience_demographics': audience_demographics,
                'top_comments': top_comments,
                'performance_insights': performance_insights,
                'donation_analytics': donation_analytics
            }
        })
        
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f'Get post analytics error: {str(e)}')
        logger.error(f'Traceback: {traceback.format_exc()}')
        
        # Return detailed error in development
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while fetching post analytics.',
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)


@login_required
def user_activities(request):
    """
    Display all user activities (post interactions, church follows, bookings, etc.)
    """
    from .activity_tracker import ActivityTracker
    from accounts.models import UserActivity
    
    # Get filter parameter
    activity_filter = request.GET.get('filter', 'all')
    
    # Get post interactions
    if activity_filter == 'posts':
        activities = ActivityTracker.get_post_interactions(request.user, limit=50)
    elif activity_filter == 'churches':
        activities = ActivityTracker.get_church_interactions(request.user, limit=50)
    elif activity_filter == 'auth':
        # Get authentication activities instead
        activities = UserActivity.objects.filter(user=request.user).order_by('-created_at')[:50]
    else:
        # Get all interactions
        activities = ActivityTracker.get_user_activities(request.user, limit=50)
    
    # Get activity statistics
    stats = ActivityTracker.get_activity_stats(request.user)
    
    # Get recent authentication activities for sidebar
    recent_auth_activities = UserActivity.objects.filter(
        user=request.user
    ).order_by('-created_at')[:3]
    
    context = {
        'activities': activities,
        'activity_filter': activity_filter,
        'stats': stats,
        'recent_auth_activities': recent_auth_activities,
        'active': 'activities',
        'is_admin_mode': bool(request.session.get('super_admin_mode', False)) if getattr(request.user, 'is_superuser', False) else False,
    }
    
    return render(request, 'user_activities.html', context)


@login_required
def global_search(request):
    """
    Global search API for topbar search - returns churches and posts
    """
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({
            'success': True,
            'churches': [],
            'posts': []
        })
    
    # Split query into words for flexible matching
    words = query.split()
    
    # Search churches (parishes) - match any word
    church_query = Q()
    for word in words:
        church_query |= Q(name__icontains=word) | Q(description__icontains=word) | Q(address__icontains=word)
    
    churches = Church.objects.filter(church_query).filter(is_active=True).distinct()[:5]
    
    # Search posts - match any word
    post_query = Q()
    for word in words:
        post_query |= Q(content__icontains=word) | Q(church__name__icontains=word)
    
    posts = Post.objects.filter(post_query).filter(is_active=True).select_related('church').order_by('-created_at').distinct()[:5]
    
    # Format church results
    church_results = []
    for church in churches:
        church_results.append({
            'id': church.id,
            'name': church.name,
            'slug': church.slug,
            'address': church.address or '',
            'logo_url': church.logo.url if church.logo else None,
            'url': reverse('core:church_detail', kwargs={'slug': church.slug})
        })
    
    # Format post results
    post_results = []
    for post in posts:
        # Truncate content for preview
        content_preview = post.content[:100] + '...' if len(post.content) > 100 else post.content
        
        post_results.append({
            'id': post.id,
            'content': content_preview,
            'church_name': post.church.name,
            'church_slug': post.church.slug,
            'created_at': post.created_at.isoformat(),
            'url': f"{reverse('core:church_detail', kwargs={'slug': post.church.slug})}#post-card-{post.id}"
        })
    
    return JsonResponse({
        'success': True,
        'churches': church_results,
        'posts': post_results
    })